#****************************************************************************************************************************************************************************************************************************************************************
# Nombre: Buscador.Py                                                                                                                                                                                                                                           *       
# Descripcion: Aplicación de busqueda de equipos en listado de equipos                                                                                                                                                                                          *
#               La idea de este soft es facilitar la busqueda de activos en la planilla de equipos electronicos que disponemos.                                                                                                                                 *
#               La funcion es simple, por medio de los modulos openpyxl navegamos la planilla de excel existente con los datos que tenemos de los equipos. Buscamos por medio del nombre del equipo que se compone por: Area-Tipo de equipo-N° designado        *
#               por medio del modulo Tkinter le hicimos la interfaz grafica para que el usuario pueda utilizar la misma sin tener que ir ingresando lineas de comando
#****************************************************************************************************************************************************************************************************************************************************************
import cmd
import datetime
import os as os
import sys as sys
import time
#****************************************************************************************************************************************************************************************************************************************************************
import tkinter as tk
from os import replace
from os import system as system
from select import select
from sys import *
from tkinter import *
from tkinter import (Entry, Grid, Image, StringVar, Text, Variable,commondialog, font, messagebox, scrolledtext,simpledialog, tix, ttk)
from tkinter.constants import *
from tkinter.tix import (STATUS, ButtonBox, ComboBox, LabelEntry, LabelFrame,Meter, PhotoImage)
from tkinter.ttk import (Entry, Progressbar, Sizegrip, Spinbox, Style,Treeview, setup_master)
from turtle import color, delay, title, width
#****************************************************************************************************************************************************************************************************************************************************************
import openpyxl
from openpyxl import Workbook, load_workbook
#****************************************************************************************************************************************************************************************************************************************************************
from Variables import *
#****************************************************************************************************************************************************************************************************************************************************************
# Defina Interfaz Grafica                                                                                                                                                                                                                                       *
#****************************************************************************************************************************************************************************************************************************************************************
#****************************************************************************************************************************************************************************************************************************************************************
mybus = Tk()                                                                                                                        # MyApp es el nombre de la planilla a visualizar                                                                            *                                                                                                             
mybus.title("Buscador")                                                                                                             # Defino el titulo del programa                                                                                             *
H=700                                                                                                                               # Defino altura de la ventana                                                                                               *
W=1000                                                                                                                              # Defino ancho de la ventana                                                                                                *
mybus.minsize(W,H)                                                                                                                  # Defino el Tamaño minimo e inicial de plantilla                                                                            *
mybus.frame()
mybus.resizable(False,False)
mybus.config(background=Fondo)                                                                                                      # Defino color de Fondo de pantalla "Myapp"                                                                                 *
wb=openpyxl.load_workbook(str(lineas[0])[:-1])                                                                                      # Genero el vinculo y abro el Excel que esta definido dentro del String dentro del Path                                     *       
#****************************************************************************************************************************************************************************************************************************************************************
# Lineas de comando para poder centrar la ventana en la mitad de la pantalla (En teoria)
#****************************************************************************************************************************************************************************************************************************************************************
Screen_width    = mybus.winfo_screenwidth()                                                                                         # En estas lineas de comando, su supone que centra la ventana de la aplicación                                              *
Screen_height   = mybus.winfo_screenheight()                                                                                        # aun esta en modo prueba. Posiblemente lo elimine al finalizar el proyecto                                                 *
X_cordinate     = int((Screen_width/2) - (W/2))
Y_cordinate     = int((Screen_height/2) - (H/2)-25)
mybus.geometry("{}x{}+{}+{}".format(W,H, X_cordinate, Y_cordinate))
#****************************************************************************************************************************************************************************************************************************************************************
Myimg=      PhotoImage(file=(user   + (str(lineas[3]))[:-1]))                                                                       # Variable para imagen del Boton de Busqueda. Se define la ruta en el programa.                                             *
Mylogo=     PhotoImage(file=(user   + (str(lineas[4]))[:-1]))                                                                       # Variable para imagen del Logo del Icono a usar                                                                            *
lbl_lable=  Label(mybus,image=Mylogo,border=0).place(x=-10,y=-10)                                                                   # Se define en Lable para poder utilizar "mylogo" como fondo de ventana.                                                    *
# Definimos el cuadro de entrada de Texto para buscar Celda                                                                                                                                                                                                     *
#****************************************************************************************************************************************************************************************************************************************************************
Entrada=    ttk.Entry(mybus,font=("Arial",12),width=20)                                                                             # Defino el renglon de entrada de datos para comparar                                                                       *
Entrada.place(x=10,y=35)                                                                                                            # Se define posición del renglon de "Entrada"                                                                               *
Entrada.insert(0,"Ingrese Valor de Celda")                                                                                          # Se coloca un texto inicial para que se muestre al iniciar el programa                                                     *
Entrada.get().upper()                                                                                                               # Configuro para que todo el dato ingresado se configure para poner en Mayusculas                                           *
Entrada.focus()                                                                                                                     # Se configura para tener el Foco al iniciar el programa                                                                    *
#****************************************************************************************************************************************************************************************************************************************************************
# Diceñamos Arbol de seleccion de WorkSheet
#****************************************************************************************************************************************************************************************************************************************************************
                                                                                                                                    # NOTA: En caso de requerir mas tablas para mostrar solo agregar como variable el nombre de la Tabla para poder procesar    *
Tree=               ttk.Treeview(mybus,height=4,show='tree')                                                                        # Creamos el Listado del Arbol para poder seleccionar el Servidor a Realizar conexion remota                                *   
Tree.place          (x=780,y=320)
Tree.tag_configure  ('arbol',background=Ftree,font=("Snap ITC",11),foreground='lightblue')                                          # Codificamos el estilo del tag para poder decorar mejor el mismo.                                                          *
Tree.tag_configure  ('arbol2',background=Ftree,font=("Snap ITC",11),foreground='red')                                               # Codificamos otro estilo para ver si podemos modificar el estilo cuando se hace foco sobre el Tag                          *
estilo='arbol'

t1_image=       PhotoImage(file=(user+ (str(lineas[7]))[:-1]))                                                                      # Definimos el icono a mostrar como "CCTV"                                                                                  *
t2_image=       PhotoImage(file=(user+ (str(lineas[8]))[:-1]))                                                                      # Icono a mostrar como "CA"                                                                                                 *
t3_image=       PhotoImage(file=(user+ (str(lineas[9]))[:-1]))                                                                      # Icono a mostrar como "Lectoras"                                                                                           *
t4_image=       PhotoImage(file=(user+ (str(lineas[10]))[:-1]))                                                                     # Icono a mostrar como "Avigilon"                                                                                           *
#****************************************************************************************************************************************************************************************************************************************************************
def Open_Pyxl():                                                                                                                    # Se optimiza la seleccion de la WorkSheet en este sub codigo. Genero todo el proceso de configuración del mismo para poder *
                                                                                                                                    #optimizar el codigo y que no tenga que ejecutar varias veces el mismo codigo en distintas partes del mismo                 *
    global ws,wbs,reg_ws,Mylogo                                                                                                            
        
    wbs = Tree.item(Tree.focus(),option='text')                                                                                     # Defino wbs como el ITem seleccionado del Tree que armamos previamente.                                                    *
    ws=wb[wbs]                                                                                                                      # Defino ws para poder utilizar como Worksheet del Libro / planilla
    if str(ws)!=str(reg_ws):                                                                                                        # Verifico que el valor de ws no sea igual al del control reg_ws, en este caso le paso el valor al control y lo almaceno    *
        reg_ws=str(ws)                                                                                                              #para poder volver a comparar en la siguiente ves que requiera.                                                             *
    return (ws,wbs)

#****************************************************************************************************************************************************************************************************************************************************************
# Definimos StringVar para Respuestas                                                                                                                                                                                                                           *
#   Nombre      |   Tipo de Variable|                                                                                                                                                                                                                           *   
#****************************************************************************************************************************************************************************************************************************************************************
Res_dispositivo=    StringVar()                                                                                                     # Variable para respuesta de Etiqueta de "Dispositivo"                                                                      *
Res_ubi=            StringVar()                                                                                                     # Variable para respuesta de Etiqueta de "Ubicacion"                                                                        *
Res_equipo=         StringVar()                                                                                                     # Variable para respuesta de Etiqueta de "Equipo"                                                                           *
Res_nombre=         StringVar()                                                                                                     # Variable para respuesta de Etiqueta de "Nombre"                                                                           *
Res_marca=          StringVar()                                                                                                     # Variable para respuesta de Etiqueta de "Marca"                                                                            *
Res_ip=             StringVar()                                                                                                     # Variable para respuesta de Etiqueta de "Ip"                                                                               *
Res_serial=         StringVar()                                                                                                     # Variable para respuesta de Etiqueta de "serial"                                                                           *
Res_usuario=        StringVar()                                                                                                     # Variable para respuesta de Etiqueta de "Usuario"                                                                          *
Res_password=       StringVar()                                                                                                     # Variable para respuesta de Etiqueta de "Contraseña"                                                                       *
Res_server=         StringVar()                                                                                                     # Variable para respuesta de Etiqueta de "Servidores"                                                                       *
#****************************************************************************************************************************************************************************************************************************************************************
def item_seleccionado (event):                                                                                                      # Evento creado para poder probar el cambio de estilo el Item del Tree (Modo Prueba)                                        *
    estilo='arbol2'

Tree.tag_bind("mytag", "<<TreeviewSelect>>", item_seleccionado)                                                                     # Se define el nombre del evento, Tipo de disparador y nombre del evento                                                    *

Tree.insert     ('', tk.END, text=S_t1, iid=0, open=False,image=t1_image,tags=(estilo,"mytag"))                                     # Defino el 1er tag
Tree.insert     ('', tk.END, text=S_t2, iid=1, open=False,image=t2_image,tags='arbol')                                              # Defino el 2do tag
Tree.insert     ('', tk.END, text=S_t3, iid=2, open=False,image=t3_image,tags='arbol')                                              # Defino el 3er tag
Tree.insert     ('', tk.END, text=S_t4, iid=3, open=False,image=t4_image,tags='arbol')                                              # Defino el 4to tag
#****************************************************************************************************************************************************************************************************************************************************************
#****************************************************************************************************************************************************************************************************************************************************************
# Definimos Los Labels de Etiquetas y de respuestas                                                                                                                                                                                                             *                  *
# Nombre            |   Definicion de Label                                                                                                                                                                                                                     *
#****************************************************************************************************************************************************************************************************************************************************************
lbl_Dispositivo=        tk.Label(mybus,text="Dispositivo"                   ,bg=Fondo,font=Fuente).place(x=10,y=10)                 # Label que indica el "Dispositivo"                                                                                         *
lbl_Ubicacion=          tk.Label(mybus,text="Ubicacion del Dispositivo =>"  ,bg=Fondo,font=Fuente).place(x=10,y=Renglon*2)          # Label que indica el "Ubicacion"                                                                                           *
lbl_equipo=             tk.Label(mybus,text="Tipo de Equipo =>"             ,bg=Fondo,font=Fuente).place(x=10,y=Renglon*3)          # Label que indica el "Tipo de Equipo"                                                                                      *
lbl_nombre=             tk.Label(mybus,text="Nombre del equipo =>"          ,bg=Fondo,font=Fuente).place(x=10,y=Renglon*4)          # Label que indica el "Nombre del Equipo"                                                                                   *
lbl_marca=              tk.Label(mybus,text="Marca de Equipo =>"            ,bg=Fondo,font=Fuente).place(x=10,y=Renglon*5)          # Label que indica el "Marca del Equipo"                                                                                    *
lbl_Serial=             tk.Label(mybus,text="Serial del Dispositivo =>"     ,bg=Fondo,font=Fuente).place(x=10,y=Renglon*6)          # Label que indica el "N° de Serie del Equipo"                                                                              *
lbl_ip=                 tk.Label(mybus,text="IP de Equipo =>"               ,bg=Fondo,font=Fuente).place(x=10,y=Renglon*7)          # Label que indica el "IP del Equipo"                                                                                       *
lbl_usuario=            tk.Label(mybus,text="Usuario del equipo =>"         ,bg=Fondo,font=Fuente).place(x=10,y=Renglon*8)          # Label que indica el "Usuario de Configuración"                                                                            *
lbl_Password=           tk.Label(mybus,text="Password de Equipo =>"         ,bg=Fondo,font=Fuente).place(x=10,y=Renglon*9)          # Label que indica el "Password de Configuración"                                                                           *

lbl_res_ubicacion=      tk.Label(mybus,textvariable=Res_ubi                 ,font=Fuente,width=50).place(x=300,y=Renglon*2)         # Label de respuesta que indica "Ubicacion"                                                                                 * 
lbl_res_equipo=         tk.Label(mybus,textvariable=Res_equipo              ,font=Fuente,width=50).place(x=300,y=Renglon*3)         # Label de respuesta que indica "Equipo"                                                                                    *
lbl_res_nombre=         tk.Label(mybus,textvariable=Res_nombre              ,font=Fuente,width=50).place(x=300,y=Renglon*4)         # Label de respuesta que indica "Nombre"                                                                                    *
lbl_res_marca=          tk.Label(mybus,textvariable=Res_marca               ,font=Fuente,width=50).place(x=300,y=Renglon*5)         # Label de respuesta que indica "Marca"                                                                                     *
lbl_res_Serial=         tk.Label(mybus,textvariable=Res_serial              ,font=Fuente,width=50).place(x=300,y=Renglon*6)         # Label de respuesta que indica "Serial"                                                                                    *
lbl_res_Ip=             tk.Label(mybus,textvariable=Res_ip                  ,font=Fuente,width=50).place(x=300,y=Renglon*7)         # Label de respuesta que indica "IP"                                                                                        *
lbl_res_Usuario=        tk.Label(mybus,textvariable=Res_usuario             ,font=Fuente,width=50).place(x=300,y=Renglon*8)         # Label de respuesta que indica "Usuario"                                                                                   *
lbl_res_Password=       tk.Label(mybus,textvariable=Res_password            ,font=Fuente,width=50).place(x=300,y=Renglon*9)         # Label de respuesta que indica "Password"                                                                                  *
lbl_res_Server=         tk.Label(mybus,textvariable=Res_server              ,font=Fuente,width=50).place(x=300,y=Renglon*10)        # Label de respuesta que indica "Servidor"                                                                                  *
#****************************************************************************************************************************************************************************************************************************************************************                                                                                                                
#****************************************************************************************************************************************************************************************************************************************************************
# Funciones del Programa a utilizar                                                                                                                                                                                                                             *
#****************************************************************************************************************************************************************************************************************************************************************   
def Ayuda():
    import Ayuda
#****************************************************************************************************************************************************************************************************************************************************************                        
def Siguiente():                                                                                                                    # Mostramos el valor siguiente al que tenemos en el buscador                                                                *
    
    global num,dispositivo                                                                                                          # Tomo las Variables Globales y le permito el uso dentro del subprograma                                                    *
    num+=1                                                                                                                          # Aumento la variable "num"                                                                                                 *
    Open_Pyxl()
    dispositivo=ws.cell(row=num,column=4)                                                                                           # Defino la variable para poder compara con la variable "Entrada" teniendo en cuenta los parametros "num" como renglones    *
                                                                                                                                    #y como columna una sola definitiva para poder buscar solamente por nombre de equipo                                        *   

    if dispositivo.value=="":                                                                                                       # Verifico que la variable no este vacia                                                                                    *
        messagebox.showerror(message="Final de listado.\n No hay datos similares")                                                  # En caso de que se encuentra vacia, mostramos mensaje indicando que llego al final del listado de equipos                  *
    else:
        Imprimir()                                                                                                                  # En caso de que tengamos información dentro de ese renglon, pasamos los datos al sub Imprimi()                             *
#****************************************************************************************************************************************************************************************************************************************************************    
def Anterior():                                                                                                                     # Mostramos el valor anteroor al que tenemos en el buscador                                                                 *
    
    global num, dispositivo                                                                                                         # Tomo las Variables Globales y le permito el uso dentro del subprograma                                                    *
    Open_Pyxl()
    if num>0:                                                                                                                       # En caso de que el Num sea mayor a 0, reducimos el valor de la variable                                                    * 
        num-=1                                                                                                                      # Se reduce el valor de la variable                                                                                         *
        Imprimir()                                                                                                                  # Ejecutamos el Sub Imprimir                                                                                                *
    else:
        messagebox.showerror(title= "Error al Buscar Valor",message="No se encuentra valor anterior al existente")                  # Caso que num sea =/< a 0, mostramos un mensaje de error para indicar que ya no se puede reducir mas esa variable          *  
#****************************************************************************************************************************************************************************************************************************************************************
def Buscar():                                                                                                                       # Buscamos el valor ingresado dentro del listado de equipos que tenemos en el listado                                       *
    
    global num,limite                                                                                                               # Importo variables                                                                                                         *
    global dispositivo
    num=1                                                                                                                           # Configuro la variable en valor incial, para poder buscar desde el inicio del listado                                      *
    
    Open_Pyxl()
    current_item = Tree.focus()                                                                                                     # Configuro la variable con el Focus del Tree, para poder utilizarlo como indicador de Ws.                                  *
    wbs = Tree.item(current_item,option='text')                                                                                     # Selecciono el item "Current_item" y lo configuro como "Text" y lo paso a la variable para utilizar                        *
    
    if wbs =="":
        messagebox.showerror(title="Error de Busqueda",message="Falta seleccionar dentro del arbol la subcategoria")                # Comparo el Valor obtenido de Wbs y en caso de estar vacio, Muestro este mensjae de error                              *
    else:
        wtree=wbs                                                                                                                   # Caso contrario, paso "wbs" a "wtree" que es un str para poder utilizarla en el programa.                              * 
        ws=wb[wbs]                                                                                                                  # Buscamos la WorkSheet correspondiente a lo selecionado                                                                    *

    dato=Entrada.get().upper()                                                                                                      # Tomo lo escrito en el Entry y lo paso a Upper para tener un control del ingreso                                           *
    dispositivo=ws.cell(row=num,column=4)                                                                                           # Tomo lo escrito dentro del Cell del listado segun el Num que es = al renglon del listado                                  *
    
    while dato!= dispositivo.value:                                                                                                 # Mientras que el valor de la Celda (Dispositivo) sea distinto a la entrada (Dato) aumenta en 1 el valor de Num (Renglon)   *                                  
        dispositivo=ws.cell(row=num,column=4)                                                                                       # vuelvo a marcar la franja en donde tengo que buscar en las tablas.                                                        *
        if num==limite:                                                                                                             # Comparamos los parametros Num con Limite, ya que sino la busqueda pasaria por toda la planilla inecesariamente            *
            num=1                                                                                                                   # Cuando llegamos al "Limite", reiniciamos la Variable Num para volver a buscar                                             *
            messagebox.showerror(title="Error 421",message="Por favor, verifique el dato a buscar")                                 # Mostramos un Mensaje de Error informando lo sucedido                                                                      *
                                                                                                                                    #,image="C:\user\desktop\codigo\nuclear.ico")
            return
        else:                                                                                                                       # Cuando la variable sea menor o distinta al Limite, aumentamos en 1 para navegar la planilla                               *
            num=num+1
        
    else: 
        num-=1                                                                                                                      # Cuando el Dato y el Dispositivo son iguales, volvemos un renglon para atras y mostramos ese renglon                       *
        Imprimir()                                                                                                                  # Ejecutamos el SubPrograma Imprimir()                                                                                      *

    
#****************************************************************************************************************************************************************************************************************************************************************        
def Imprimir():                                                                                                                     # Se imprimen los datos de la planilla encontrados segun lo Buscar(),Siguiente( y Anterior())                               *
    
    global dato,num,dispositivo                                                                                                     # Importamos las variables a utilizar / procesar                                                                            *
    
    print("valor de WS "+str(ws))
    
    dato=Entrada.get().upper()                                                                                                      # Agarro el texto ingresado y lo convierto en Mayusculas, solo para poder tener todo standarizado                           *
    dispositivo=ws.cell(row=num,column=4)                                                                                           # Tomo el valor de la celda para comparar                                                                                   *
    
    
    Mensaje=(str(Hini)+" => "+user[9:30] + " esta buscado el dato " +dato+" que esta es el registro numero => "+str(num)+"worksheet => "+wbs+"\n")                                                                                                              
                                                                                                                                    # Escribimos Horario, usuario, dato, renglon y tabla buscado para poder tener registro de lo realizado.                     *
    #import WLog                                                                                                                    # Configuramos el Mensjae para poder poner en el Log y tener registro de lo realizado                                       *
    
    #dispositivo=        ws.cell(row=num,column=4)                                                                                  # Cargamos los valores de la planilla en as variables segun su posicionamiento. Dispositivo                                 *
    #nombre=             ws.cell(row=num,column=4)                                                                                  # Nombre del dispositivo en el listado                                                                                      *
    #equipo=             ws.cell(row=num,column=2)                                                                                  # Tipo de Equipo                                                                                                            *  
    #ubicacion=          ws.cell(row=num,column=3)                                                                                  # Ubicación del Equipo                                                                                                      *
    #marca=              ws.cell(row=num,column=6)                                                                                  # Marca del Equipo                                                                                                          *
    #modelo=             ws.cell(row=num,column=7)                                                                                  # Modelo del Equipo                                                                                                         *
    #serial=             ws.cell(row=num,column=8)                                                                                  # N° de serie                                                                                                               *
    #ip=                 ws.cell(row=num,column=11)                                                                                 # N° de IP                                                                                                                  *
    #usuario=            ws.cell(row=num,column=12)                                                                                 # Usuario para ingresar a la configuración                                                                                  *
    #password=           ws.cell(row=num,column=13)                                                                                 # Password de ingreso a equipo                                                                                              *
    #server=             ws.cell(row=num,column=14)                                                                                 # Servidor a la cual esta conectado                                                                                         *

                                                                                                                                    # Seteamos los valores de las respuestas, en este caso ruteamos directamente la ubicacion, en lugar de cargalo en una var   *
    Res_nombre.set      (ws.cell(row=num,column=4).value)                                                                           # y luego setear el StringVar correspondiente. En esta caso es el Nombre                                                    *
    Res_equipo.set      (ws.cell(row=num,column=2).value)                                                                           # Tipo de Equipo                                                                                                            *
    Res_ubi.set         (str(ws.cell(row=num,column=3).value))                                                                      # Ubicación del Equipo                                                                                                      *
    Res_marca.set       (str(ws.cell(row=num,column=6).value) +"   =>  "+ str(ws.cell(row=num,column=7).value))                     # Marca del Equipo y modelo del equipo concatenado                                                                          *
    Res_ip.set          (ws.cell(row=num,column=11).value)                                                                          # N° de IP                                                                                                                  *
    Res_serial.set      (ws.cell(row=num,column=8).value)                                                                           # N° de serie                                                                                                               *
    Res_usuario.set     (ws.cell(row=num,column=12).value)                                                                          # Usuario para ingresar a la configuración                                                                                  *
    Res_password.set    (ws.cell(row=num,column=13).value)                                                                          # Password de ingreso a equipo                                                                                              *
    Res_server.set      (ws.cell(row=num,column=14).value)                                                                          # Servidor a la cual esta conectado                                                                                         *
#****************************************************************************************************************************************************************************************************************************************************************    
def Modifi():
    
    global wtree

    qst=messagebox.askokcancel(title="Modificar Parametros",message="Esta seguro que decea Modificar los Parametros???")
    if qst==True:
        current_item = Tree.focus()    
        wbs = Tree.item(current_item,option='text') 
        wtree=str(wbs)
        #import Destroy
        import Modificar
        
    else:
        Mensaje="Se cancelo modificación solicitada"
        messagebox.showerror(title="Modificación",message=Mensaje)
#****************************************************************************************************************************************************************************************************************************************************************        
def Conectar():

    wb=openpyxl.load_workbook(str(lineas[0])[:-1])
    ws=wb["CCTV"]
    server=ws.cell(row=num,column=14) 
    servidor='cmd /k "mstsc -v ' + server.value + ':4489'
    
    #import Destroy
    Salir()
    os.system(servidor)
    
    return 
#****************************************************************************************************************************************************************************************************************************************************************    
def Salir():
    
    Mensaje=(str(Hini)+" => Se cierra aplicación "+user[9:]+"\n")
    import WLog
    mybus.destroy()
    print("Se cierra Mybus")
    return
#****************************************************************************************************************************************************************************************************************************************************************
# Definimos los botoes a ver en la ventana inicial                                                                                                                                                                                                              *
#****************************************************************************************************************************************************************************************************************************************************************
boton=      tk.Button(mybus,text="Buscar",      activebackground="#ABCDEF",background="#838B8B",command=Buscar,width=180,image=Myimg).place(x=780,y=60)                               # Creo Boton "planilla" para procesar las plantillas Requeridas para informe.   *
salir=      tk.Button(mybus,text="Salir",       activebackground="#BABABA",command=Salir,justify='center',width=23).place(x=790,y=650)                                                 # Creo un Boton para cerrar la aplicación                                       *
bsiguiente= tk.Button(mybus,text="Siguiente",   activebackground="#ABABAB",background="#838383",command=Siguiente,width=11,state='active').place(x=780,y=150)
banterios=  tk.Button(mybus,text="Previo",      activebackground="#ABABAB",background="#838383",command=Anterior,width=11,state='active').place(x=880,y=150)
bmodificar= tk.Button(mybus,text="Modificar",   activebackground="#ABABAB",background="#838383",command=Modifi,width=25,state='active').place(x=780,y=180)
bconectar=  tk.Button(mybus,text="Conectar",    activebackground="#ABABAB",background="#838383",command=Conectar,width=25,state='active').place(x=780,y=500)
bhelp=      tk.Button(mybus,text="Ayuda",       background="#838383",command=Ayuda,width=5,).place(x=15,y=650)                                                                         # Creo el Boton de "Ayuda" para mostrar el Txt correspondiente                  *
#****************************************************************************************************************************************************************************************************************************************************************                     
#****************************************************************************************************************************************************************************************************************************************************************
log.close()
mybus.mainloop()
#****************************************************************************************************************************************************************************************************************************************************************


