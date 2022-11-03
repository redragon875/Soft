#****************************************************************************************************************************************************************************************************************************************************************
# Nombre: Buscador.Py                                                                                                                                                                                                                                           *       
# Descripcion: Aplicación de busqueda de equipos en listado de equipos                                                                                                                                                                                          *
#               La idea de este soft es facilitar la busqueda de activos en la planilla de equipos electronicos que disponemos.                                                                                                                                 *
#               La funcion es simple, por medio de los modulos openpyxl navegamos la planilla de excel existente con los datos que tenemos de los equipos. Buscamos por medio del nombre del equipo que se compone por: Area-Tipo de equipo-N° designado        *
#               por medio del modulo Tkinter le hicimos la interfaz grafica para que el usuario pueda utilizar la misma sin tener que ir ingresando lineas de comando
#****************************************************************************************************************************************************************************************************************************************************************
import cmd
import os as os
from select import select
import sys as sys
from sys import *
from os import replace, system as system
import datetime,time
#****************************************************************************************************************************************************************************************************************************************************************
import tkinter as tk
from tkinter import *
from tkinter import Entry, Grid, Image, StringVar, Text, Variable, messagebox, ttk, scrolledtext, simpledialog, tix, font, commondialog
from tkinter.ttk import Progressbar, Style, Treeview, setup_master,Sizegrip,Entry,Spinbox
from tkinter.tix import STATUS, LabelEntry,LabelFrame,Meter, ButtonBox,PhotoImage,ComboBox
from tkinter.constants import *
from turtle import color, delay, title, width
#****************************************************************************************************************************************************************************************************************************************************************
import openpyxl
from openpyxl import Workbook,load_workbook
#****************************************************************************************************************************************************************************************************************************************************************
#****************************************************************************************************************************************************************************************************************************************************************
#                                                                                                               Declaración de Variables                                                                                                                        *
#****************************************************************************************************************************************************************************************************************************************************************
#****************************************************************************************************************************************************************************************************************************************************************
new_user=os.environ['USERPROFILE']                                                                                                  # Identifico el "UserProfile" de la pc para poder encontrar las carpetas instaladas en el Setup.py                          *
user=new_user.replace("\\","/")                                                                                                     # Reemplazo "\" por "/" dado a que no reconocen la ruta en Python                                                           *

path=open(user+'/Desktop/Soft/Path.txt')                                                                                            # Busco el archivo donde se encuentran las rutas preestablecidas para encontrar los archivos. (Las cuales se pueden         *
lineas=path.readlines()                                                                                                             # modificar en caso que asi se quiera)                                                                                      *

log=open(user  + (str(lineas[1])[:-1]),mode="a")                                                                                    # Ruta de Archivo donde se encuentran los Logs de eventos                                                                   *
#****************************************************************************************************************************************************************************************************************************************************************
dia=        datetime.datetime.today().day                                                                                           # Variable de Dia                                                                                                           *
mes=        datetime.datetime.today().month                                                                                         # Variable de Mes                                                                                                           *
Hini=       datetime.datetime.now()                                                                                                 # Horario de inicio                                                                                                         *
Mensaje=    str()                                                                                                                   # Variable que utilizo para confeccionar los mensajes a mostrar en la aplicación                                            *
servidor=   str()                                                                                                                   # Variable que utilizo para poder realizar el MSTSC por el CMD                                                              *
dato=       str()                                                                                                                   # Variable para poder obtener datos y poder comparar                                                                        *
num=        int(1)                                                                                                                  # Variable para poder navegar por el listado Excel                                                                          *
limite=     3000                                                                                                                    # Limite de renglones para buscar                                                                                           *
Mensaje=    str()                                                                                                                   # Variable para poder mostrar mensaje                                                                                       *    
#****************************************************************************************************************************************************************************************************************************************************************
Texto=str()                                                                                                                         # Variable para poder escribir e imprimir en pantalla o CMD                                                                 *
sdia=int()                                                                                                                          # Variable para buscar "Día"                                                                                                *
smes=int()                                                                                                                          # Variable para buscar "Mes"                                                                                                *
Step=int()                                                                                                                          # Variable para marcar el paso a ejecutar en el programa                                                                    *
Fondo='#5B5B5B'                                                                                                                     # Color Amarillo de Prosegur'#FFCC01'
Ftree='#444444'
#****************************************************************************************************************************************************************************************************************************************************************
var=int()                                                                                                                           # Variable de uso general para pruebas                                                                                      *
#****************************************************************************************************************************************************************************************************************************************************************
renglon=30                                                                                                                          # Defino el valor a utilizar commo Renglones para poder las etiquetas sobre la ventana del soft                             *
wtree=str()                                                                                                                         # Utilizo este Str para poder mover el dato del Tree a "Modificar.py" y poder modificar dentro de las otras Tablas          
#****************************************************************************************************************************************************************************************************************************************************************
#****************************************************************************************************************************************************************************************************************************************************************
# Defina Interfaz Grafica                                                                                                                                                                                                                                       *
#****************************************************************************************************************************************************************************************************************************************************************
#****************************************************************************************************************************************************************************************************************************************************************
myapp = Tk()                                                                                                                        # MyApp es el nombre de la planilla a visualizar                                                                            *                                                                                                             
myapp.title("Buscador")                                                                                                             # Defino el titulo del programa                                                                                             *
h=700                                                                                                                               # Defino altura de la ventana                                                                                               *
w=1000                                                                                                                              # Defino ancho de la ventana                                                                                                *
myapp.minsize(w,h)                                                                                                                  # Defino el Tamaño minimo e inicial de plantilla                                                                            *
myapp.frame()
myapp.resizable(False,False)
myapp.config(background=Fondo)                                                                                                      # Defino color de Fondo de pantalla "Myapp"                                                                                 *
#****************************************************************************************************************************************************************************************************************************************************************
# Lineas de comando para poder centrar la ventana en la mitad de la pantalla (En teoria)
#****************************************************************************************************************************************************************************************************************************************************************
screen_width    = myapp.winfo_screenwidth()                                                                                         # En estas lineas de comando, su supone que centra la ventana de la aplicación                                              *
screen_height   = myapp.winfo_screenheight()                                                                                        # aun esta en modo prueba. Posiblemente lo elimine al finalizar el proyecto                                                 *
x_cordinate     = int((screen_width/2) - (w/2))
y_cordinate     = int((screen_height/2) - (h/2)-25)
myapp.geometry("{}x{}+{}+{}".format(w,h, x_cordinate, y_cordinate))
#****************************************************************************************************************************************************************************************************************************************************************
myimg=PhotoImage(file=(user+"\Desktop\Soft\lupa.png"))                                                                              # Variable para imagen del Boton de Busqueda. Se define la ruta en el programa.                                             *
mylogo=PhotoImage(file=(user  + (str(lineas[4]))[:-1]))                                                                             # Variable para imagen del Logo del Icono a usar                                                                            *
lbl_lable=ttk.Label(myapp,image=mylogo,border=0).place(x=-10,y=-10)                                                                 # Se define en Lable para poder utilizar "mylogo" como fondo de ventana.                                                    *
#****************************************************************************************************************************************************************************************************************************************************************
# Definimos StringVar para Respuestas                                                                                                                                                                                                                           *
#   Nombre      |   Tipo de Variable|                                                                                                                                                                                                                           *   
#****************************************************************************************************************************************************************************************************************************************************************
res_dispositivo=    StringVar()                                                                                                     # Variable para respuesta de Etiqueta de "Dispositivo"                                                                      *
res_ubi=            StringVar()                                                                                                     # Variable para respuesta de Etiqueta de "Ubicacion"                                                                        *
res_equipo=         StringVar()                                                                                                     # Variable para respuesta de Etiqueta de "Equipo"                                                                           *
res_nombre=         StringVar()                                                                                                     # Variable para respuesta de Etiqueta de "Nombre"                                                                           *
res_marca=          StringVar()                                                                                                     # Variable para respuesta de Etiqueta de "Marca"                                                                            *
res_ip=             StringVar()                                                                                                     # Variable para respuesta de Etiqueta de "Ip"                                                                               *
res_serial=         StringVar()                                                                                                     # Variable para respuesta de Etiqueta de "serial"                                                                           *
res_usuario=        StringVar()                                                                                                     # Variable para respuesta de Etiqueta de "Usuario"                                                                          *
res_password=       StringVar()                                                                                                     # Variable para respuesta de Etiqueta de "Contraseña"                                                                       *
res_server=         StringVar()                                                                                                     # Variable para respuesta de Etiqueta de "Servidores"                                                                       *
#****************************************************************************************************************************************************************************************************************************************************************
# Definimos el cuadro de entrada de Texto para buscar Celda                                                                                                                                                                                                     *
#****************************************************************************************************************************************************************************************************************************************************************
Entrada=    ttk.Entry(myapp,font=("Arial",12),width=20)                                                                             # Defino el renglon de entrada de datos para comparar                                                                       *
Entrada.place(x=10,y=35)                                                                                                            # Se define posición del renglon de "Entrada"                                                                               *
Entrada.insert(0,"Ingrese Valor de Celda")                                                                                          # Se coloca un texto inicial para que se muestre al iniciar el programa                                                     *
Entrada.get().upper()                                                                                                               # Configuro para que todo el dato ingresado se configure para poner en Mayusculas                                           *
Entrada.focus()                                                                                                                     # Se configura para tener el Foco al iniciar el programa                                                                    *
#****************************************************************************************************************************************************************************************************************************************************************
# Definimos Los Labels de Etiquetas y de respuestas                                                                                                                                                                                                             *                  *
# Nombre            |   Definicion de Label                                                                                                                                                                                                                     *
#****************************************************************************************************************************************************************************************************************************************************************
lbl_Dispositivo=        tk.Label(myapp,text="Dispositivo"                   ,bg=Fondo,font=("Arial",12)).place(x=10,y=10)           # Label que indica el "Dispositivo"                                                                                         *
lbl_Ubicacion=          tk.Label(myapp,text="Ubicacion del Dispositivo =>"  ,bg=Fondo,font=("Arial",12)).place(x=10,y=renglon*2)    # Label que indica el "Ubicacion"                                                                                           *
lbl_equipo=             tk.Label(myapp,text="Tipo de Equipo =>"             ,bg=Fondo,font=("Arial",12)).place(x=10,y=renglon*3)    # Label que indica el "Tipo de Equipo"                                                                                      *
lbl_nombre=             tk.Label(myapp,text="Nom    bre del equipo =>"          ,bg=Fondo,font=("Arial",12)).place(x=10,y=renglon*4)    # Label que indica el "Nombre del Equipo"                                                                                   *
lbl_marca=              tk.Label(myapp,text="Marca de Equipo =>"            ,bg=Fondo,font=("Arial",12)).place(x=10,y=renglon*5)    # Label que indica el "Marca del Equipo"                                                                                    *
lbl_Serial=             tk.Label(myapp,text="Serial del Dispositivo =>"     ,bg=Fondo,font=("Arial",12)).place(x=10,y=renglon*6)    # Label que indica el "N° de Serie del Equipo"                                                                              *
lbl_ip=                 tk.Label(myapp,text="IP de Equipo =>"               ,bg=Fondo,font=("Arial",12)).place(x=10,y=renglon*7)    # Label que indica el "IP del Equipo"                                                                                       *
lbl_usuario=            tk.Label(myapp,text="Usuario del equipo =>"         ,bg=Fondo,font=("Arial",12)).place(x=10,y=renglon*8)    # Label que indica el "Usuario de Configuración"                                                                            *
lbl_Password=           tk.Label(myapp,text="Password de Equipo =>"         ,bg=Fondo,font=("Arial",12)).place(x=10,y=renglon*9)    # Label que indica el "Password de Configuración"                                                                           *

lbl_res_ubicacion=      tk.Label(myapp,textvariable=res_ubi                 ,font=("Arial",12),width=50).place(x=300,y=renglon*2)   # Label de respuesta que indica "Ubicacion"                                                                                 * 
lbl_res_equipo=         tk.Label(myapp,textvariable=res_equipo              ,font=("Arial",12),width=50).place(x=300,y=renglon*3)   # Label de respuesta que indica "Equipo"                                                                                    *
lbl_res_nombre=         tk.Label(myapp,textvariable=res_nombre              ,font=("Arial",12),width=50).place(x=300,y=renglon*4)   # Label de respuesta que indica "Nombre"                                                                                    *
lbl_res_marca=          tk.Label(myapp,textvariable=res_marca               ,font=("Arial",12),width=50).place(x=300,y=renglon*5)   # Label de respuesta que indica "Marca"                                                                                     *
lbl_res_Serial=         tk.Label(myapp,textvariable=res_serial              ,font=("Arial",12),width=50).place(x=300,y=renglon*6)   # Label de respuesta que indica "Serial"                                                                                    *
lbl_res_Ip=             tk.Label(myapp,textvariable=res_ip                  ,font=("Arial",12),width=50).place(x=300,y=renglon*7)   # Label de respuesta que indica "IP"                                                                                        *
lbl_res_Usuario=        tk.Label(myapp,textvariable=res_usuario             ,font=("Arial",12),width=50).place(x=300,y=renglon*8)   # Label de respuesta que indica "Usuario"                                                                                   *
lbl_res_Password=       tk.Label(myapp,textvariable=res_password            ,font=("Arial",12),width=50).place(x=300,y=renglon*9)   # Label de respuesta que indica "Password"                                                                                  *
lbl_res_Server=         tk.Label(myapp,textvariable=res_server              ,font=("Arial",12),width=50).place(x=300,y=renglon*10)  # Label de respuesta que indica "Servidor"                                                                                  *
#****************************************************************************************************************************************************************************************************************************************************************                                                                                                                
#****************************************************************************************************************************************************************************************************************************************************************
# Diceñamos Arbol de seleccion de WorkSheet
#****************************************************************************************************************************************************************************************************************************************************************
s_t1="CCTV"                                                                                                                         # Colocamos el Nombre de la Tabla de archivo para poder visualizar dentro del Tree. Esta Tabla se Llama "CCTV"              *
s_t2="CA"                                                                                                                           # Tabla "CA"                                                                                                                *
s_t3="Lectoras"                                                                                                                     # Tabla "Lectoras"                                                                                                          *
s_t4="Avigilon"                                                                                                                     # Tabla "Avigilon"                                                                                                          *
                                                                                                                                    # NOTA: En caso de requerir mas tablas para mostrar solo agregar como variable el nombre de la Tabla para poder procesar    *
                               
tree=               ttk.Treeview(myapp,height=4,show='tree')                                                                        # Creamos el Listado del Arbol para poder seleccionar el Servidor a Realizar conexion remota                                *   
tree.place          (x=780,y=320)
tree.tag_configure  ('arbol',background=Ftree,font=("Snap ITC",11),foreground='lightblue')                                          # Codificamos el estilo del tag para poder decorar mejor el mismo.                                                          *
tree.tag_configure  ('arbol2',background=Ftree,font=("Snap ITC",11),foreground='red')                                               # Codificamos otro estilo para ver si podemos modificar el estilo cuando se hace foco sobre el Tag                          *
estilo='arbol'

t1_image=       PhotoImage(file=(user+"/Desktop/Soft/CCTV.png"))                                                                    # Definimos el icono a mostrar como "CCTV"                                                                                  *
t2_image=       PhotoImage(file=(user+"/Desktop/Soft/electrician-tools.png"))                                                       # Icono a mostrar como "CA"                                                                                                 *
t3_image=       PhotoImage(file=(user+"/Desktop/Soft/HID.png"))                                                                     # Icono a mostrar como "Lectoras"                                                                                           *
t4_image=       PhotoImage(file=(user+"/Desktop/Soft/CCTV_2.png"))                                                                  # Icono a mostrar como "Avigilon"                                                                                           *

def item_seleccionado (event):                                                                                                      # Evento creado para poder probar el cambio de estilo el Item del Tree (Modo Prueba)                                        *
    estilo='arbol2'
    print('Item Seleccionado')

tree.tag_bind("mytag", "<<TreeviewSelect>>", item_seleccionado)                                                                     # Se define el nombre del evento, Tipo de disparador y nombre del evento                                                    *

tree.insert     ('', tk.END, text=s_t1, iid=0, open=False,image=t1_image,tags=(estilo,"mytag"))                                     # Defino el 1er tag
tree.insert     ('', tk.END, text=s_t2, iid=1, open=False,image=t2_image,tags='arbol')                                              # Defino el 2do tag
tree.insert     ('', tk.END, text=s_t3, iid=2, open=False,image=t3_image,tags='arbol')                                              # Defino el 3er tag
tree.insert     ('', tk.END, text=s_t4, iid=3, open=False,image=t4_image,tags='arbol')                                              # Defino el 4to tag
#****************************************************************************************************************************************************************************************************************************************************************
# Funciones del Programa a utilizar                                                                                                                                                                                                                             *
#****************************************************************************************************************************************************************************************************************************************************************   
def Ayuda():
    system('start %userprofile%\Desktop\Soft\Help.txt')                                                                             # Abro por Sistema el Txt de ayuda del programa                                                                             *
    return
#****************************************************************************************************************************************************************************************************************************************************************                        
def Siguiente():                                                                                                                    # Mostramos el valor siguiente al que tenemos en el buscador                                                                *
    
    global num,dispositivo                                                                                                          # Tomo las Variables Globales y le permito el uso dentro del subprograma                                                    *
    
    num+=1                                                                                                                          # Aumento la variable "num"                                                                                                 *

    wb=openpyxl.load_workbook(str(lineas[0])[:-1])                                                                                  # Genero el vinculo y abro el Excel que esta definido dentro del String dentro del Path                                     *     
    current_item = tree.focus()    
    wbs = tree.item(current_item,option='tex    t') 
    ws=wb[wbs]                                                                                                                      # Abro la Hoja (ws) CCTV para poder trabajar con esos datos
    dispositivo=ws.cell(row=num,column=4)                                                                                           # Defino la variable para poder compara con la variable "Entrada" teniendo en cuenta los parametros "num" como renglones    *
                                                                                                                                    #y como columna una sola definitiva para poder buscar solamente por nombre de equipo                                        *   
   
    if dispositivo.value=="":                                                                                                       # Verifico que la variable no este vacia                                                                                    *
        messagebox.showerror(message="Final de listado sin datos similares")                                                        # En caso de que se encuentra vacia, mostramos mensaje indicando que llego al final del listado de equipos                  *
    else:
        Imprimir()                                                                                                                  # En caso de que tengamos información dentro de ese renglon, pasamos los datos al sub Imprimi()                             *
#****************************************************************************************************************************************************************************************************************************************************************    
def Anterior():                                                                                                                     # Mostramos el valor anteroor al que tenemos en el buscador                                                                 *
    
    global num, dispositivo                                                                                                         # Tomo las Variables Globales y le permito el uso dentro del subprograma                                                    *
    
    if num>0:                                                                                                                       # En caso de que el Num sea mayor a 0, reducimos el valor de la variable                                                    * 
        num-=1                                                                                                                      # Se reduce el valor de la variable                                                                                         *
        Imprimir()                                                                                                                  # Ejecutamos el Sub Imprimir                                                                                                *
    else:
        tk.messagebox.showerror(message="No se encuentra valor menor")                                                              # Caso que num sea =/< a 0, mostramos un mensaje de error para indicar que ya no se puede reducir mas esa variable          *  
#****************************************************************************************************************************************************************************************************************************************************************
def Buscar():                                                                                                                       # Buscamos el valor ingresado dentro del listado de equipos que tenemos en el listado                                       *
    
    global num,limite                                                                                                               # Importo variables                                                                                                         *
    global dispositivo
    num=1                                                                                                                           # Configuro la variable en valor incial, para poder buscar desde el inicio del listado                                      *
    
    current_item = tree.focus()                                                                                                     # Configuro la variable con el Focus del Tree, para poder utilizarlo como indicador de Ws.                                  *
    wbs = tree.item(current_item,option='text')                                                                                     # Selecciono el item "Current_item" y lo configuro como "Text" y lo paso a la variable para utilizar                        *
    wtree=wbs                                                                                                                       # Paso "wbs" a "wtree" que es un str
    
    #wb=openpyxl.load_workbook(user + str(lineas[0])[:-1])                                                                          # Corresponde a la ruta en caso de tenerlo en el servidor                                                                   *
    wb=openpyxl.load_workbook(str(lineas[0])[:-1])                                                                                  # Corresponde a la ruta en caso de tenerlo en la pc en forma local                                                          *
    ws=wb[wbs]
   
    dato=Entrada.get().upper()                                                                                                      # Tomo lo escrito en el Entry y lo paso a Upper para tener un control del ingreso                                           *
    dispositivo=ws.cell(row=num,column=4)                                                                                           # Tomo lo escrito dentro del Cell del listado segun el Num que es = al renglon del listado                                  *
    
    while dato!= dispositivo.value:                                                                                                 # Mientras que el valor de la Celda (Dispositivo) sea distinto a la entrada (Dato) aumenta en 1 el valor de Num (Renglon)   *                                  
        dispositivo=ws.cell(row=num,column=4)                                                                                       # vuelvo a marcar la franja en donde tengo que buscar en las tablas.                                                        *
        if num==limite:                                                                                                             # Comparamos los parametros Num con Limite, ya que sino la busqueda pasaria por toda la planilla inecesariamente            *
            num=1                                                                                                                   # Cuando llegamos al "Limite", reiniciamos la Variable Num para volver a buscar                                             *
            messagebox.showerror(title="Error 421",message="Por favor, verifique el dato a buscar")                                 # Mostramos un Mensaje de Error informando lo sucedido                                                                      *
                                                                                                                                            #,image="C:\user\desktop\codigo\nuclear.ico")
            return
        else:                                                                                                                       # Cuando la variable sea menor o distinta al Limite, aumentamos en 1 para navegar la planilla                               *
            num=num+1
        
    else: 
        num-=1                                                                                                                      # Cuando el Dato y el Dispositivo son iguales, volvemos un renglon para atras y mostramos ese renglon                       *
        Imprimir()                                                                                                                  # Ejecutamos el SubPrograma Imprimir()                                                                                      *
#****************************************************************************************************************************************************************************************************************************************************************        
def Imprimir():                                                                                                                     # Se imprimen los datos de la planilla encontrados segun lo Buscar(),Siguiente( y Anterior())                               *
    
    global dato,num,dispositivo                                                                                                     # Importamos las variables a utilizar / procesar                                                                            *
    
    current_item = tree.focus()                                                                                                     # Similar al Buscar(), uso el Item Seleccionado del Arbol como dato para identificar la tabla a buscar                      *
    wbs = tree.item(current_item,option='text') 
    #wb=openpyxl.load_workbook(user + str(lineas[0])[:-1])
    wb=openpyxl.load_workbook(str(lineas[0])[:-1])                                                                                  # Abro la Planilla de Excel (Se especifica ruta en el Path adjunto en el Archivo)                                           *
    ws=wb[wbs]                                                                                                                      # Seleccionamos la Tabla de acuerdo a lo seleccionado en le Arbol de Tablas                                                 *
    wtree=wbs                                                                                                                       # pasamo el valor al Str() para poder trabajar sin problemas                                                                *
    
    dato=Entrada.get().upper()                                                                                                      # Agarro el texto ingresado y lo convierto en Mayusculas, solo para poder tener todo standarizado                           *
    dispositivo=ws.cell(row=num,column=4)                                                                                           # Tomo el valor de la celda para comparar                                                                                   *
    
    log=open(user  + (str(lineas[1])[:-1]),mode="a")                                                                                # Abrimos el Archivo .txt que utilizamos para registrar los logs. Este esta enrutado en el Path                             *
    log.write(str(Hini)+" => "+user[9:30] + " esta buscado el dato " +dato+" que esta es el registro numero => "+str(num)+"worksheet => "+wbs+"\n")
    log.close()                                                                                                                     # Escribimos Horario, usuario, dato, renglon y tabla buscado para poder tener registro de lo realizado.
    
    dispositivo=    ws.cell(row=num,column=4)
    nombre=         ws.cell(row=num,column=4)
    equipo=         ws.cell(row=num,column=2)
    ubicacion=      ws.cell(row=num,column=3)
    marca=          ws.cell(row=num,column=6)
    modelo=         ws.cell(row=num,column=7)
    ip=             ws.cell(row=num,column=11)
    serial=         ws.cell(row=num,column=8)
    usuario=        ws.cell(row=num,column=12)
    password=       ws.cell(row=num,column=13)
    server=         ws.cell(row=num,column=14)    
           
    res_nombre.set      (nombre.value)
    res_equipo.set      (equipo.value)
    res_ubi.set         (str(ubicacion.value))
    res_marca.set       (str(marca.value) +"   =>  "+ str(modelo.value)) 
    res_ip.set          (ip.value)
    res_serial.set      (serial.value)
    res_usuario.set     (usuario.value)
    res_password.set    (password.value)
    res_server.set      (server.value)
#****************************************************************************************************************************************************************************************************************************************************************    
def Modifi():
    
    global wtree

    qst=messagebox.askokcancel(title="Modificar Parametros",message="Esta seguro que decea Modificar los Parametros???")
    if qst==True:
        current_item = tree.focus()    
        wbs = tree.item(current_item,option='text') 
        wtree=str(wbs)
        myapp.destroy()
        import Modificar
        
    else:
        Mensaje="Se cancelo modificación solicitada"
        messagebox.showerror(title="Modificación",message=Mensaje)
#****************************************************************************************************************************************************************************************************************************************************************        
def Conectar():

    wb=openpyxl.load_workbook(str(lineas[0])[:-1])
    ws=wb["CCTV"]
    server=ws.cell(row=num,column=14) 
    servidor='cmd /k "mstsc -v ' + server.value + ':4489'
    print(servidor)
    myapp.destroy()
    os.system(servidor)
    
    return 
#****************************************************************************************************************************************************************************************************************************************************************    
def Salir():
    
    log=open(user  + (str(lineas[1])[:-1]),mode="a")
    log.write(str(Hini)+" => Se cierra aplicación "+user[9:]+"\n")
    log.close()
    Destroy()
    
def Destroy():
    myapp.destroy()
#****************************************************************************************************************************************************************************************************************************************************************
# Definimos los botoes a ver en la ventana inicial                                                                                                                                                                                                              *
#****************************************************************************************************************************************************************************************************************************************************************
boton=      tk.Button(myapp,text="Buscar",activebackground="#ABCDEF",background="#838B8B",command=Buscar,width=180,image=myimg).place(x=780,y=60)                               # Creo Boton "planilla" para procesar las plantillas Requeridas para informe.   *
salir=      tk.Button(myapp,text="Salir",activebackground="#BABABA",command=Salir,justify='center',width=23).place(x=790,y=650)                                                 # Creo un Boton para cerrar la aplicación                                       *
bsiguiente= tk.Button(myapp,text="Siguiente",activebackground="#ABABAB",background="#838383",command=Siguiente,width=11,state='active').place(x=780,y=150)
banterios=  tk.Button(myapp,text="Previo",activebackground="#ABABAB",background="#838383",command=Anterior,width=11,state='active').place(x=880,y=150)
bmodificar= tk.Button(myapp,text="Modificar",activebackground="#ABABAB",background="#838383",command=Modifi,width=25,state='active').place(x=780,y=180)
bhelp=      tk.Button(myapp,text="Ayuda",background="#838383",command=Ayuda,width=5,).place(x=15,y=650)                                                                         # Creo el Boton de "Ayuda" para mostrar el Txt correspondiente                  *
bconectar=  tk.Button(myapp,text="Conectar",activebackground="#ABABAB",background="#838383",command=Conectar,width=25,state='active').place(x=780,y=500)
#****************************************************************************************************************************************************************************************************************************************************************                     
#****************************************************************************************************************************************************************************************************************************************************************
log.close()
myapp.mainloop()
#****************************************************************************************************************************************************************************************************************************************************************


