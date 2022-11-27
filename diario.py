#****************************************************************************************************************************************************************************************************************************************************************
# Nombre: Buscador.Py                                                                                                                                                                                                                                           *       
# Descripcion: Aplicación para poder monitorear y visualizar las tareas diarias a realizar                                                                                                                                                                      *
#               Consiste en un desarroyo que asiste a la visualización de tareas programadas y carga de tareas realizadas en el dia. Tambien permite que se mofiquen tareas anteriores pudiendo grabar las mismas desde la app.                                 *
# #**************************************************************************************************************************************************************************************************************************************************************
from calendar import weekday
import cmd
import os as os
import sys as sys
from sys import *
from os import replace, system as system
import datetime,time
#****************************************************************************************************************************************************************************************************************************************************************
import tkinter as tk
from tkinter import *
from tkinter import Entry, Grid, Image, StringVar, Text, Variable, messagebox, ttk, scrolledtext, simpledialog, tix, font, commondialog
from tkinter import dialog
from tkinter.ttk import Notebook, Progressbar, setup_master,Sizegrip,Entry,Checkbutton,Treeview
from tkinter.constants import *
from turtle import bgcolor, color, delay, heading, title, width
#****************************************************************************************************************************************************************************************************************************************************************
import openpyxl
from openpyxl import Workbook,load_workbook
from openpyxl.styles import PatternFill
#****************************************************************************************************************************************************************************************************************************************************************
from Variables import *
#****************************************************************************************************************************************************************************************************************************************************************
#                                                                                                               Declaración de Variables                                                                                                                        *
#****************************************************************************************************************************************************************************************************************************************************************
#****************************************************************************************************************************************************************************************************************************************************************
#****************************************************************************************************************************************************************************************************************************************************************
# Defina Interfaz Grafica                                                                                                                                                                                                                                       *
#****************************************************************************************************************************************************************************************************************************************************************
#****************************************************************************************************************************************************************************************************************************************************************
mydia = Tk()                                                                                                                        # MyApp es el nombre de la planilla a visualizar                                                                            *                                                                                                             
mydia.title("Diario")                                                                                                               # Defino el titulo del programa                                                                                             *
h=700
w=1350
mydia.minsize(w,h)                                                                                                                  # Defino el Tamaño minimo e inicial de plantilla                                                                            *
#myapp.maxsize(1050,900)                                                                                                            # Defino el Tamaño al Maximar                                                                                               *
mydia.resizable(False,False)
#myapp.geometry('1000x500')                                                                                                         # Defino el Tamaño de la Grilla para poder colocar el resto de los objetos                                                  *
mydia.config(background=Fondo)                                                                                                      # Defino color de Fondo de pantalla "Myapp"                                                                                 *
#****************************************************************************************************************************************************************************************************************************************************************
#   Lineas de comando para poder centrar la ventana en la mitad de la pantalla (En teoria)
#****************************************************************************************************************************************************************************************************************************************************************
screen_width =  mydia.winfo_screenwidth()                                                                                           # Definimos las dimenciones de la ventana de la aplicacion.
screen_height = mydia.winfo_screenheight()
x_cordinate =   int((screen_width/2) - (w/2))
y_cordinate =   int((screen_height/2) - (h/2)-26)
mydia.geometry("{}x{}+{}+{}".format(w,h, x_cordinate, y_cordinate))
#****************************************************************************************************************************************************************************************************************************************************************
myimg=      PhotoImage(file=(user+"/Desktop/Soft/actualizar.png"))                                                                  # Variable para imagen del Boton de Busqueda. Se define la ruta en el programa. Tendria que ver de ponerlo en el Path       *                                                                                                           # Defino el valor a utilizar commo Renglones para poder las etiquetas sobre la ventana del soft                             *
#****************************************************************************************************************************************************************************************************************************************************************
mylogo=     PhotoImage(file=(user  + (str(lineas[4]))[:-1]))                                                                        # Variable para imagen del Logo del Icono a usar                                                                            *
lbl_lable=ttk.Label(mydia,image=mylogo,border=0).place(x=-10,y=-10)   
#****************************************************************************************************************************************************************************************************************************************************************
sct=scrolledtext.ScrolledText(mydia)                                                                                                # Creamos el "ScrolledText", es para poder mostrar los mensajes de progreso del soft                                        *
sct.place(x=800,y=300,width=500,height=200)                                                                                         # Configuramos la posición del "ScrolledText"                                                                               *
#****************************************************************************************************************************************************************************************************************************************************************
#****************************************************************************************************************************************************************************************************************************************************************
# Definimos Variables para respuestas                                                                                                                                                                                                                           *
#****************************************************************************************************************************************************************************************************************************************************************
res_celda=  StringVar()                                                                                                             # Variable para respuesta de Etiqueta de "Celda"                                                                            *
res_t1=     StringVar()                                                                                                             # Variable para respuesta de Etiqueta de "T1"                                                                               *
res_t2=     StringVar()                                                                                                             # Variable para respuesta de Etiqueta de "T2"                                                                               *
res_t3=     StringVar()                                                                                                             # Variable para respuesta de Etiqueta de "T3"                                                                               *
res_t4=     StringVar()                                                                                                             # Variable para respuesta de Etiqueta de "T4"                                                                               *
res_t5=     StringVar()                                                                                                             # Variable para respuesta de Etiqueta de "T5"                                                                               *
res_t6=     StringVar()                                                                                                             # Variable para respuesta de Etiqueta de "T6"                                                                               *
res_t7=     StringVar()                                                                                                             # Variable para respuesta de Etiqueta de "T7"                                                                               *
#****************************************************************************************************************************************************************************************************************************************************************
l_t1=       StringVar()                                                                                                             # Variable para manipular el Label relacionado a "T1"                                                                       *
l_t2=       StringVar()                                                                                                             # Variable para manipular el Label relacionado a "T2"                                                                       *
l_t3=       StringVar()                                                                                                             # Variable para manipular el Label relacionado a "T3"                                                                       *
l_t4=       StringVar()                                                                                                             # Variable para manipular el Label relacionado a "T4"                                                                       *
l_t5=       StringVar()                                                                                                             # Variable para manipular el Label relacionado a "T5"                                                                       *
l_t6=       StringVar()                                                                                                             # Variable para manipular el Label relacionado a "T6"                                                                       *
l_t7=       StringVar()                                                                                                             # Variable para manipular el Label relacionado a "T7"                                                                       *
#****************************************************************************************************************************************************************************************************************************************************************
cdia=       IntVar()                                                                                                                # Variable para respuesta de Etiqueta de "cdia"                                                                             *
dcontrol=   IntVar()                                                                                                                # Variable para respuesta de Etiqueta de "dcontrol"                                                                         *
controldia= IntVar()                                                                                                                # Variable para utilizar de control para poder hacer calculo                                                                *
ddia=       StringVar()                                                                                                             # Utilizamos esta variable para mostrar el String el Dia de la semana.                                                      *
#****************************************************************************************************************************************************************************************************************************************************************
dcontrol=Dia+22                                                                                                                     # En este sector le sumamos 22 para poder ver los rengloes con los controles efectuados                                     * 
ddia.set(str(Dia))                                                                                                                  # Configuramos la variable para poder ver el dia en que nos encontramos                                                     *
print("dia => "+str(ddia.get()))                                                                                                    # Muestra en el CMD el valor del dia                                                                                        *
#****************************************************************************************************************************************************************************************************************************************************************
# Defino dia de la semana para poder visualizar tareas rutinarias                                                                                                                                                                                               *
#****************************************************************************************************************************************************************************************************************************************************************
# Detalles;                                                                                                                                                                                                                                                     *
# En esta parte definimos los vectores para convertir los valores del valor del dia a el nombre de los dias propiamente dichos y asi poder cotejarlos en la planilla correspondiente.                                                                           *
#****************************************************************************************************************************************************************************************************************************************************************
semana=str(Hini.strftime('%w'))                                                                                                     # Deinimos el dia de la Semana segun el valor dela variable que proporciona el sistema.                                     *
dsemana=StringVar()                                                                                                                 # Definimos variable para poder comparar                                                                                    *

if semana=="1":
    dsemana.set('Lunes')                                                                                                            # Si la Var es "1" es Lunes                                                                                                 *
elif semana=="2":
    dsemana.set('Martes')                                                                                                           # Si la Var es "2" es Martes                                                                                                *
elif semana=="3":
    dsemana.set('Miercoles')                                                                                                        # Si la Var es "3" es Miercoles                                                                                             *
elif semana=="4":
    dsemana.set('Jueves')                                                                                                           # Si la Var es "4" es Jueves                                                                                                *
elif semana=="5":
    dsemana.set('Viernes')                                                                                                          # Si la Var es "5" es Viernes                                                                                               *
elif semana=="6":
    dsemana.set('Sabado')                                                                                                           # Si la Var es "6" es Sabado                                                                                                *
else:
    dsemana.set("Domingo")                                                                                                          # Si no cumple con las otras opciones es Domingo                                                                            *
#****************************************************************************************************************************************************************************************************************************************************************
# Detalles:                                                                                                                                                                                                                                                     *
# Escribimos en el log, el registro de la fecha y el usuario en que se ejecuta.                                                                                                                                                                                 *
#****************************************************************************************************************************************************************************************************************************************************************
Mensaje=("Hoy es=> " + str(dsemana.get())+" "+str(Dia)+" / "+str(Mes)+"\n")                                                         # Imprimimos el numero y mes en que se ejecuta el programa                                                                  *
print (Mensaje)
import WLog                                                                                                                         # Cerramos el Log                                                                                                           *
#****************************************************************************************************************************************************************************************************************************************************************
if Dia <15:                                                                                                                         # De acuerdo a la planilla que diceñamos, si el dia supera los 15 debemos modificar el valor para poder ver que tareas hacer*
    cdia.set(Dia)                                                                                                                   # hasta el 15 respetamos el valor del dia                                                                                   *
    ddia.set(Dia)
else:                                                                                                                               # cuando supera el 15 le restamos 15 para volver al listado desde el inicio                                                 *
    cdia.set(Dia-15)                                                                                                                                                                                                                                           #*
    ddia.set(Dia)
    print("Dia es => "+ str(ddia.get()))
    print ("Cdia => "+ str(cdia.get()))
#****************************************************************************************************************************************************************************************************************************************************************
wb=openpyxl.load_workbook(str(lineas[0])[:-1])                                                                                      # Corresponde a la ruta en caso de tenerlo en la pc en forma local                                                          *
ws=wb["Calendario"]
 
t1=         ws.cell     (row=int(semana),column=2)                                                                                  # Tomo el valor de la celda para la variables del calendario Tareas semanales                                               *
t2=         ws.cell     (row=int(semana),column=3)
t3=         ws.cell     (row=int(semana),column=4)
t4=         ws.cell     (row=int(semana),column=5)
t5=         ws.cell     (row=cdia.get(),column=2)                                                                                   # A partir de esta variable tomamos las tareas mensuales definidas                                                          *
t6=         ws.cell     (row=cdia.get(),column=3)
t7=         ws.cell     (row=cdia.get(),column=4)
#****************************************************************************************************************************************************************************************************************************************************************
res_t1.set  (t1.value)                                                                                                              # Configuramos el VarString con relacion al valor de la celda                                                               *    
res_t2.set  (t2.value)                                                                                                              # Estos StringVar sirven para tomar los valores de la tabla y poder ponerlos en los labels de respuesta                     *
res_t3.set  (t3.value)
res_t4.set  (t4.value)
res_t5.set  (t5.value)
res_t6.set  (t6.value)
res_t7.set  (t7.value)
#****************************************************************************************************************************************************************************************************************************************************************
# Servidores a Realizar Mantenimiento                                                                                                                                                                                                                           *
#****************************************************************************************************************************************************************************************************************************************************************
s_t1=       ws.cell     (row=int(semana),column=7)                                                                                  # Esta variable toma los valores correspondientes a los nombres de los Servidores correspondientes a los puntos de tareas   *
s_t2=       ws.cell     (row=int(semana),column=8)                                                                                  #diarias correspondientes a la semana                                                                                       *
s_t3=       ws.cell     (row=int(semana),column=9) 
s_t4=       ws.cell     (row=cdia.get(),column=7)
s_t5=       ws.cell     (row=cdia.get(),column=8)
s_t6=       ws.cell     (row=cdia.get(),column=9)                                                                                            # A partir de esta variable se toma los servidores correspondientes a las tareas mensuales                                  *
print ("Semana => "+ str(semana))   
#****************************************************************************************************************************************************************************************************************************************************************
# Se configura visual de Arbol para poder realizar conexiones a Servidor Directamente                                                                                                                                                                           *
#****************************************************************************************************************************************************************************************************************************************************************
# Detalles:                                                                                                                                                                                                                                                     *
# Se configura el arbol para poder mostrar el listado de servidores en relación a las tareas correspondientes de los dias. Lo que hacemos es tomar el valor de los servidores relacionados a las tareas asignadas en la planilla y se van agregando             *
# En caso de que no tenga servidor, se saltea el nivel y seguimos con el siguiente servidor                                                                                                                                                                     *
#****************************************************************************************************************************************************************************************************************************************************************
tree=ttk.Treeview(mydia,height=8,show='tree',selectmode='browse')                                                                   # Creamos el Listado del Arbol para poder seleccionar el Servidor a Realizar conexion remota                                *
pos=int(0)                                                                                                                          # Inicializamos la variable posición en 0. Con esta var vamos a navegar por los "Renglones" del arbol                       *

print ("valor de S_t1 => " + str(s_t1.value))

if s_t1 !="-":                                                                                                                      # En caso de que la variable sea distinta a "-", insertamos ese valor en el arbol para seleccionarlo                        *
    tree.insert('', tk.END, text=str(s_t1.value), iid=pos, open=False)                                                              # Colocamos el valor en el arbol y luego aumentamos la poasicion para colocar el siguiente valor.Sino se respeta el nivel   *
    pos=pos+1
if s_t2 !="-":                                                                                                                      # En esta caso, se repite la misma secuencia que la anterior, si es distinto se ingresa el valor en el nivel que marco en el*
    tree.insert('', tk.END, text=s_t2.value, iid=pos, open=False)                                                                   #nivel anterior. En caso de que no tenga valor, no se aumenta y se pondra el siguiente valor                                *
    pos=pos+1                                                                                                                       
if s_t3 !="-":
    tree.insert('', tk.END, text=s_t3.value, iid=pos, open=False)
    pos=pos+1
if s_t4 !="-":
    tree.insert('', tk.END, text=s_t4.value, iid=pos, open=False)
    pos=pos+1
if s_t5 !="-":
    tree.insert('', tk.END, text=s_t5.value, iid=pos, open=False)
    pos=pos+1
if s_t6 !="-":
    tree.insert('', tk.END, text=s_t6.value, iid=pos, open=False)   
    
tree.place(x=900,y=80)                                                                                                              # Definimos la posición del arbol dentro de la ventana de la app                                                            *
mydia.update()                                                                                                                      # Hacemos un apdate de la ventana                                                                                           *
#****************************************************************************************************************************************************************************************************************************************************************
# Defino las variables a utilizar para los CheckBox                                                                                                                                                                                                             *
# Detalles:                                                                                                                                                                                                                                                     *
#   Estas variables las utilizamos para validar los checkbox para ver el estado de las tareas                                                                                                                                                                   *
#****************************************************************************************************************************************************************************************************************************************************************
var_t1=     IntVar()                                                                                                                # Definimos las variable para los checkbox de validacion de estado de Tareas.                                               *
var_t2=     IntVar()
var_t3=     IntVar()
var_t4=     IntVar()
var_t5=     IntVar()
var_t6=     IntVar()
var_t7=     IntVar()
#****************************************************************************************************************************************************************************************************************************************************************
# Definimos Los Checkbotons a Mostrar                                                                                                                                                                                                                           *
#****************************************************************************************************************************************************************************************************************************************************************
chk_t1=     tk.Checkbutton(mydia        ,activebackground="#FF0000",        background="#0B0B0B",       onvalue=1,      offvalue=0,     variable=var_t1)        .place(x=500,y=Renglon*3) # Check botons creados para validar o no las tareas del dia, cada uno por cada tarea                            *                                                                                                                 
chk_t2=     tk.Checkbutton(mydia        ,activebackground="#FF0000",        background="#0B0B0B",       onvalue=1,      offvalue=0,     variable=var_t2)        .place(x=500,y=Renglon*4)
chk_t3=     tk.Checkbutton(mydia        ,activebackground="#FF0000",        background="#0B0B0B",       onvalue=1,      offvalue=0,     variable=var_t3)        .place(x=500,y=Renglon*5)
chk_t4=     tk.Checkbutton(mydia        ,activebackground="#FF0000",        background="#0B0B0B",       onvalue=1,      offvalue=0,     variable=var_t4)        .place(x=500,y=Renglon*6)
chk_t5=     tk.Checkbutton(mydia        ,activebackground="#FF0000",        background="#0B0B0B",       onvalue=1,      offvalue=0,     variable=var_t5)        .place(x=500,y=Renglon*7)
chk_t6=     tk.Checkbutton(mydia        ,activebackground="#FF0000",        background="#0B0B0B",       onvalue=1,      offvalue=0,     variable=var_t6)        .place(x=500,y=Renglon*8)
chk_t7=     tk.Checkbutton(mydia        ,activebackground="#FF0000",        background="#0B0B0B",       onvalue=1,      offvalue=0,     variable=var_t7)        .place(x=500,y=Renglon*9)
#****************************************************************************************************************************************************************************************************************************************************************
# Defino Variables para control de estado de Celdas de Excel                                                                                                                                                                                                    *
# Detalles:                                                                                                                                                                                                                                                     *
#   Aca se tomas los valores de las celdas de la planilla para mostrar en la app                                                                                                                                                                                *
#****************************************************************************************************************************************************************************************************************************************************************
c_t1=       ws.cell(row=dcontrol,column=2)                                                                                      # Variable para tomar y verificar el estado de las tareas realizadas. Esto nos sirve para controlar tareas pendientes           *
c_t2=       ws.cell(row=dcontrol,column=3)                                                                                      # a realizar y que se continuaran dias posteriores                                                                              *
c_t3=       ws.cell(row=dcontrol,column=4)
c_t4=       ws.cell(row=dcontrol,column=5)
c_t5=       ws.cell(row=dcontrol,column=8)
c_t6=       ws.cell(row=dcontrol,column=9)
#****************************************************************************************************************************************************************************************************************************************************************
var_t1.set  (c_t1.value)                                                                                                        # Seteamos los valores de las variables en realacion a los valores que se toman desde el Excel y asi podemos hacer un control   *
var_t2.set  (c_t2.value)
var_t3.set  (c_t3.value)
var_t4.set  (c_t4.value)
var_t5.set  (c_t5.value)
var_t6.set  (c_t6.value)
#****************************************************************************************************************************************************************************************************************************************************************
#Defino variables para poder limpiar en el Excel las celdas de control marcadas                                                                                                                                                                                 *
#****************************************************************************************************************************************************************************************************************************************************************
# Detalles:                                                                                                                                                                                                                                                     *
#   Tomamos la celda y limpiamos el estado de la actividad que quedo configurado de actividades anteriores. Esto sucede cuando se configura el estado de la tarea mesual y semanal y queda. Por lo que limpiamos el estado y colocamos el valor como si         *
# la tarea se encontraria pendiente
#****************************************************************************************************************************************************************************************************************************************************************
clr_t1=       ws.cell(row=(dcontrol+1),column=2)                                                                                # Con estas varibles lo que hacemos es limpiar los valores que quedaron almacenadas de trabajos realizados pero en la celdas    *
clr_t2=       ws.cell(row=(dcontrol+1),column=3)                                                                                # siguientes a las de la fecha. Al ir completando semana tras semana y configurando el valor de trarea realizada o no quedan    *
clr_t3=       ws.cell(row=(dcontrol+1),column=4)                                                                                # marcadas y en caso de transcurrido un tiempo se vuelve a la misma celda definida en el mes. Por lo que para no tener falsas   *
clr_t4=       ws.cell(row=(dcontrol+1),column=5)                                                                                # lecturas, corregimos estos valores con esta var
clr_t5=       ws.cell(row=(dcontrol+1),column=8)
clr_t6=       ws.cell(row=(dcontrol+1),column=9)

clr_t1=clr_t2=clr_t3=clr_t4=clr_t5=clr_t6=0
#****************************************************************************************************************************************************************************************************************************************************************
# Definimos Labels para Agregar Tareas Manuales.                                                                                                                                                                                                                *
#****************************************************************************************************************************************************************************************************************************************************************
# Detalles:                                                                                                                                                                                                                                                     *
# En esta parte del Scrip lo que hacemos es tomar los valores de la planilla en relacion a las Tareas realizadas y las visualizamos en al aplicacion, para esto creamos varias StringVars para tomar los valores y poder modificarlos o guardarlos en la        *
# planilla. Con esto podemos ver las tareas y poder anotar las nuevas                                                                                                                                                                                           *
#****************************************************************************************************************************************************************************************************************************************************************    
cn=             IntVar()
wb=             openpyxl.load_workbook(str(lineas[0])[:-1])                                                                         # Corresponde a la ruta en caso de tenerlo en la pc en forma local                                                          *
ws=             wb["Tareas"]                                                                                                        # Abrimos la hoja "Tareas"
dispositivo=    ws.cell(row=num,column=1)                                                                                           # Creamos la variable para poder tomar el valor del disposicitivo y poder cotejarlo con lo ingresado                        *

while dispositivo.value!=None:                                                                                                      # Verificamos que el renglon contenga información para poder ver el contenido del mismo. Desde este punto podemos navegar   *
        num=num+1                                                                                                                   #por el listado de tareas Realizadas.                                                                                       *
        dispositivo=ws.cell(row=num,column=1)    
#****************************************************************************************************************************************************************************************************************************************************************
TFecha=     StringVar   (value=(ws.cell(column=1,row=num-1)).value)                                                                 # Definimos el StringVar para mostrar la Fecha,tomando el valor de fecha desde la planilla                                  *
TTarea=     StringVar   (value=(ws.cell(column=2,row=num-1)).value)                                                                 # Definimos el StringVar para mostrar la Tarea a / o realizada                                                              *
TEqui=      StringVar   (value=(ws.cell(column=3,row=num-1)).value)                                                                 # Definimos el StringVar para mostrar el Equipo a/en donde se realiza las tareas                                            *
TSector=    StringVar   (value=(ws.cell(column=4,row=num-1)).value)                                                                 # Definimos el StringVar para mostrar la Sector que corresponde el equipo                                                   *
TStatus=    StringVar   (value=(ws.cell(column=5,row=num-1)).value)                                                                 # Definimos el StringVar para mostrar el Estado en que se encuentra la tarea correspondiente                                *
TAsig=      StringVar   (value=(ws.cell(column=6,row=num-1)).value)                                                                 # Definimos el StringVar para mostrar a quien se le Asigno la tarea indicanda                                               *
TNotas=     StringVar   (value=(ws.cell(column=7,row=num-1)).value)                                                                 # Definimos el StringVar para mostrar las Notas agregadas a las tarea                                                       *
TInicio=    StringVar   (value=(ws.cell(column=8,row=num-1)).value)                                                                 # Definimos el StringVar para mostrar la fecha de inicio de tareas                                                          *
TFin=       StringVar   (value=(ws.cell(column=9,row=num-1)).value)                                                                 # Definimos el StringVar para mostrar la fecha de finalizacion de la tarea                                                  *
TStat=      StringVar   ()                                                                                                          # Definimos esta variable para poder configurar y setear el Estado de la tarea                                              * 
#****************************************************************************************************************************************************************************************************************************************************************
print       ("TTarea es => " + str(TTarea.get()))
#****************************************************************************************************************************************************************************************************************************************************************
# Definimos Sub Rutina de Status para poder cambiar el color de Labels                                                                                                                                                                                          *
#****************************************************************************************************************************************************************************************************************************************************************
# Detalles:                                                                                                                                                                                                                                                     *
# En esta parte del codigo lo que hacemos el verificar el estado de las tareas y poder moficarlo para poder visualizar en ves de un numero una palabra.                                                                                                         *
#****************************************************************************************************************************************************************************************************************************************************************
if TStatus.get()=='0':                                                                                                              # Dependiendo del Nivel de Status es que vamos a pintar los Labels de las Tareas                                            *
    TStat.set('Abierto')                                                                                                            # En caso de que figure "0", la tarea esta "Abierto", por lo que se pinta de color Rojo                                     *
    Flbl='#FF0000'
elif TStatus.get()=='1':                                                                                                            # En caso de que figure "1", la tarea esta "Trabajando", por lo que se pinta de color Amarillo                              *
    TStat.set('Trabajando')
    Flbl='#FFFF00'
elif TStatus.get()=='2':                                                                                                            # En caso de que figure "2", la tarea esta "Listo", por lo que se pinta de color Verde                                      *
    TStat.set('Listo')
    Flbl='#00FF00'
else:
    Flbl='#0000FF'                                                                                                                  # En caso de que no figure nada, la tarea esta "Abierto", por lo que se pinta de color Azul                                 *
#****************************************************************************************************************************************************************************************************************************************************************
# Funciones del Programa a utilizar                                                                                                                                                                                                                             *
##****************************************************************************************************************************************************************************************************************************************************************                        
#   Detalle:                                                                                                                                                                                                                                                    *
#           Esta parte del codigo se utiliza para realizar cambios de las tareas de acuerdo al dia. Modificamos el dia aumentando uno y asi podemos navegar para poder analizar las tareas programadas                                                          *
#****************************************************************************************************************************************************************************************************************************************************************
def Siguiente():
    
    global var_t1,var_t2,var_t3,var_t4,var_t5,var_t6
    global t1,t2,t3,t4,t5,t6
    global s_t1,s_t2,s_t3,s_t4,s_t5,s_t6
    global num,Adia,semana                                                                                                                # Tomo las Variables Globales y le permito el uso dentro del subprograma                                                *
    pos=1
    Adia=Adia+1

    control=cdia.get()                                                                                                                  # Utilizamos esta variable para tomar el valor de "Cdia" y asi podemos manipular la misma                               *
    control=control+1                                                                                                                   # Aumento en 1 la variable para poder ver el dia siguiente                                                              *
    ddia.set(Dia + Adia)                                                                                                                 # Esta variable la utilizamos para poder mostrar en el Label de Dia                                                     *
    
    if int(semana) < 7:
        semana= int(semana) + 1
        print (int(semana))
    else:
        semana = 1
        print (int(semana))
    
    if (Mes%2)==0:                                                                                                                      # Verificamos cuantos dias tiene el mes. En caso de que el mes tenga 30 dias                                          *
        cdia.set(control)
        print ("control "+str(control)+" Mes => "+ str(Mes))                                                                            # Mostramos mensaje, mostrando el valor de "control"                                                                    *
    else:                                                                                                                               # En caso controrio, 
        print("Mes 31 dias"+str(control)+" Mes => "+ str(Mes))
        cdia.set(control)
        print ("control "+str(control))
    
    if control<31:
        cdia.set(control)
        print ("control "+str(control))
    else:
        control=1
        cdia.set(control)
        print (control) 

    Act_Tree() 
#****************************************************************************************************************************************************************************************************************************************************************    
def Anterior():
       
    global var_t1,var_t2,var_t3,var_t4,var_t5,var_t6
    global t1,t2,t3,t4,t5,t6
    global num                                                                                                                          # Tomo las Variables Globales y le permito el uso dentro del subprograma                                                    *
    pos=int(0)
    
    wb=openpyxl.load_workbook(str(lineas[0])[:-1])                                                                                      # Corresponde a la ruta en caso de tenerlo en la pc en forma local                                                          *
    ws=wb["Calendario"]
    num=1
    control=cdia.get()
    cdia.set(control-num)
    ddia.set(control-num)
    
    if (cdia.get()<1) & (Smes==30):
        cdia.set(control+Smes)
    elif (cdia.get()<1) & (Smes==31):
        cdia.set(control+29)  
        
    if int(semana) > 1:
        semana= int(semana) - 1
        print (int(semana))
    else:
        semana = 7
        print (int(semana))

    Act_Tree()
#****************************************************************************************************************************************************************************************************************************************************************
def Actualizar():
    
    global var_t1,var_t2,var_t3,var_t4,var_t5,var_t6,var_t7
    global t1,t2,t3,t4,t5,t6
    global num
   
    sct.delete('1.0', END)                                                                                                          # Limpieamos el Scrolltext, para poder visualizar las tareas correctamente armadas
    sct.insert(INSERT,"Se Realiza Mantenimiento de: ")                                                                              # Incertamos la primer linea del Scroll text
                   
    wb=openpyxl.load_workbook(str(lineas[0])[:-1])                                                                                  # Abrimos el Excel de Base para
    ws=wb["Calendario"]                                                 
    
#****************************************************************************************************************************************************************************************************************************************************************
#   Validacion Visual de Celdas en Excel
#****************************************************************************************************************************************************************************************************************************************************************
    if var_t1.get()==1:                                                                                                                     # Validacion de tarea de renglon T1, en caso de que sea Verdadera, se pinta la tarea de Verde en Excel               *        
        t1=ws.cell(row=int(semana),column=2)
        ws[t1.coordinate].fill=PatternFill(patternType='solid',fgColor='FF0000')
        lbl_t1=tk.Label(mydia,textvariable=res_t1,background="#00FF00",font=("Arial",12),width=50).place(x=10,y=Renglon*3)
        sct.insert(INSERT,"\n \t* "+t1.value)
        c_t1.value="1"
        ws[c_t1.coordinate]="1"
    else:                                                                                                                                   # En caso de que sea Falso, se pinta la tarea de Rojo                                                             *
        t1=ws.cell(row=int(semana),column=2)
        ws[t1.coordinate].fill=PatternFill(patternType='solid', fgColor='FF0000')
        lbl_t1=tk.Label(mydia,textvariable=res_t1,background='#FF0000',font=("Arial",12),width=50).place(x=10,y=Renglon*3)
        c_t1.value="0" 
        ws[c_t1.coordinate]="0"
        
    if var_t2.get()==1:                                                                                                                     # Validacion del resto de las tareas de renglones T2 a T8                                                           * 
        t2=ws.cell(row=int(semana),column=3)
        ws[t2.coordinate].fill=PatternFill(patternType='solid', fgColor='00FF00')
        lbl_t2=tk.Label(mydia,textvariable=res_t2,background="#00FF00",font=("Arial",12),width=50).place(x=10,y=Renglon*4)
        sct.insert(INSERT,"\n \t* "+t2.value)
        c_t2.value="1"
        ws[c_t2.coordinate]="1"
    else:                                                                                                                                   
        t2=ws.cell(row=int(semana),column=3)
        ws[t2.coordinate].fill=PatternFill(patternType='solid', fgColor='FF0000')
        lbl_t2=tk.Label(mydia,textvariable=res_t2,background="#FF0000",font=("Arial",12),width=50).place(x=10,y=Renglon*4)
        c_t2.value="0"
        ws[c_t2.coordinate]="1"
         
    if var_t3.get()==1:                                                                                                                 
        t3=ws.cell(row=int(semana),column=4)
        ws[t3.coordinate].fill=PatternFill(patternType='solid', fgColor='00FF00')
        lbl_t3=tk.Label(mydia,textvariable=res_t3,background="#00FF00",font=("Arial",12),width=50).place(x=10,y=Renglon*5)
        sct.insert(INSERT,"\n \t* "+t3.value)
        c_t3.value="1"
        ws[c_t3.coordinate]="1"
    else:                                                                                                                               
        t3=ws.cell(row=int(semana),column=4)
        ws[t3.coordinate].fill=PatternFill(patternType='solid', fgColor='FF0000')
        lbl_t3=tk.Label(mydia,textvariable=res_t3,background="#FF0000",font=("Arial",12),width=50).place(x=10,y=Renglon*5)
        c_t3.value="0"
        ws[c_t3.coordinate]="0"
        
    if var_t4.get()==1:                                                                                                                 
        t4=ws.cell(row=int(semana),column=5)
        ws[t4.coordinate].fill=PatternFill(patternType='solid', fgColor='00FF00')
        lbl_t4=tk.Label(mydia,textvariable=res_t4,background="#00FF00",font=("Arial",12),width=50).place(x=10,y=Renglon*6)
        sct.insert(INSERT,"\n \t* "+t4.value)
        c_t4.value="1"
        ws[c_t4.coordinate]="1"
    else:                                                                                                                               
        t4=ws.cell(row=int(semana),column=5)
        ws[t4.coordinate].fill=PatternFill(patternType='solid', fgColor='FF0000')
        lbl_t4=tk.Label(mydia,textvariable=res_t4,background="#FF0000",font=("Arial",12),width=50).place(x=10,y=Renglon*6)
        c_t4.value="0"
        ws[c_t4.coordinate]="0"
    
    if var_t5.get()==1:                                                                                                                 
        t5 =ws.cell(row=cdia.get(),column=2)
        ws[t5.coordinate].fill=PatternFill(patternType='solid', fgColor='00FF00')
        lbl_t5=tk.Label(mydia,textvariable=res_t5,background="#00FF00",font=("Arial",12),width=50).place(x=10,y=Renglon*7)
        sct.insert(INSERT,"\n \t* "+t5.value)
        c_t5.value="1"
        ws[c_t5.coordinate]="1"
    else:                                                                                                                               
        t5=ws.cell(row=cdia.get(),column=2)
        ws[t5.coordinate].fill=PatternFill(patternType='solid', fgColor='FF0000')
        lbl_t5=tk.Label(mydia,textvariable=res_t5,background="#FF0000",font=("Arial",12),width=50).place(x=10,y=Renglon*7)
        c_t5.value="0"
        ws[c_t5.coordinate]="0"

    if var_t6.get()==1:                                                                                                                 
        t6=ws.cell(row=cdia.get(),column=3)
        ws[t6.coordinate].fill=PatternFill(patternType='solid', fgColor='00FF00')
        lbl_t6=tk.Label(mydia,textvariable=res_t6,background="#00FF00",font=("Arial",12),width=50).place(x=10,y=Renglon*8)
        sct.insert(INSERT,"\n \t* "+t6.value)
        c_t6.value="1"
        ws[c_t6.coordinate]="1"
    else:                                                                                                                               
        t6=ws.cell(row=cdia.get(),column=3)
        ws[t6.coordinate].fill=PatternFill(patternType='solid', fgColor='FF0000')
        lbl_t6=tk.Label(mydia,textvariable=res_t6,background="#FF0000",font=("Arial",12),width=50).place(x=10,y=Renglon*8)
        c_t6.value="0"
        ws[c_t6.coordinate]="0"
#****************************************************************************************************************************************************************************************************************************************************************
    res_t1.set      (t1.value)
    res_t2.set      (t2.value)
    res_t3.set      (t3.value)
    res_t4.set      (t4.value)
    res_t5.set      (t5.value)
    res_t6.set      (t6.value)  
    res_t7.set      (t7.value) 
#****************************************************************************************************************************************************************************************************************************************************************
    l_t1.set        (ws.cell(row=(int(semana)),column=2))
    l_t2.set        (ws.cell(row=(int(semana)),column=3))
    l_t3.set        (ws.cell(row=(int(semana)),column=4))
    l_t4.set        (ws.cell(row=(int(semana)),column=5))
    l_t5.set        (ws.cell(row=(cdia.get()),column=2))
    l_t6.set        (ws.cell(row=(cdia.get()),column=3))
    l_t7.set        (ws.cell(row=(cdia.get()),column=4))
#**************************************************************************************************************************************************************************************************************************************************************** 
    ws.cell         (row=(int(semana)+1),column=2).fill=      PatternFill(patternType='solid', fgColor='FFFFFF')
    ws.cell         (row=(int(semana)+1),column=3).fill=      PatternFill(patternType='solid', fgColor='FFFFFF')
    ws.cell         (row=(int(semana)+1),column=4).fill=      PatternFill(patternType='solid', fgColor='FFFFFF')
    ws.cell         (row=(int(semana)+1),column=5).fill=      PatternFill(patternType='solid', fgColor='FFFFFF')
    ws.cell         (row=(cdia.get()+1),column=2).fill=       PatternFill(patternType='solid', fgColor='FFFFFF')
    ws.cell         (row=(cdia.get()+1),column=3).fill=       PatternFill(patternType='solid', fgColor='FFFFFF')
    ws.cell         (row=(cdia.get()+1),column=4).fill=       PatternFill(patternType='solid', fgColor='FFFFFF')
    ws.cell         (row=(cdia.get()+1),column=5).fill=       PatternFill(patternType='solid', fgColor='FFFFFF')
#****************************************************************************************************************************************************************************************************************************************************************
    mydia.update()  
    wb.save(str(lineas[0])[:-1])
#****************************************************************************************************************************************************************************************************************************************************************       
def Nuevo():
    global num,dispositivo
    
    dispositivo=ws.cell(row=num,column=1)
    while dispositivo.value!=None:
        num=num+1
        dispositivo=ws.cell(row=num-1,column=1)

    print('Nuevo Numero =>'+str(num))
    ws.cell     (column=1,row=num).value=       TFecha  .get()
    ws.cell     (column=2,row=num).value=       TTarea  .get()
    ws.cell     (column=3,row=num).value=       TEqui   .get()
    ws.cell     (column=4,row=num).value=       TSector .get()
    ws.cell     (column=5,row=num).value=       TStatus .get()
    ws.cell     (column=6,row=num).value=       TAsig   .get()
    ws.cell     (column=7,row=num).value=       TTarea  .get()
    ws.cell     (column=8,row=num).value=       TInicio .get()
    ws.cell     (column=9,row=num).value=       TFin    .get()
    ws.cell     (column=10,row=num).value=      TNotas  .get()
    
    wb.save(str(lineas[0])[:-1])
#****************************************************************************************************************************************************************************************************************************************************************
def NSig():
    global num

    wb=openpyxl.load_workbook(str(lineas[0])[:-1])                                                                                      # Corresponde a la ruta en caso de tenerlo en la pc en forma local                                                      *
    ws=wb["Tareas"]

    num=num+1
#****************************************************************************************************************************************************************************************************************************************************************    
#   Variable    Tipo        Codificación                                                                                                                                                                                                                        *
#****************************************************************************************************************************************************************************************************************************************************************
    TFecha  =   StringVar   (value=(ws.cell(column=1,row=num)).value)       #*
    TTarea  =   StringVar   (value=(ws.cell(column=2,row=num)).value)       #*   
    TEqui   =   StringVar   (value=(ws.cell(column=3,row=num)).value)       #*
    TSector =   StringVar   (value=(ws.cell(column=4,row=num)).value)       #*
    TStatus =   StringVar   (value=(ws.cell(column=5,row=num)).value)       #*
    TAsig   =   StringVar   (value=(ws.cell(column=6,row=num)).value)       #*
    TNotas  =   StringVar   (value=(ws.cell(column=7,row=num)).value)       #*
    TInicio =   StringVar   (value=(ws.cell(column=8,row=num)).value)       #*
    TFin    =   StringVar   (value=(ws.cell(column=9,row=num)).value)       #*
#****************************************************************************************************************************************************************************************************************************************************************    
    if TStatus.get()=='0':
        TStat.set('Abierto')
        Flbl='#FF0000'
    elif TStatus.get()=='1':
        TStat.set('Trabajando')
        Flbl='#FFFF00'
    elif TStatus.get()=='2':
        TStat.set('Listo')
        Flbl='#00FF00'
    else:
        Flbl='#0000FF' 
#****************************************************************************************************************************************************************************************************************************************************************    
#   Nombre          Tipo        Codificación                                                                                     Ubicacion                                                                                                                      *
#****************************************************************************************************************************************************************************************************************************************************************
    lbl_TFecha=     tk.Label    (mydia,text="Fecha"         ,background=Flbl,font=("Arial",12),width=20)                            .place(x=10,    y=Renglon*11)
    lbl_TEqui=      tk.Label    (mydia,text="Equipo"        ,background=Flbl,font=("Arial",12),width=20)                            .place(x=10,    y=Renglon*12)
    lbl_TSector=    tk.Label    (mydia,text="Sector"        ,background=Flbl,font=("Arial",12),width=20)                            .place(x=10,    y=Renglon*13)
    lbl_TAsig=      tk.Label    (mydia,text="Asignado"      ,background=Flbl,font=("Arial",12),width=20)                            .place(x=10,    y=Renglon*14)
    lbl_TTarea=     tk.Label    (mydia,text="Tarea"         ,background=Flbl,font=("Arial",12),width=20)                            .place(x=10,    y=Renglon*15)
    lbl_TInicio=    tk.Label    (mydia,text="Inicio"        ,background=Flbl,font=("Arial",12),width=20)                            .place(x=550,   y=Renglon*11)
    lbl_TFin=       tk.Label    (mydia,text="Fin"           ,background=Flbl,font=("Arial",12),width=20)                            .place(x=550,   y=Renglon*13)
    lbl_TStatus=    tk.Label    (mydia,text="Status"        ,background=Flbl,font=("Arial",12),width=20,justify='center')           .place(x=550,   y=Renglon*15)
    lbl_TNotas=     tk.Label    (mydia,text="Notas"         ,background=Flbl,font=("Arial",12),width=57)                            .place(x=10,    y=Renglon*17)

    Ent_TFecha=     tk.Entry    (mydia,textvariable=TFecha  ,background=Fondo,font=("Arial",12),width=30)                           .place(x=250,   y=Renglon*11)
    Ent_TEqui=      tk.Entry    (mydia,textvariable=TEqui   ,background=Fondo,font=("Arial",12),width=30)                           .place(x=250,   y=Renglon*12)
    Ent_TSector=    tk.Entry    (mydia,textvariable=TSector ,background=Fondo,font=("Arial",12),width=30)                           .place(x=250,   y=Renglon*13)
    Ent_TAsig=      tk.Entry    (mydia,textvariable=TAsig   ,background=Fondo,font=("Arial",12),width=30)                           .place(x=250,   y=Renglon*14)
    Ent_TTarea=     tk.Entry    (mydia,textvariable=TTarea  ,background=Fondo,font=("Arial",12),width=57)                           .place(x=10,    y=Renglon*16)
    Ent_TInicio=    tk.Entry    (mydia,textvariable=TInicio ,background=Fondo,font=("Arial",12),width=20)                           .place(x=550,   y=Renglon*12)
    Ent_TFin=       tk.Entry    (mydia,textvariable=TFin    ,background=Fondo,font=("Arial",12),width=20)                           .place(x=550,   y=Renglon*14)
    Ent_TStatus=    tk.Entry    (mydia,textvariable=TStatus ,background=Fondo,font=("Arial",12),width=20,justify='center')          .place(x=550,   y=Renglon*16)
    Ent_TNotas=     tk.Entry    (mydia,textvariable=TNotas  ,background=Fondo,font=("Arial",12),width=57)                           .place(x=10,    y=Renglon*18)   # Imprimo el nombre de Dia en Ingles. Se puede generar el alternativo por medio de un switch*
#****************************************************************************************************************************************************************************************************************************************************************
def NAnt():
    global num
    wb=openpyxl.load_workbook(str(lineas[0])[:-1])                                                                                      # Corresponde a la ruta en caso de tenerlo en la pc en forma local                                                          *
    ws=wb["Tareas"]
    
    num=num-1
    print ("Valor del Renglon" + str(num))
#****************************************************************************************************************************************************************************************************************************************************************    
#   Variable    Tipo        Codificación                                                                                                                                                                                                                        *
#****************************************************************************************************************************************************************************************************************************************************************
    TFecha=     StringVar   (value=(ws.cell(column=1,row=num)).value)
    TTarea=     StringVar   (value=(ws.cell(column=2,row=num)).value)
    TEqui=      StringVar   (value=(ws.cell(column=3,row=num)).value)
    TSector=    StringVar   (value=(ws.cell(column=4,row=num)).value)
    TStatus=    StringVar   (value=(ws.cell(column=5,row=num)).value)
    TAsig=      StringVar   (value=(ws.cell(column=6,row=num)).value)
    TNotas=     StringVar   (value=(ws.cell(column=7,row=num)).value)
    TInicio=    StringVar   (value=(ws.cell(column=8,row=num)).value)
    TFin=       StringVar   (value=(ws.cell(column=9,row=num)).value)
#****************************************************************************************************************************************************************************************************************************************************************        
    print('TTarea => '+TTarea.get())
    print('TStatus =>'+TStatus.get())
    print('Numero =>'+ str(num))
    
    if TStatus.get()=='0':
        TStat.set('Abierto')
        Flbl='#FF0000'
    elif TStatus.get()=='1':
        TStat.set('Trabajando')
        Flbl='#FFFF00'
    elif TStatus.get()=='2':
        TStat.set('Listo')
        Flbl='#00FF00'
    else:
        Flbl='#0000FF'            

#****************************************************************************************************************************************************************************************************************************************************************    
#   Nombre          Tipo        Codificación                                                                                     Ubicacion                                                                                                                      *
#****************************************************************************************************************************************************************************************************************************************************************
    lbl_TFecha=     tk.Label    (mydia,text="Fecha"     ,background=Flbl,font=("Arial",12),width=20)                                .place(x=10,    y=Renglon*11)
    lbl_TEqui=      tk.Label    (mydia,text="Equipo"    ,background=Flbl,font=("Arial",12),width=20)                                .place(x=10,    y=Renglon*12)
    lbl_TSector=    tk.Label    (mydia,text="Sector"    ,background=Flbl,font=("Arial",12),width=20)                                .place(x=10,    y=Renglon*13)
    lbl_TStatus=    tk.Label    (mydia,text="Status"    ,background=Flbl,font=("Arial",12),width=20,justify='center')               .place(x=550,   y=Renglon*15)
    lbl_TAsig=      tk.Label    (mydia,text="Asignado"  ,background=Flbl,font=("Arial",12),width=20)                                .place(x=10,    y=Renglon*14)
    lbl_TTarea=     tk.Label    (mydia,text="Tarea"     ,background=Flbl,font=("Arial",12),width=20)                                .place(x=10,    y=Renglon*15)
    lbl_TInicio=    tk.Label    (mydia,text="Inicio"    ,background=Flbl,font=("Arial",12),width=20)                                .place(x=550,   y=Renglon*11)
    lbl_TFin=       tk.Label    (mydia,text="Fin"       ,background=Flbl,font=("Arial",12),width=20)                                .place(x=550,   y=Renglon*13)
    lbl_TNotas=     tk.Label    (mydia,text="Notas"     ,background=Flbl,font=("Arial",12),width=57)                                .place(x=10,    y=Renglon*17)

    Ent_TFecha=     tk.Entry    (mydia,textvariable=TFecha  ,background=Fondo,font=("Arial",12),width=30)                           .place(x=250,   y=Renglon*11)
    Ent_TEqui=      tk.Entry    (mydia,textvariable=TEqui   ,background=Fondo,font=("Arial",12),width=30)                           .place(x=250,   y=Renglon*12)
    Ent_TSector=    tk.Entry    (mydia,textvariable=TSector ,background=Fondo,font=("Arial",12),width=30)                           .place(x=250,   y=Renglon*13)
    Ent_TStatus=    tk.Entry    (mydia,textvariable=TStatus ,background=Fondo,font=("Arial",12),width=20,justify='center')          .place(x=550,   y=Renglon*16)
    Ent_TAsig=      tk.Entry    (mydia,textvariable=TAsig   ,background=Fondo,font=("Arial",12),width=30)                           .place(x=250,   y=Renglon*14)
    Ent_TTarea=     tk.Entry    (mydia,textvariable=TTarea  ,background=Fondo,font=("Arial",12),width=57)                           .place(x=10,    y=Renglon*16)
    Ent_TInicio=    tk.Entry    (mydia,textvariable=TInicio ,background=Fondo,font=("Arial",12),width=20)                           .place(x=550,   y=Renglon*12)
    Ent_TFin=       tk.Entry    (mydia,textvariable=TFin    ,background=Fondo,font=("Arial",12),width=20)                           .place(x=550,   y=Renglon*14)
    Ent_TNotas=     tk.Entry    (mydia,textvariable=TNotas  ,background=Fondo,font=("Arial",12),width=57)                           .place(x=10,    y=Renglon*18)   # Imprimo el nombre de Dia en Ingles. Se puede generar el alternativo por medio de un switch*
#****************************************************************************************************************************************************************************************************************************************************************
def Nguardar():
    global num,dispositivo
    global TFecha,TTarea
    
    dispositivo=ws.cell(row=num,column=1)
    while dispositivo.value!=None:
        num=num+1
        dispositivo=ws.cell(row=num-1,column=1)
    
    VFecha=     str(Ent_TFecha)
    VTarea=     str(Ent_TTarea)
    VEqui=      str()
    VSector=    str()
    VStatus=    str()
    VAsig=      str()
    VTarea=     str()
    VInicio=    str()
    VFin=       str()
    Vnotas=     str()

    #TFecha      .set    (str(VFecha.value))
    #TTarea      .set    (value=Ent_TTarea)
    #TEqui       .set    (value=Ent_TEqui)
    #TSector     .set    (value=Ent_TSector)
    #TStatus     .set    (value=Ent_TStatus)
    #TAsig       .set    (value=Ent_TAsig)
    #TTarea      .set    (value=Ent_TTarea)
    #TInicio     .set    (value=Ent_TInicio)
    #TFin        .set    (value=Ent_TFin)
    #TNotas      .set    (value=Ent_TNotas)

    print('Nuevo Numero =>'+str(num)+"valor de entrada "+str(VFecha))
    
    ws.cell     (column=1,row=num).value=       VFecha
    #ws.cell     (column=2,row=num).value=       TTarea  .get()
    #ws.cell     (column=3,row=num).value=       TEqui   .get()
    #ws.cell     (column=4,row=num).value=       TSector .get()
    #ws.cell     (column=5,row=num).value=       TStatus .get()
    #ws.cell     (column=6,row=num).value=       TAsig   .get()
    #ws.cell     (column=7,row=num).value=       TTarea  .get()
    #ws.cell     (column=8,row=num).value=       TInicio .get()
    #ws.cell     (column=9,row=num).value=       TFin    .get()
    #ws.cell     (column=10,row=num).value=      TNotas  .get()
    
    print('Nuevo Numero =>'+str(num)+" Fecha "+str(VFecha))
    
    num=num+1
    messagebox.showinfo(title="Guardado",message="Se guardo información")
    
    wb.save(str(lineas[0])[:-1])
    mydia.update()
#****************************************************************************************************************************************************************************************************************************************************************
def Act_Tree():
 
    wb=openpyxl.load_workbook(str(lineas[0])[:-1])                                                                                      # Corresponde a la ruta en caso de tenerlo en la pc en forma local                                                      *
    ws=wb["Calendario"]
#****************************************************************************************************************************************************************************************************************************************************************
#   Nombre    / Valor                                                                                                                                                                                                                                           *
#****************************************************************************************************************************************************************************************************************************************************************
    
    t1=         ws.cell(row=int(semana),column=2)                                                                                           # Tomo el valor de la celda para la variables                                                                               *     
    t2=         ws.cell(row=int(semana),column=3)
    t3=         ws.cell(row=int(semana),column=4)
    t4=         ws.cell(row=int(semana),column=5)
    t5=         ws.cell(row=cdia.get(),column=2)
    t6=         ws.cell(row=cdia.get(),column=3)
    t7=         ws.cell(row=cdia.get(),column=4)
    
    res_t1.set  (t1.value)
    res_t2.set  (t2.value)
    res_t3.set  (t3.value)
    res_t4.set  (t4.value)
    res_t5.set  (t5.value)
    res_t6.set  (t6.value)
    res_t7.set  (t7.value)
      
      
    s_t1=       ws.cell(row=int(semana),column=7)
    s_t2=       ws.cell(row=int(semana),column=8)
    s_t3=       ws.cell(row=int(semana),column=9)
    s_t4=       ws.cell(row=cdia.get(),column=2)
    s_t5=       ws.cell(row=cdia.get(),column=3)
    s_t6=       ws.cell(row=cdia.get(),column=4)
    s_t7=       ws.cell(row=cdia.get(),column=5)        
    
    for i in tree.get_children():
         tree.delete(i)
    mydia.update()
    pos=int(0)
    
    if s_t1.value !="-":
        tree.insert('', tk.END, text=s_t1.value, iid=pos, open=False,)
        pos=pos+1     
    if s_t2.value !="-":
        tree.insert('', tk.END, text=s_t2.value, iid=pos, open=False)
        pos=pos+1
    if s_t3.value !="-":
        tree.insert('', tk.END, text=s_t3.value, iid=pos, open=False)
        pos=pos+1
    if s_t4.value !="-":
        tree.insert('', tk.END, text=s_t4.value, iid=pos, open=False)
        pos=pos+1
    if s_t5.value!="-":
        tree.insert('', tk.END, text=s_t5.value, iid=pos, open=False)
        pos=pos+1
#    if s_t6.value != "-":
#        tree.insert('', tk.END, text=s_t6.value, iid=pos, open=False)
#        pos=pos+1
#    if s_t7.value != "-":
#        tree.insert('', tk.END, text=s_t7.value, iid=pos, open=False)
#        pos=pos+1   
           
    mydia.update() 
#****************************************************************************************************************************************************************************************************************************************************************
def Conectar():
    
    global tree,sct
    
    current_item = tree.focus()    
    data = tree.item(current_item,option='text') 
    server=data
    Mensaje=(str(Hini)+" => "+ user[9:]+"Se conecta al servidor " + str(server) + "\n")
    import WLog
    servidor='cmd /k "mstsc -v ' + str(server) + ':4489'
    sct.insert(INSERT,"\n\t* "+ server)
    time.sleep(1)
    os.system(servidor)                                                                                                                 #\n Se realiza mantenimiento remoto en:                                                                                 *
    return                                                                                                                                                                                                                                                     #*
#**************************************************************************************************************************************************************************************************************************************************************** 
def Imprimir():                                                                                                                     # Parte del codigo para imprimir los resultados en la pantalla                                                              * 
    global dato,num,dispositivo
    
    wb=openpyxl.load_workbook(str(lineas[0])[:-1])
    ws=wb["Calendario"]
    
    sct.insert(INSERT,"\n Se imprime informe ")                                                                                     # Se agrega el Texto al Bloque de notas que se encuetra en la pantalla                                                      *
    
    Mensaje=(str(Hini)+" => "+user[9:30]+ " " + sct.get("1.0",END) + "\n"+(100*"*")+"\n")
    import WLog
#****************************************************************************************************************************************************************************************************************************************************************
def Ayuda():                                                                    #           Boton para abrir el .txt donde detalla asistencias y detalles de la aplicación. Se define los detalles de la aplicación y funcion de la misma                       *
    import Ayuda
#****************************************************************************************************************************************************************************************************************************************************************    
def Salir():
    
    Mensaje=(str(Hini)+" => Se cierra aplicación "+user[9:]+"\n"+(100*"*")+"\n")
    import WLog
    Destroy()
#****************************************************************************************************************************************************************************************************************************************************************    
def Destroy():
    mydia.destroy()
#****************************************************************************************************************************************************************************************************************************************************************
# Definimos Los Labels a Mostrar 
#****************************************************************************************************************************************************************************************************************************************************************
#****************************************************************************************************************************************************************************************************************************************************************    
#   Nombre          Tipo        Codificación                                                                                     Ubicacion                                                                                                                      *
#****************************************************************************************************************************************************************************************************************************************************************
lbl_TFecha=     tk.Label    (mydia,text="Fecha"         ,background=Flbl,font=("Arial",12),width=20)                            .place(x=10,    y=Renglon*11)
lbl_TEqui=      tk.Label    (mydia,text="Equipo"        ,background=Flbl,font=("Arial",12),width=20)                            .place(x=10,    y=Renglon*12)
lbl_TSector=    tk.Label    (mydia,text="Sector"        ,background=Flbl,font=("Arial",12),width=20)                            .place(x=10,    y=Renglon*13)
lbl_TAsig=      tk.Label    (mydia,text="Asignado"      ,background=Flbl,font=("Arial",12),width=20)                            .place(x=10,    y=Renglon*14)
lbl_TTarea=     tk.Label    (mydia,text="Tarea"         ,background=Flbl,font=("Arial",12),width=20)                            .place(x=10,    y=Renglon*15)
lbl_TInicio=    tk.Label    (mydia,text="Inicio"        ,background=Flbl,font=("Arial",12),width=20)                            .place(x=550,   y=Renglon*11)
lbl_TFin=       tk.Label    (mydia,text="Fin"           ,background=Flbl,font=("Arial",12),width=20)                            .place(x=550,   y=Renglon*13)
lbl_TStatus=    tk.Label    (mydia,text="Status"        ,background=Flbl,font=("Arial",12),width=20,justify='center')           .place(x=550,   y=Renglon*15)
lbl_TNotas=     tk.Label    (mydia,text="Notas"         ,background=Flbl,font=("Arial",12),width=57)                            .place(x=10,    y=Renglon*17)

Ent_TFecha=     tk.Entry    (mydia,textvariable=TFecha  ,background=Fondo,font=("Arial",12),width=30)                           .place(x=250,   y=Renglon*11)
Ent_TEqui=      tk.Entry    (mydia,textvariable=TEqui   ,background=Fondo,font=("Arial",12),width=30)                           .place(x=250,   y=Renglon*12)
Ent_TSector=    tk.Entry    (mydia,textvariable=TSector ,background=Fondo,font=("Arial",12),width=30)                           .place(x=250,   y=Renglon*13)
Ent_TAsig=      tk.Entry    (mydia,textvariable=TAsig   ,background=Fondo,font=("Arial",12),width=30)                           .place(x=250,   y=Renglon*14)
Ent_TTarea=     tk.Entry    (mydia,textvariable=TTarea  ,background=Fondo,font=("Arial",12),width=57)                           .place(x=10,    y=Renglon*16)
Ent_TInicio=    tk.Entry    (mydia,textvariable=TInicio ,background=Fondo,font=("Arial",12),width=20)                           .place(x=550,   y=Renglon*12)
Ent_TFin=       tk.Entry    (mydia,textvariable=TFin    ,background=Fondo,font=("Arial",12),width=20)                           .place(x=550,   y=Renglon*14)
Ent_TStatus=    tk.Entry    (mydia,textvariable=TStatus ,background=Fondo,font=("Arial",12),width=20,justify='center')          .place(x=550,   y=Renglon*16)
Ent_TNotas=     tk.Entry    (mydia,textvariable=TNotas  ,background=Fondo,font=("Arial",12),width=57)                           .place(x=10,    y=Renglon*18)       # Imprimo el nombre de Dia en Ingles. Se puede generar el alternativo por medio de un switch*
#****************************************************************************************************************************************************************************************************************************************************************    
lbl_res_celda=  tk.Label    (mydia,     text="Dia",                 background=Fondo,       font=("Arial",12))                    .place(x=10,    y=Renglon)
lbl_dia=        tk.Label    (mydia,     textvariable=ddia,          background=Fondo,       font=("Arial",12),      width=25)     .place(x=50,    y=Renglon)
lbl_t1=         tk.Label    (mydia,     textvariable=res_t1,        background=Fondo,       font=("Arial",12),      width=50)     .place(x=10,    y=Renglon*3)
lbl_t2=         tk.Label    (mydia,     textvariable=res_t2,        background=Fondo,       font=("Arial",12),      width=50)     .place(x=10,    y=Renglon*4)
lbl_t3=         tk.Label    (mydia,     textvariable=res_t3,        background=Fondo,       font=("Arial",12),      width=50)     .place(x=10,    y=Renglon*5)
lbl_t4=         tk.Label    (mydia,     textvariable=res_t4,        background=Fondo,       font=("Arial",12),      width=50)     .place(x=10,    y=Renglon*6)
lbl_t5=         tk.Label    (mydia,     textvariable=res_t5,        background=Fondo,       font=("Arial",12),      width=50)     .place(x=10,    y=Renglon*7)
lbl_t6=         tk.Label    (mydia,     textvariable=res_t6,        background=Fondo,       font=("Arial",12),      width=50)     .place(x=10,    y=Renglon*8)
lbl_t7=         tk.Label    (mydia,     textvariable=res_t7,        background=Fondo,       font=("Arial",12),      width=50)     .place(x=10,    y=Renglon*9)
#****************************************************************************************************************************************************************************************************************************************************************
mydia.update()
#****************************************************************************************************************************************************************************************************************************************************************
# Definimos los botoes a ver en la ventana inicial                                                                                                                                                                                                              *
#****************************************************************************************************************************************************************************************************************************************************************
boton=          tk.Button(mydia,text="Buscar",      activebackground="#ABCDEF",     background="#838B8B",       command=Actualizar,     width=155,      image=myimg)            .place(x=1150,y=60)                  # Creo Boton "planilla" para procesar las plantillas Requeridas para informe*
salir=          tk.Button(mydia,text="Salir",       activebackground="#BABABA",                                 command=Salir,          width=7,        justify='center')       .place(x=1250,y=610)                 # Creo un Boton para cerrar la aplicación                                   *
bhelp=          tk.Button(mydia,text="Ayuda",                                       background="#838383",       command=Ayuda,          width=5)                                .place(x=1200,y=610)                 # Creo el Boton de "Ayuda" para mostrar el Txt correspondiente              *
bsiguiente=     tk.Button(mydia,text="Siguiente",   activebackground="#ABABAB",     background="#838383",       command=Siguiente,      width=10,       state='active')         .place(x=1230,y=150)
banterios=      tk.Button(mydia,text="Previo",      activebackground="#ABABAB",     background="#838383",       command=Anterior,       width=10,       state='active')         .place(x=1150,y=150)
bconectar=      tk.Button(mydia,text="Conectar",    activebackground="#ABABAB",     background="#838383",       command=Conectar,       width=22,       state='active')         .place(x=1150,y=180)
bimprimir=      tk.Button(mydia,text="Imprimir",    activebackground="#ABABAB",     background="#838383",       command=Imprimir,       width=22,       state='active')         .place(x=1150,y=250)
#****************************************************************************************************************************************************************************************************************************************************************                     
bnuevo=         tk.Button(mydia,text="Nueva Linea",activebackground="#ABABAB",background="#838383",command=Nuevo,width=15,state='active')       .place(x=10,y=640)
bn_siguente=    tk.Button(mydia,text="Siguiente Linea",activebackground="#ABABAB",background="#838383",command=NSig,width=15,state='active')    .place(x=240,y=640)
bn_anterior=    tk.Button(mydia,text="Linea Anterior",activebackground="#ABABAB",background="#838383",command=NAnt,width=15,state='active')     .place(x=360,y=640)
bn_guardar=     tk.Button(mydia,text="Guardar",activebackground="#ABABAB",background="#838383",command=Nguardar,width=6,state='active')         .place(x=475,y=640)
#****************************************************************************************************************************************************************************************************************************************************************                     
#****************************************************************************************************************************************************************************************************************************************************************
log.close()
mydia.mainloop()
#****************************************************************************************************************************************************************************************************************************************************************