#****************************************************************************************************************************************************************************************************************************************************************
# Nombre: Modificar.Py                                                                                                                                                                                                                                           *       
# Descripcion: Aplicación de busqueda de equipos en listado de equipos                                                                                                                                                                                          *
# #**************************************************************************************************************************************************************************************************************************************************************
import cmd
from dis import dis
from operator import eq
import os as os
from sre_parse import State
import string
import sys as sys
from sys import *
from os import replace, system as system
import datetime,time
#****************************************************************************************************************************************************************************************************************************************************************
import tkinter as tk
from tkinter import *
from tkinter import Entry, Grid, Image, StringVar, Text, Variable, messagebox, ttk, scrolledtext, simpledialog, tix, font, commondialog
from tkinter.ttk import Progressbar, setup_master,Sizegrip,Entry
from tkinter.tix import STATUS, LabelEntry,LabelFrame,Meter, ButtonBox,PhotoImage,ComboBox
from tkinter.constants import *
from tokenize import Double, String
from turtle import color, delay, title
from typing import Any
#****************************************************************************************************************************************************************************************************************************************************************
import openpyxl
from openpyxl import Workbook,load_workbook
#****************************************************************************************************************************************************************************************************************************************************************
from buscador import num
#****************************************************************************************************************************************************************************************************************************************************************
#                                                                                                               Declaración de Variables                                                                                                                        *
#****************************************************************************************************************************************************************************************************************************************************************
new_user=os.environ['USERPROFILE']                                                                                                  # Identifico el "UserProfile" de la pc para poder encontrar las carpetas instaladas en el Setup.py                          *
user=new_user.replace("\\","/")                                                                                                     # Reemplazo "\" por "/" dado a que no reconocen la ruta en Python                                                           *

#path=open(new_user + '/Desktop/Codigo/Paths.txt','r')                                                                              # Busco el archivo donde se encuentran las rutas preestablecidas para encontrar los archivos. (Las cuales se pueden         *
path=open(user+'/Desktop/Soft/Path.txt')                                                                              
lineas=path.readlines()                                                                                                             # modificar en caso que asi se quiera)                                                                                      *

log=open(user  + (str(lineas[1])[:-1]),mode="a")                                                                                   # Ruta de Archivo donde se encuentran los Logs de eventos                                                                   *
#****************************************************************************************************************************************************************************************************************************************************************
dia=datetime.datetime.today().day                                                                                                   # Variable de Dia                                                                                                           *
mes=datetime.datetime.today().month                                                                                                 # Variable de Mes                                                                                                           *
Hini=datetime.datetime.now()                                                                                                        # Horario de inicio                                                                                                         *
Mensaje=str()                                                                                                                       # Variable que utilizo para confeccionar los mensajes a mostrar en la aplicación                                            *
servidor=str()
#****************************************************************************************************************************************************************************************************************************************************************
#****************************************************************************************************************************************************************************************************************************************************************
Texto=str()                                                                                                                         # Variable para poder escribir e imprimir en pantalla o CMD                                                                 *
sdia=int()                                                                                                                          # Variable para buscar "Día"                                                                                                *
smes=int()                                                                                                                          # Variable para buscar "Mes"                                                                                                *
Step=int()                                                                                                                          # Variable para marcar el paso a ejecutar en el programa                                                                    *
Fondo='#5B5B5B'                                                                                                                     # Color Amarillo de Prosegur'#FFCC01'

#****************************************************************************************************************************************************************************************************************************************************************
#   Variables para uso del programa                                                                                                                                                                                                                             *
#****************************************************************************************************************************************************************************************************************************************************************
var=int()                                                                                                                           # Variable de uso general para pruebas                                                                                      *
dato=str()                                                                                                                          # Variable para poder obtener datos y poder comparar                                                                        *
num=int(1)                                                                                                                          # Variable para poder navegar por el listado Excel                                                                          *
limite=3000                                                                                                                         # Limite de renglones para buscar                                                                                           *
Mensaje=str()                                                                                                                       # Variable para poder mostrar mensaje                                                                                       *    
#****************************************************************************************************************************************************************************************************************************************************************
#****************************************************************************************************************************************************************************************************************************************************************
# Defina Interfaz Grafica                                                                                                                                                                                                                                       *
#****************************************************************************************************************************************************************************************************************************************************************
#****************************************************************************************************************************************************************************************************************************************************************
myapp = Tk()                                                                                                                        # MyApp es el nombre de la planilla a visualizar                                                                            *                                                                                                             
myapp.title("Modificar")                                                                                                             # Defino el titulo del programa                                                                                             *
h=500
w=1000
myapp.minsize(w,h)                                                                                                                  # Defino el Tamaño minimo e inicial de plantilla                                                                            *
myapp.frame()
myapp.resizable(False,False)
myapp.config(background=Fondo)                                                                                                      # Defino color de Fondo de pantalla "Myapp"                                                                                 *
#****************************************************************************************************************************************************************************************************************************************************************
#   Lineas de comando para poder centrar la ventana en la mitad de la pantalla (En teoria)
#****************************************************************************************************************************************************************************************************************************************************************
screen_width = myapp.winfo_screenwidth()
screen_height = myapp.winfo_screenheight()
x_cordinate = int((screen_width/2) - (w/2))
y_cordinate = int((screen_height/2) - (h/2)-25)
myapp.geometry("{}x{}+{}+{}".format(w,h, x_cordinate, y_cordinate))
#****************************************************************************************************************************************************************************************************************************************************************
mylogo=PhotoImage(file=(user  + (str(lineas[4]))[:-1]))                                                                             # Variable para imagen del Logo del Icono a usar                                                                            *
lbl_lable=ttk.Label(myapp,image=mylogo,border=0).place(x=-10,y=-10)                                                                 # Se define en Lable para poder utilizar "mylogo" como fondo de ventana.                                                    *
#****************************************************************************************************************************************************************************************************************************************************************
renglon=30                                                                                                                          # Defino el valor a utilizar commo Renglones para poder las etiquetas sobre la ventana del soft                             *
posicion=150
ancho=37
bh=20
#****************************************************************************************************************************************************************************************************************************************************************
# Definimos Variables para respuestas                                                                                                                                                                                                                           *
#****************************************************************************************************************************************************************************************************************************************************************
res_dispositivo=StringVar()                                                                                                         # Variable para respuesta de Etiqueta de "Dispositivo"                                                                      *
res_ubi=StringVar()                                                                                                                 # Variable para respuesta de Etiqueta de "Ubicacion"                                                                        *
res_equipo=StringVar()                                                                                                              # Variable para respuesta de Etiqueta de "Equipo"                                                                           *
res_nombre=StringVar()                                                                                                              # Variable para respuesta de Etiqueta de "Nombre"                                                                           *
res_marca=StringVar()                                                                                                               # Variable para respuesta de Etiqueta de "Marca"                                                                            *
res_modelo=StringVar()
res_ip=StringVar()                                                                                                                  # Variable para respuesta de Etiqueta de "Ip"                                                                               *
res_serial=StringVar()                                                                                                              # Variable para respuesta de Etiqueta de "serial"                                                                           *
res_usuario=StringVar()                                                                                                             # Variable para respuesta de Etiqueta de "Usuario"                                                                          *
res_password=StringVar()                                                                                                            # Variable para respuesta de Etiqueta de "Contraseña"                                                                       *
res_server=StringVar()
#****************************************************************************************************************************************************************************************************************************************************************
mod_ubi=StringVar()                                                                                                                 # Variable para Modificar de Etiqueta de "Ubicacion"                                                                        *
mod_equipo=str()                                                                                                              # Variable para Modificar de Etiqueta de "Equipo"                                                                           *
mod_nombre=StringVar()                                                                                                              # Variable para Modificar de Etiqueta de "Nombre"                                                                           *
mod_marca=StringVar()                                                                                                               # Variable para Modificar de Etiqueta de "Marca"                                                                            *
mod_ip=StringVar()                                                                                                                  # Variable para Modificar de Etiqueta de "Ip"                                                                               *
mod_serial=StringVar()                                                                                                              # Variable para Modificar de Etiqueta de "serial"                                                                           *
mod_usuario=StringVar()                                                                                                             # Variable para Modificar de Etiqueta de "Usuario"                                                                          *
mod_password=StringVar()                                                                                                            # Variable para Modificar de Etiqueta de "Contraseña"                                                                       *
mod_server=StringVar()
#****************************************************************************************************************************************************************************************************************************************************************
simb_disp=IntVar()                                                                                                                  # Validador de Check Boton de "Dispositivo"                                                                      *
simb_ubi=IntVar()                                                                                                                   # Validador de Check Boton de "Ubicacion"                                                                        *
simb_equipo=IntVar()                                                                                                                # Validador de Check Boton de "Equipo"                                                                           *
simb_nombre=IntVar()                                                                                                                # Validador de Check Boton de "Nombre"                                                                           *
simb_marca=IntVar()                                                                                                                 # Validador de Check Boton de "Marca"                                                                            *
simb_modelo=IntVar()                                                                                                                # Valodador de Check Boton de "Modelo"
simb_ip=IntVar()                                                                                                                    # Validador de Check Boton de "Ip"                                                                               *
simb_serial=IntVar()                                                                                                                # Validador de Check Boton de "serial"                                                                           *
simb_user=IntVar()                                                                                                                  # Validador de Check Boton de "Usuario"                                                                          *
simb_pass=IntVar()                                                                                                                  # Validador de Check Boton de "Contraseña"                                                                       *
simb_server=IntVar()
#****************************************************************************************************************************************************************************************************************************************************************

#****************************************************************************************************************************************************************************************************************************************************************
lbl_Ubicacion=tk.Label(myapp,text="Ubicacion =>",background=Fondo,font=("Arial",12),width=15).place(x=10,y=renglon*2)
lbl_res_ubicacion=tk.Label(myapp,textvariable=res_ubi,font=("Arial",12),background="#AAAAAA",width=ancho).place(x=posicion,y=renglon*2)
lbl_equipo=tk.Label(myapp,text="Equipo =>",background=Fondo,font=("Arial",12),width=15).place(x=10,y=renglon*3)
lbl_res_equipo=tk.Label(myapp,textvariable=res_equipo,font=("Arial",12),background="#AAAAAA",width=ancho).place(x=posicion,y=renglon*3)
lbl_nombre=tk.Label(master=myapp,text="Nombre =>",background=Fondo,font=("Arial",12),width=15).place(x=10,y=renglon*4)
lbl_res_nombre=tk.Label(myapp,textvariable=res_nombre,font=("Arial",12),background="#AAAAAA",width=ancho).place(x=posicion,y=renglon*4)
lbl_marca=tk.Label(myapp,text="Marca =>",background=Fondo,font=("Arial",12),width=15).place(x=10,y=renglon*5)
lbl_res_marca=tk.Label(myapp,textvariable=res_marca,font=("Arial",12),background="#AAAAAA",width=ancho).place(x=posicion,y=renglon*5)
lbl_marca=tk.Label(myapp,text="Modelo =>",background=Fondo,font=("Arial",12),width=15).place(x=10,y=renglon*6)
lbl_res_marca=tk.Label(myapp,textvariable=res_modelo,font=("Arial",12),background="#AAAAAA",width=ancho).place(x=posicion,y=renglon*6)
lbl_Serial=tk.Label(myapp,text="Serial =>",background=Fondo,font=("Arial",12),width=15).place(x=10,y=renglon*7)
lbl_res_Serial=tk.Label(myapp,textvariable=res_serial,font=("Arial",12),background="#AAAAAA",width=ancho).place(x=posicion,y=renglon*7)
lbl_ip=tk.Label(myapp,text="IP =>",background=Fondo,font=("Arial",12),width=15).place(x=10,y=renglon*8)
lbl_res_Ip=tk.Label(myapp,textvariable=res_ip,font=("Arial",12),background="#AAAAAA",width=ancho).place(x=posicion,y=renglon*8)
lbl_usuario=tk.Label(master=myapp,text="Usuario =>",background=Fondo,font=("Arial",12),width=15).place(x=10,y=renglon*9)
lbl_res_Usuario=tk.Label(myapp,textvariable=res_usuario,font=("Arial",12),background="#AAAAAA",width=ancho).place(x=posicion,y=renglon*9)
lbl_Password=tk.Label(myapp,text="Password =>",background=Fondo,font=("Arial",12),width=15).place(x=10,y=renglon*10)
lbl_res_Password=tk.Label(myapp,textvariable=res_password,font=("Arial",12),background="#AAAAAA",width=ancho).place(x=posicion,y=renglon*10)
lbl_res_Server=tk.Label(myapp,textvariable=res_server,font=("Arial",12),background="#AAAAAA",width=ancho).place(x=posicion,y=renglon*11)
#****************************************************************************************************************************************************************************************************************************************************************                                                                                                                
#****************************************************************************************************************************************************************************************************************************************************************                                                                                                                
modifi=450
#****************************************************************************************************************************************************************************************************************************************************************
E_mod_ubicacion=ttk.Entry(myapp,textvariable=res_ubi,font=("Arial",12),foreground="#FF0000",width=ancho,justify='center').place(x=posicion+modifi,y=renglon*2)                                            # Defino el renglon de entrada de datos para comparar y Se define posición del renglon de "Entrada"              *
E_mod_equipo=ttk.Entry(myapp,textvariable=res_equipo,font=("Arial",12),background="#AAAAAA",width=ancho,justify='center').place(x=posicion+modifi,y=renglon*3)
E_mod_nombre=ttk.Entry(myapp,textvariable=res_nombre,font=("Arial",12),background="#AAAAAA",width=ancho,justify='center').place(x=posicion+modifi,y=renglon*4)
E_mod_marca=ttk.Entry(myapp,textvariable=res_marca,font=("Arial",12),background="#AAAAAA",width=ancho,justify='center').place(x=posicion+modifi,y=renglon*5)
E_mod_modelo=ttk.Entry(myapp,textvariable=res_modelo,font=("Arial",12),background="#AAAAAA",width=ancho,justify='center').place(x=posicion+modifi,y=renglon*6)
E_mod_Serial=ttk.Entry(myapp,textvariable=res_serial,font=("Arial",12),background="#AAAAAA",width=ancho,justify='center').place(x=posicion+modifi,y=renglon*7)
E_mod_Ip=ttk.Entry(myapp,textvariable=res_ip,font=("Arial",12),background="#AAAAAA",width=ancho,justify='center').place(x=posicion+modifi,y=renglon*8)
E_mod_Usuario=ttk.Entry(myapp,textvariable=res_usuario,font=("Arial",12),background="#AAAAAA",width=ancho,justify='center').place(x=posicion+modifi,y=renglon*9)
E_mod_Password=ttk.Entry(myapp,textvariable=res_password,font=("Arial",12),background="#AAAAAA",width=ancho,justify='center').place(x=posicion+modifi,y=renglon*10)
E_mod_Server=ttk.Entry(myapp,textvariable=res_server,font=("Arial",12),background="#AAAAAA",width=ancho,justify='center').place(x=posicion+modifi,y=renglon*11)
#****************************************************************************************************************************************************************************************************************************************************************

#****************************************************************************************************************************************************************************************************************************************************************
# Funciones del Programa a utilizar                                                                                                                                                                                                                             *
#****************************************************************************************************************************************************************************************************************************************************************   
def Ayuda():
    import Ayuda
#****************************************************************************************************************************************************************************************************************************************************************                        
def Siguiente():
    
    global num,dispositivo                                                                                                          # Tomo las Variables Globales y le permito el uso dentro del subprograma                                                    *
    
    num+=1                                                                                                                          # Aumento la variable "num"                                                                                                 *

    wb=openpyxl.load_workbook(str(lineas[0])[:-1])                                                                                  # Genero el vinculo y abro el Excel que esta definido dentro del String dentro del Path                                     *     
    ws=wb["CCTV"]                                                                                                                   # Abro la Hoja (ws) CCTV para poder trabajar con esos datos
    dispositivo=ws.cell(row=num,column=4)                                                                                           # Defino la variable para poder compara con la variable "Entrada" teniendo en cuenta los parametros "num" como renglones    *
                                                                                                                                    #y como columna una sola definitiva para poder buscar solamente por nombre de equipo                                        *   
   
    if dispositivo.value=="":                                                                                                       # Verifico que la variable no este vacia                                                                                    *
        messagebox.showerror(message="Final de listado sin datos similares")                                                        # En caso de que se encuentra vacia, mostramos mensaje indicando que llego al final del listado de equipos                  *
    else:
        Imprimir()                                                                                                                  # En caso de que tengamos información dentro de ese renglon, pasamos los datos al sub Imprimi()                             *
#****************************************************************************************************************************************************************************************************************************************************************    
def Anterior():
    
    global num, dispositivo                                                                                                          # Tomo las Variables Globales y le permito el uso dentro del subprograma                                                    *
    
    if num>0:
        num-=1
        Imprimir()
    else:
        tk.messagebox.showerror(message="No se encuentra valor menor")    
        
#****************************************************************************************************************************************************************************************************************************************************************
def m_ubi():
    global num,limite
    global dispositivo
        
    wb=openpyxl.load_workbook(str(lineas[0])[:-1])                                                                                  # Corresponde a la ruta en caso de tenerlo en la pc en forma local                                                          *
    ws=wb["CCTV"]
    #ubicacion=ws.cell(row=num,column=3)
    #ubicacion.value=res_ubi.get().upper()
    #wb.save(str(lineas[0])[:-1])

def Modific():
    
    global num,limite
    global dispositivo
        
    wb=openpyxl.load_workbook(str(lineas[0])[:-1])                                                                                  # Corresponde a la ruta en caso de tenerlo en la pc en forma local                                                          *
    ws=wb["CCTV"]
    
    dispositivo=ws.cell(row=num,column=4)
    nombre=ws.cell(row=num,column=4)
    equipo=ws.cell(row=num,column=2)
    ubicacion=ws.cell(row=num,column=3)
    marca=ws.cell(row=num,column=6)
    modelo=ws.cell(row=num,column=7)
    ip=ws.cell(row=num,column=11)
    serial=ws.cell(row=num,column=8)
    usuario=ws.cell(row=num,column=12)
    password=ws.cell(row=num,column=13)
    server=ws.cell(row=num,column=14)    
    
    equipo.value=res_equipo.get().upper()
    nombre.value=res_nombre.get().upper()
    ubicacion.value=res_ubi.get().upper()
    marca.value=res_marca.get().upper()
    modelo.value=res_modelo.get().upper()
    ip.value=res_ip.get().upper()
    serial.value=res_serial.get().upper()
    usuario.value=res_usuario.get().upper()
    password.value=res_password.get().upper()
    server.value=res_server.get().upper()
    
    wb.save(str(lineas[0])[:-1])
    
#****************************************************************************************************************************************************************************************************************************************************************        
def Imprimir():    
    
    global dato,num,dispositivo
    
    wb=openpyxl.load_workbook(str(lineas[0])[:-1])
    ws=wb["CCTV"]
    
    log=open(user  + (str(lineas[1])[:-1]),mode="a")
    log.write(str(Hini)+" => "+user[9:30] + " esta buscado el dato " +dato+" que esta es el registro numero => "+str(num)+"\n")
    log.close()
    
    dispositivo=ws.cell(row=num,column=4)
    nombre=ws.cell(row=num,column=4)
    equipo=ws.cell(row=num,column=2)
    ubicacion=ws.cell(row=num,column=3)
    marca=ws.cell(row=num,column=6)
    modelo=ws.cell(row=num,column=7)
    ip=ws.cell(row=num,column=11)
    serial=ws.cell(row=num,column=8)
    usuario=ws.cell(row=num,column=12)
    password=ws.cell(row=num,column=13)
    server=ws.cell(row=num,column=14)    
           
    res_nombre.set(nombre.value)
    res_equipo.set(equipo.value)
    res_ubi.set(str(ubicacion.value))
    res_marca.set(marca.value)
    res_modelo.set(modelo.value)
    res_ip.set(ip.value)
    res_serial.set(serial.value)
    res_usuario.set(usuario.value)
    res_password.set(password.value)
    res_server.set(server.value)
    

def Conectar():

    wb=openpyxl.load_workbook(str(lineas[0])[:-1])
    ws=wb["CCTV"]
    server=ws.cell(row=num,column=14) 
    servidor='cmd /k "mstsc -v ' + server.value + ':4489'
    print(servidor)
    myapp.destroy()
    os.system(servidor)
    
    return 
    
def Salir():
    
    log=open(user  + (str(lineas[1])[:-1]),mode="a")
    log.write(str(Hini)+" => Se cierra aplicación "+user[9:]+"\n")
    log.close()
    Destroy()
    
def Destroy():
    myapp.destroy()

#****************************************************************************************************************************************************************************************************************************************************************
# Definimos los botoes a ver en la ventana inicial                                                                                                                                                                                                              *
#***************************************************************************************************************************************************************************************************************************************************************
altura=400
boton=tk.Button(myapp,text="Modificar",activebackground="#ABABAB",background="#838383",command=Modific,width=11,state='active').place(x=680,y=altura)                                     # Creo Boton "planilla" para procesar las plantillas Requeridas para informe.   *
salir=tk.Button(myapp,text="Salir",activebackground="#BABABA",command=Salir,justify='center',width=23).place(x=790,y=altura+50)                                                       # Creo un Boton para cerrar la aplicación                                       *
bsiguiente=tk.Button(myapp,text="Siguiente",activebackground="#ABABAB",background="#838383",command=Siguiente,width=11,state='active').place(x=780,y=altura)
banterios=tk.Button(myapp,text="Previo",activebackground="#ABABAB",background="#838383",command=Anterior,width=11,state='active').place(x=880,y=altura)
#bmodificar=tk.Button(myapp,text="Modificar",activebackground="#ABABAB",background="#838383",command=Modificar,width=25,state='active').place(x=780,y=180)
#bhelp=tk.Button(myapp,text="Ayuda",background="#838383",command=Ayuda,width=5,).place(x=15,y=650)                                                                               # Creo el Boton de "Ayuda" para mostrar el Txt correspondiente                  *
#bconectar=tk.Button(myapp,text="Conectar",activebackground="#ABABAB",background="#838383",command=Conectar,width=25,state='active').place(x=780,y=500)
#****************************************************************************************************************************************************************************************************************************************************************                     
imagen=PhotoImage(file=(user+"/Desktop/Soft/flecha.png"))
#****************************************************************************************************************************************************************************************************************************************************************
#chk_simb_disp=tk.Button(myapp,activebackground=Fondo,background="#AAAAAA",image=imagen,height=bh).place(x=posicion+350,y=renglon*2)                                            # Boton para habilitar el cambio de Texto en Listado (Dispositivo)
chk_simb_ubi=tk.Button(myapp,activebackground=Fondo,background="#AAAAAA",image=imagen,height=bh,command=m_ubi).place(x=posicion+350,y=renglon*2)                                # Boton para habilitar el cambio de Texto en Listado (Ubicación)
chk_simb_equipo=tk.Button(myapp,activebackground=Fondo,background="#AAAAAA",image=imagen,height=bh,textvariable=simb_equipo).place(x=posicion+350,y=renglon*3)                  # Boton para habilitar el cambio de Texto en Listado (Equipo)
chk_simb_nombre=tk.Button(myapp,activebackground=Fondo,background="#AAAAAA",image=imagen,height=bh,textvariable=simb_nombre).place(x=posicion+350,y=renglon*4)                  # Boton para habilitar el cambio de Texto en Listado (Nombre)
chk_simb_marca=tk.Button(myapp,activebackground=Fondo,background="#AAAAAA",image=imagen,height=bh,textvariable=simb_marca).place(x=posicion+350,y=renglon*5)                    # Boton para habilitar el cambio de Texto en Listado (Marca)
chk_simb_modelo=tk.Button(myapp,activebackground=Fondo,background="#AAAAAA",image=imagen,height=bh,textvariable=simb_modelo).place(x=posicion+350,y=renglon*6)                  # Boton para habilitar el cambio de Texto en Listado (Modelo)
chk_simb_ip=tk.Button(myapp,activebackground=Fondo,background="#AAAAAA",image=imagen,height=bh,textvariable=simb_ip).place(x=posicion+350,y=renglon*7)                          # Boton para habilitar el cambio de Texto en Listado (Ip)
chk_simb_serial=tk.Button(myapp,activebackground=Fondo,background="#AAAAAA",image=imagen,height=bh,textvariable=simb_serial).place(x=posicion+350,y=renglon*8)                  # Boton para habilitar el cambio de Texto en Listado (Serial)
chk_simb_user=tk.Button(myapp,activebackground=Fondo,background="#AAAAAA",image=imagen,height=bh,textvariable=simb_user).place(x=posicion+350,y=renglon*9)                     # Boton para habilitar el cambio de Texto en Listado (Usuario)
chk_simb_pass=tk.Button(myapp,activebackground=Fondo,background="#AAAAAA",image=imagen,height=bh,textvariable=simb_pass).place(x=posicion+350,y=renglon*10)                     # Boton para habilitar el cambio de Texto en Listado (Password)
chk_simb_server=tk.Button(myapp,activebackground=Fondo,background="#AAAAAA",image=imagen,textvariable=simb_server).place(x=posicion+350,y=renglon*11)                           # Boton para habilitar el cambio de Texto en Listado (Servidor)                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                          
#****************************************************************************************************************************************************************************************************************************************************************
log.close()
myapp.mainloop()
#****************************************************************************************************************************************************************************************************************************************************************


