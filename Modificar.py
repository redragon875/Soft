#****************************************************************************************************************************************************************************************************************************************************************
# Nombre: Modificar.Py                                                                                                                                                                                                                                          *       
# Descripcion: Aplicación para modificar información de equipos en listado de equipos                                                                                                                                                                           *
#               El funcionamiento es simple, de la busqueda de un equipo en la app "Busqueda.py* salta la info y podemos modificar los mismos parametros sin tener que buscar nuevamente.
# #**************************************************************************************************************************************************************************************************************************************************************
import os as os
import sys as sys
from sys import *
from os import replace, system as system
import datetime,time
#****************************************************************************************************************************************************************************************************************************************************************
import tkinter as tk
from tkinter import *
from tkinter import Entry, Grid, Image, StringVar, Text, Variable, messagebox, ttk, scrolledtext, simpledialog, tix, font, commondialog
from tkinter.ttk import Progressbar, setup_master,Sizegrip,Entry
from tkinter.tix import STATUS, LabelEntry,LabelFrame,Meter, ButtonBox,PhotoImage,ComboBox
from tkinter.constants import *
#****************************************************************************************************************************************************************************************************************************************************************
import openpyxl
from openpyxl import Workbook,load_workbook
#****************************************************************************************************************************************************************************************************************************************************************
from Variables import *
from buscador import *
#****************************************************************************************************************************************************************************************************************************************************************
#****************************************************************************************************************************************************************************************************************************************************************
# Defina Interfaz Grafica                                                                                                                                                                                                                                       *
#****************************************************************************************************************************************************************************************************************************************************************
#****************************************************************************************************************************************************************************************************************************************************************
Mymod = Tk()                                                                                                                        # MyMod es el nombre de la planilla a visualizar                                                                            *                                                                                                             
Mymod.title("Modificar")                                                                                                            # Defino el titulo del programa                                                                                             *
H=500                                                                                                                               # Defininos la Altura de la ventana                                                                                         *
W=1000                                                                                                                              # Definimos el ancho de la ventana                                                                                          *
Mymod.minsize(W,H)                                                                                                                  # Defino el Tamaño minimo e inicial de plantilla                                                                            *
Mymod.frame()
Mymod.resizable(False,False)                                                                                                        
Mymod.config(background=Fondo)                                                                                                      # Defino color de Fondo de pantalla "Myapp"                                                                                 *
#****************************************************************************************************************************************************************************************************************************************************************
#   Lineas de comando para poder centrar la ventana en la mitad de la pantalla (En teoria)
#****************************************************************************************************************************************************************************************************************************************************************
screen_width =  Mymod.winfo_screenwidth()
screen_height = Mymod.winfo_screenheight()
x_cordinate =   int((screen_width/2) - (W/2))
y_cordinate =   int((screen_height/2) - (H/2)-25)
Mymod.geometry("{}x{}+{}+{}".format(W,H, x_cordinate, y_cordinate))
#****************************************************************************************************************************************************************************************************************************************************************
MyFondo=    PhotoImage(file=(user   + (str(lineas[4]))[:-1]))                                                                       # Variable para imagen del Logo del Icono a usar                                                                            * 
lbl_lable=  ttk.Label(Mymod,image=MyFondo,border=0).place(x=-10,y=-10)                                                              # Se define en Lable para poder utilizar "mylogo" como fondo de ventana.                                                    *
#****************************************************************************************************************************************************************************************************************************************************************
# Definimos Variables para respuestas                                                                                                                                                                                                                           *
#****************************************************************************************************************************************************************************************************************************************************************
res_dispositivo=    StringVar()                                                                                                     # Variable para respuesta de Etiqueta de "Dispositivo"                                                                      *
res_ubi=            StringVar()                                                                                                     # Variable para respuesta de Etiqueta de "Ubicacion"                                                                        *
res_equipo=         StringVar()                                                                                                     # Variable para respuesta de Etiqueta de "Equipo"                                                                           *
res_nombre=         StringVar()                                                                                                     # Variable para respuesta de Etiqueta de "Nombre"                                                                           *
res_marca=          StringVar()                                                                                                     # Variable para respuesta de Etiqueta de "Marca"                                                                            *
res_modelo=         StringVar()                                                                                                     # Variable para respuesta de Etiqueta de "Modelo"                                                                           *
res_ip=             StringVar()                                                                                                     # Variable para respuesta de Etiqueta de "Ip"                                                                               *
res_serial=         StringVar()                                                                                                     # Variable para respuesta de Etiqueta de "serial"                                                                           *
res_usuario=        StringVar()                                                                                                     # Variable para respuesta de Etiqueta de "Usuario"                                                                          *
res_password=       StringVar()                                                                                                     # Variable para respuesta de Etiqueta de "Contraseña"                                                                       *
res_server=         StringVar()                                                                                                     # Variable para respuesta de Etiqueta de "Servidores"                                                                       *
#****************************************************************************************************************************************************************************************************************************************************************
mod_ubi=            StringVar()                                                                                                     # Variable para Modificar de Etiqueta de "Ubicacion"                                                                        *
mod_equipo=         StringVar()                                                                                                     # Variable para Modificar de Etiqueta de "Equipo"                                                                           *
mod_nombre=         StringVar()                                                                                                     # Variable para Modificar de Etiqueta de "Nombre"                                                                           *
mod_marca=          StringVar()                                                                                                     # Variable para Modificar de Etiqueta de "Marca"                                                                            *
mod_modelo=         StringVar()                                                                                                     # Variable para Modificar de Etiqueta de "Modelo"
mod_ip=             StringVar()                                                                                                     # Variable para Modificar de Etiqueta de "Ip"                                                                               *
mod_serial=         StringVar()                                                                                                     # Variable para Modificar de Etiqueta de "serial"                                                                           *
mod_usuario=        StringVar()                                                                                                     # Variable para Modificar de Etiqueta de "Usuario"                                                                          *
mod_password=       StringVar()                                                                                                     # Variable para Modificar de Etiqueta de "Contraseña"                                                                       *
mod_server=         StringVar()                                                                                                     # Variable para Modificar de Etiqueta de "Servidores"                                                                       *
#****************************************************************************************************************************************************************************************************************************************************************
simb_disp=          IntVar()                                                                                                        # Validador de Check Boton de "Dispositivo"                                                                                 *
simb_ubi=           IntVar()                                                                                                        # Validador de Check Boton de "Ubicacion"                                                                                   *
simb_equipo=        IntVar()                                                                                                        # Validador de Check Boton de "Equipo"                                                                                      *
simb_nombre=        IntVar()                                                                                                        # Validador de Check Boton de "Nombre"                                                                                      *
simb_marca=         IntVar()                                                                                                        # Validador de Check Boton de "Marca"                                                                                       *
simb_modelo=        IntVar()                                                                                                        # Valodador de Check Boton de "Modelo"                                                                                      *
simb_ip=            IntVar()                                                                                                        # Validador de Check Boton de "Ip"                                                                                          *
simb_serial=        IntVar()                                                                                                        # Validador de Check Boton de "serial"                                                                                      *
simb_user=          IntVar()                                                                                                        # Validador de Check Boton de "Usuario"                                                                                     *
simb_pass=          IntVar()                                                                                                        # Validador de Check Boton de "Contraseña"                                                                                  *
simb_server=        IntVar()                                                                                                        # Validador de Check Boton de "Servidor"                                                                                    *
#****************************************************************************************************************************************************************************************************************************************************************
# Labels para mostrar información                                                                                                                                                                                                                               *    
#****************************************************************************************************************************************************************************************************************************************************************
lbl_Ubicacion=      tk.Label(Mymod      ,text="Ubicacion =>"          ,font=Fuente        ,background=Fondo     ,width=15)      .place(x=10,y=Renglon*2)
lbl_equipo=         tk.Label(Mymod      ,text="Equipo =>"             ,font=Fuente        ,background=Fondo     ,width=15)      .place(x=10,y=Renglon*3)
lbl_nombre=         tk.Label(Mymod      ,text="Nombre =>"             ,font=Fuente        ,background=Fondo     ,width=15)      .place(x=10,y=Renglon*4)
lbl_marca=          tk.Label(Mymod      ,text="Marca =>"              ,font=Fuente        ,background=Fondo     ,width=15)      .place(x=10,y=Renglon*5)
lbl_modelo=         tk.Label(Mymod      ,text="Modelo =>"             ,font=Fuente        ,background=Fondo     ,width=15)      .place(x=10,y=Renglon*6)
lbl_Serial=         tk.Label(Mymod      ,text="Serial =>"             ,font=Fuente        ,background=Fondo     ,width=15)      .place(x=10,y=Renglon*7)
lbl_ip=             tk.Label(Mymod      ,text="IP =>"                 ,font=Fuente        ,background=Fondo     ,width=15)      .place(x=10,y=Renglon*8)
lbl_usuario=        tk.Label(Mymod      ,text="Usuario =>"            ,font=Fuente        ,background=Fondo     ,width=15)      .place(x=10,y=Renglon*9)
lbl_Password=       tk.Label(Mymod      ,text="Password =>"           ,font=Fuente        ,background=Fondo     ,width=15)      .place(x=10,y=Renglon*10)

#****************************************************************************************************************************************************************************************************************************************************************                                                                                                                
lbl_res_ubicacion=  tk.Label(Mymod      ,textvariable=res_ubi         ,font=Fuente      ,background="#AAAAAA"   ,width=Ancho)    .place(x=Posicion,y=Renglon*2)
lbl_res_equipo=     tk.Label(Mymod      ,textvariable=res_equipo      ,font=Fuente      ,background="#AAAAAA"   ,width=Ancho)    .place(x=Posicion,y=Renglon*3)
lbl_res_nombre=     tk.Label(Mymod      ,textvariable=res_nombre      ,font=Fuente      ,background="#AAAAAA"   ,width=Ancho)    .place(x=Posicion,y=Renglon*4)
lbl_res_marca=      tk.Label(Mymod      ,textvariable=res_marca       ,font=Fuente      ,background="#AAAAAA"   ,width=Ancho)    .place(x=Posicion,y=Renglon*5)
lbl_res_Modelo=     tk.Label(Mymod      ,textvariable=res_modelo      ,font=Fuente      ,background="#AAAAAA"   ,width=Ancho)    .place(x=Posicion,y=Renglon*6)
lbl_res_Serial=     tk.Label(Mymod      ,textvariable=res_serial      ,font=Fuente      ,background="#AAAAAA"   ,width=Ancho)    .place(x=Posicion,y=Renglon*7)
lbl_res_Ip=         tk.Label(Mymod      ,textvariable=res_ip          ,font=Fuente      ,background="#AAAAAA"   ,width=Ancho)    .place(x=Posicion,y=Renglon*8)
lbl_res_Usuario=    tk.Label(Mymod      ,textvariable=res_usuario     ,font=Fuente      ,background="#AAAAAA"   ,width=Ancho)    .place(x=Posicion,y=Renglon*9)
lbl_res_Password=   tk.Label(Mymod      ,textvariable=res_password    ,font=Fuente      ,background="#AAAAAA"   ,width=Ancho)    .place(x=Posicion,y=Renglon*10)
lbl_res_Server=     tk.Label(Mymod      ,textvariable=res_server      ,font=Fuente      ,background="#AAAAAA"   ,width=Ancho)    .place(x=Posicion,y=Renglon*11)
#****************************************************************************************************************************************************************************************************************************************************************                                                                                                                
modifi=450
#****************************************************************************************************************************************************************************************************************************************************************
E_mod_ubicacion=    ttk.Entry(Mymod     ,textvariable=mod_ubi        ,font=Fuente       ,background="#AAAAAA"   ,width=Ancho    ,justify='center')   .place(x=Posicion+modifi,y=Renglon*2)         # Defino el renglon de entrada de datos para comparar y      *
E_mod_equipo=       ttk.Entry(Mymod     ,textvariable=mod_equipo     ,font=Fuente       ,background="#AAAAAA"   ,width=Ancho    ,justify='center')   .place(x=Posicion+modifi,y=Renglon*3)         #Se define posición del renglon de "Entrada"                 *
E_mod_nombre=       ttk.Entry(Mymod     ,textvariable=mod_nombre     ,font=Fuente       ,background="#AAAAAA"   ,width=Ancho    ,justify='center')   .place(x=Posicion+modifi,y=Renglon*4)
E_mod_marca=        ttk.Entry(Mymod     ,textvariable=mod_marca      ,font=Fuente       ,background="#AAAAAA"   ,width=Ancho    ,justify='center')   .place(x=Posicion+modifi,y=Renglon*5)
E_mod_modelo=       ttk.Entry(Mymod     ,textvariable=mod_modelo     ,font=Fuente       ,background="#AAAAAA"   ,width=Ancho    ,justify='center')   .place(x=Posicion+modifi,y=Renglon*6)
E_mod_Serial=       ttk.Entry(Mymod     ,textvariable=mod_serial     ,font=Fuente       ,background="#AAAAAA"   ,width=Ancho    ,justify='center')   .place(x=Posicion+modifi,y=Renglon*7)
E_mod_Ip=           ttk.Entry(Mymod     ,textvariable=mod_ip         ,font=Fuente       ,background="#AAAAAA"   ,width=Ancho    ,justify='center')   .place(x=Posicion+modifi,y=Renglon*8)
E_mod_Usuario=      ttk.Entry(Mymod     ,textvariable=mod_usuario    ,font=Fuente       ,background="#AAAAAA"   ,width=Ancho    ,justify='center')   .place(x=Posicion+modifi,y=Renglon*9)
E_mod_Password=     ttk.Entry(Mymod     ,textvariable=mod_password   ,font=Fuente       ,background="#AAAAAA"   ,width=Ancho    ,justify='center')   .place(x=Posicion+modifi,y=Renglon*10)
E_mod_Server=       ttk.Entry(Mymod     ,textvariable=mod_server     ,font=Fuente       ,background="#AAAAAA"   ,width=Ancho    ,justify='center')   .place(x=Posicion+modifi,y=Renglon*11)
#****************************************************************************************************************************************************************************************************************************************************************
def Var_local():
    dispositivo=        ws.cell(row=num,column=4)                                                                                  # Cargamos los valores de la planilla en as variables segun su posicionamiento. Dispositivo                                      *    
    nombre=             ws.cell(row=num,column=4)                                                                                  # Nombre del dispositivo en el listado                                                                                           *
    equipo=             ws.cell(row=num,column=2)                                                                                  # Tipo de Equipo                                                                                                                 *    
    ubicacion=          ws.cell(row=num,column=3)                                                                                  # Ubicación del Equipo                                                                                                           *
    marca=              ws.cell(row=num,column=6)                                                                                  # Marca del Equipo                                                                                                               *
    modelo=             ws.cell(row=num,column=7)                                                                                  # Modelo del Equipo                                                                                                              *
    serial=             ws.cell(row=num,column=8)                                                                                  # N° de serie                                                                                                                    *
    ip=                 ws.cell(row=num,column=11)                                                                                 # N° de IP                                                                                                                       *
    usuario=            ws.cell(row=num,column=12)                                                                                 # Usuario para ingresar a la configuración                                                                                       *
    password=           ws.cell(row=num,column=13)                                                                                 # Password de ingreso a equipo                                                                                                   *
    server=             ws.cell(row=num,column=14)                                                                                 # Servidor a la cual esta conectado                                                                                              *
    
    return (nombre,dispositivo,equipo,ubicacion,marca,modelo,serial,ip,usuario,password,server)
#****************************************************************************************************************************************************************************************************************************************************************
# Funciones del Programa a utilizar                                                                                                                                                                                                                             *
#****************************************************************************************************************************************************************************************************************************************************************   

#****************************************************************************************************************************************************************************************************************************************************************                        
def Siguiente():
    
    global num,dispositivo                                                                                                          # Tomo las Variables Globales y le permito el uso dentro del subprograma                                                    *
    
    num+=1                                                                                                                          # Aumento la variable "num"                                                                                                 *
    print (num)
    if dispositivo.value=="":                                                                                                       # Verifico que la variable no este vacia                                                                                    *
        messagebox.showerror(message="Final de listado sin datos similares")                                                        # En caso de que se encuentra vacia, mostramos mensaje indicando que llego al final del listado de equipos                  *
    else:
        Imprimir()                                                                                                                  # En caso de que tengamos información dentro de ese renglon, pasamos los datos al sub Imprimi()                             *
#****************************************************************************************************************************************************************************************************************************************************************    
def Anterior():
    
    global num, dispositivo                                                                                                         # Tomo las Variables Globales y le permito el uso dentro del subprograma                                                    *
    
    if num>0:                                                                                                                       # Verificamos que la posición sea mayor a 0 para por lo menos estar como minimo en el 1er renglon 
        num-=1
        Imprimir()
    else:
        messagebox.showerror(message="No se encuentra valor menor")    
#****************************************************************************************************************************************************************************************************************************************************************
def m_ubi():
     
    wb=openpyxl.load_workbook(str(lineas[0])[:-1])                                                                                  # Corresponde a la ruta en caso de tenerlo en la pc en forma local                                                          *
    ws=wb["CCTV"]
#****************************************************************************************************************************************************************************************************************************************************************
def Modific():                                                                                                                      # Parte del codigo donde se Modifica los valores del dispositivo y lo almacena en la planilla
    
    global num,limite
    global dispositivo

    if res_nombre.get() !=mod_nombre.get():
        E_mod_nombre=    ttk.Entry(Mymod     ,textvariable=mod_nombre        ,font=Fuente       ,foreground="#FF0000"   ,width=Ancho    ,justify='center')   .place(x=Posicion+modifi,y=Renglon*4)
        res_nombre.set  (mod_nombre.get().upper())
    else:
        E_mod_nombre=    ttk.Entry(Mymod     ,textvariable=mod_nombre        ,font=Fuente                               ,width=Ancho    ,justify='center')   .place(x=Posicion+modifi,y=Renglon*4)
         
    
    if res_equipo.get() !=mod_equipo.get():
        E_mod_equipo=    ttk.Entry(Mymod     ,textvariable=mod_equipo        ,font=Fuente       ,foreground="#FF0000"   ,width=Ancho    ,justify='center')   .place(x=Posicion+modifi,y=Renglon*3)
        res_equipo.set  (mod_equipo.get().upper())
    else:
        E_mod_equipo=    ttk.Entry(Mymod     ,textvariable=mod_equipo        ,font=Fuente                               ,width=Ancho    ,justify='center')   .place(x=Posicion+modifi,y=Renglon*3)
    
    if res_ubi.get()    != mod_ubi.get():
        E_mod_ubicacion=    ttk.Entry(Mymod     ,textvariable=mod_ubi        ,font=Fuente       ,foreground="#FF0000"   ,width=Ancho    ,justify='center')   .place(x=Posicion+modifi,y=Renglon*2)
        res_ubi.set     (mod_ubi.get().upper())
    else:
        E_mod_ubicacion=    ttk.Entry(Mymod     ,textvariable=mod_ubi        ,font=Fuente                               ,width=Ancho    ,justify='center')   .place(x=Posicion+modifi,y=Renglon*2)
        
    if res_marca.get()  !=mod_marca.get():
        E_mod_marca=        ttk.Entry(Mymod     ,textvariable=mod_marca      ,font=Fuente       ,foreground="#FF0000"   ,width=Ancho    ,justify='center')   .place(x=Posicion+modifi,y=Renglon*5)
        res_marca.set   (mod_marca.get().upper())
    else:
        E_mod_marca=        ttk.Entry(Mymod     ,textvariable=mod_marca      ,font=Fuente                               ,width=Ancho    ,justify='center')   .place(x=Posicion+modifi,y=Renglon*5)
        
    if res_modelo.get() !=mod_modelo.get():
        E_mod_modelo=       ttk.Entry(Mymod     ,textvariable=mod_modelo     ,font=Fuente       ,foreground="#FF0000"   ,width=Ancho    ,justify='center')   .place(x=Posicion+modifi,y=Renglon*6)
        res_modelo.set  (mod_modelo.get().upper())
    else:
        E_mod_modelo=       ttk.Entry(Mymod     ,textvariable=mod_modelo     ,font=Fuente                               ,width=Ancho    ,justify='center')   .place(x=Posicion+modifi,y=Renglon*6)
    
    if res_ip.get()     !=mod_ip.get():
        E_mod_Ip=           ttk.Entry(Mymod     ,textvariable=mod_ip         ,font=Fuente       ,foreground="#FF0000"   ,width=Ancho    ,justify='center')   .place(x=Posicion+modifi,y=Renglon*8)
        res_ip.set      (mod_ip.get().upper())
    else:
        E_mod_Ip=           ttk.Entry(Mymod     ,textvariable=mod_ip         ,font=Fuente                               ,width=Ancho    ,justify='center')   .place(x=Posicion+modifi,y=Renglon*8)
        
    if res_serial.get() !=mod_serial.get():
        E_mod_Serial=       ttk.Entry(Mymod     ,textvariable=mod_serial     ,font=Fuente       ,foreground="#FF0000"   ,width=Ancho    ,justify='center')   .place(x=Posicion+modifi,y=Renglon*7)
        res_serial.set  (mod_serial.get().upper())
    else:
        E_mod_Serial=       ttk.Entry(Mymod     ,textvariable=mod_serial     ,font=Fuente                               ,width=Ancho    ,justify='center')   .place(x=Posicion+modifi,y=Renglon*7)
        
    if res_usuario.get()    !=mod_usuario.get():
        E_mod_Usuario=      ttk.Entry(Mymod     ,textvariable=mod_usuario    ,font=Fuente       ,foreground="#FF0000"   ,width=Ancho    ,justify='center')   .place(x=Posicion+modifi,y=Renglon*9)
        res_usuario.set (mod_usuario.get().upper())
    else:
        E_mod_Usuario=      ttk.Entry(Mymod     ,textvariable=mod_usuario    ,font=Fuente                               ,width=Ancho    ,justify='center')   .place(x=Posicion+modifi,y=Renglon*9)
        
    if res_password.get()   !=mod_password.get():
        E_mod_Password=     ttk.Entry(Mymod     ,textvariable=mod_password   ,font=Fuente      ,foreground="#FF0000"   ,width=Ancho    ,justify='center')   .place(x=Posicion+modifi,y=Renglon*10)
        res_password.set(mod_password.get().upper())
    else:
        E_mod_Password=     ttk.Entry(Mymod     ,textvariable=mod_password   ,font=Fuente                               ,width=Ancho    ,justify='center')   .place(x=Posicion+modifi,y=Renglon*10)
        
    if res_server.get() !=mod_server.get():
        E_mod_Server=       ttk.Entry(Mymod     ,textvariable=mod_server     ,font=Fuente       ,foreground="#FF0000"   ,width=Ancho    ,justify='center')   .place(x=Posicion+modifi,y=Renglon*11)
        res_server.set  (mod_server.get().upper())
    else:
         E_mod_Server=       ttk.Entry(Mymod     ,textvariable=mod_server     ,font=Fuente                              ,width=Ancho    ,justify='center')   .place(x=Posicion+modifi,y=Renglon*11)
           
    
    wb.save(str(lineas[0])[:-1])
    
#****************************************************************************************************************************************************************************************************************************************************************        
def Imprimir():    
    
    global Dato,dispositivo

    m_ubi()
    Dato=dispositivo.value
    
    Mensaje=(" => "+user[9:30] + " esta buscado el dato " + Dato +" que esta es el registro numero => \n")
    import WLog 
        
    dispositivo=        ws.cell(row=num,column=4)                                                                                   # Cargamos los valores de la planilla en as variables segun su posicionamiento. Dispositivo                                 *
    nombre=             ws.cell(row=num,column=4)                                                                                   # Nombre del dispositivo en el listado                                                                                      *
    equipo=             ws.cell(row=num,column=2)                                                                                   # Tipo de Equipo                                                                                                            *
    ubicacion=          ws.cell(row=num,column=3)                                                                                   # Ubicación del Equipo                                                                                                      *
    marca=              ws.cell(row=num,column=6)                                                                                   # Marca del Equipo                                                                                                          *
    modelo=             ws.cell(row=num,column=7)                                                                                   # Modelo del Equipo                                                                                                         *
    serial=             ws.cell(row=num,column=8)                                                                                   # N° de serie                                                                                                               *
    ip=                 ws.cell(row=num,column=11)                                                                                  # N° de IP                                                                                                                  *
    usuario=            ws.cell(row=num,column=12)                                                                                  # Usuario para ingresar a la configuración                                                                                  *
    password=           ws.cell(row=num,column=13)                                                                                  # Password de ingreso a equipo                                                                                              *
    server=             ws.cell(row=num,column=14)                                                                                  # Servidor a la cual esta conectado                                                                                         *
    
    res_nombre  .set    (nombre.value)                                                                                              # Seteamos el valor de la variable a los StringVar correspondientes para mostrar en la ventana                                  *
    res_equipo  .set    (equipo.value)
    res_ubi     .set    (str(ubicacion.value))
    res_marca   .set    (marca.value)
    res_modelo  .set    (modelo.value)
    res_ip      .set    (ip.value)
    res_serial  .set    (serial.value)
    res_usuario .set    (usuario.value)
    res_password.set    (password.value)
    res_server  .set    (server.value)
    
    
    if mod_nombre.get()=="" :                                                   
        mod_nombre      .set(res_nombre.get())
    
    if mod_equipo.get()=="":
        mod_equipo      .set(res_equipo.get())
    
    if mod_ubi.get()=="":
        mod_ubi         .set(res_ubi.get())
    
    if mod_marca.get()=="":
        mod_marca       .set(res_marca.get())
    
    if mod_modelo.get()=="":
        mod_modelo      .set(res_modelo.get())
    
    if mod_ip.get()=="":
        mod_ip          .set(res_ip.get())
    
    if mod_serial.get()=="":
        mod_serial      .set(res_serial.get())
    
    if mod_usuario.get()=="":
        mod_usuario     .set(res_usuario.get())
        
    if mod_password.get()=="":
        mod_password    .set(res_password.get())
        
    if mod_server.get()=="":
        mod_server      .set(res_server.get())
#****************************************************************************************************************************************************************************************************************************************************************
def Salir():                                                                                                                        # Definimos el subcodigo para cerrar la ventana y registrar el Log                                                          *
    Mensaje=(" => Se cierra aplicación "+user[9:]+"\n")                                                                             # Definimos el mensaje a registrar en el Log                                                                                *
    import WLog                                                                                                                     # Ejecutamos el codigo Wlog                                                                                                 *
    Mymod.destroy()
    import buscador                                                                                                                 # Volvemos a la ventana del Buscador                                                                                        *
#****************************************************************************************************************************************************************************************************************************************************************    
def Destroy():
    Mymod.destroy()
#****************************************************************************************************************************************************************************************************************************************************************
# Definimos los botoes a ver en la ventana inicial                                                                                                                                                                                                              *
#****************************************************************************************************************************************************************************************************************************************************************
altura=400
boton=      tk.Button(Mymod,text="Modificar"    ,activebackground="#ABABAB"     ,background="#838383"   ,command=Modific    ,width=11   ,state='active')       .place(x=880,y=altura)               # Creo Boton "planilla" para procesar las plantillas Requeridas para informe.   *
salir=      tk.Button(Mymod,text="Salir"        ,activebackground="#BABABA"     ,background="#838383"   ,command=Salir      ,width=23   ,justify='center')       .place(x=790,y=altura+50)          # Creo un Boton para cerrar la aplicación                                       *
#****************************************************************************************************************************************************************************************************************************************************************                     
imagen=     PhotoImage(file=(user+"/Desktop/Soft/flecha.png"))
#****************************************************************************************************************************************************************************************************************************************************************
#chk_simb_disp= tk.Button(myapp ,activebackground=Fondo,background="#AAAAAA",image=imagen,height=bh).place(x=posicion+350,y=renglon*2)                                            # Boton para habilitar el cambio de Texto en Listado (Dispositivo)            *
chk_simb_ubi=   tk.Button(Mymod     ,activebackground=Fondo     ,background="#AAAAAA"   ,image=imagen   ,height=Bh      ,command=Modific)       .place(x=Posicion+350,y=Renglon*2)  # Boton para habilitar el cambio de Texto en Listado (Ubicación)    *
chk_simb_equipo=tk.Button(Mymod     ,activebackground=Fondo     ,background="#AAAAAA"   ,image=imagen   ,height=Bh      ,command=Modific)       .place(x=Posicion+350,y=Renglon*3)  # Boton para habilitar el cambio de Texto en Listado (Equipo)       *
chk_simb_nombre=tk.Button(Mymod     ,activebackground=Fondo     ,background="#AAAAAA"   ,image=imagen   ,height=Bh      ,command=Modific)       .place(x=Posicion+350,y=Renglon*4)  # Boton para habilitar el cambio de Texto en Listado (Nombre)       *
chk_simb_marca= tk.Button(Mymod     ,activebackground=Fondo     ,background="#AAAAAA"   ,image=imagen   ,height=Bh      ,command=Modific)       .place(x=Posicion+350,y=Renglon*5)  # Boton para habilitar el cambio de Texto en Listado (Marca)        *
chk_simb_modelo=tk.Button(Mymod     ,activebackground=Fondo     ,background="#AAAAAA"   ,image=imagen   ,height=Bh      ,command=Modific)       .place(x=Posicion+350,y=Renglon*6)  # Boton para habilitar el cambio de Texto en Listado (Modelo)       *
chk_simb_ip=    tk.Button(Mymod     ,activebackground=Fondo     ,background="#AAAAAA"   ,image=imagen   ,height=Bh      ,command=Modific)       .place(x=Posicion+350,y=Renglon*7)  # Boton para habilitar el cambio de Texto en Listado (Ip)           *
chk_simb_serial=tk.Button(Mymod     ,activebackground=Fondo     ,background="#AAAAAA"   ,image=imagen   ,height=Bh      ,command=Modific)       .place(x=Posicion+350,y=Renglon*8)  # Boton para habilitar el cambio de Texto en Listado (Serial)       *
chk_simb_user=  tk.Button(Mymod     ,activebackground=Fondo     ,background="#AAAAAA"   ,image=imagen   ,height=Bh      ,command=Modific)       .place(x=Posicion+350,y=Renglon*9)  # Boton para habilitar el cambio de Texto en Listado (Usuario)      *
chk_simb_pass=  tk.Button(Mymod     ,activebackground=Fondo     ,background="#AAAAAA"   ,image=imagen   ,height=Bh      ,command=Modific)       .place(x=Posicion+350,y=Renglon*10) # Boton para habilitar el cambio de Texto en Listado (Password)     *
chk_simb_server=tk.Button(Mymod     ,activebackground=Fondo     ,background="#AAAAAA"   ,image=imagen   ,height=Bh      ,command=Modific)       .place(x=Posicion+350,y=Renglon*11) # Boton para habilitar el cambio de Texto en Listado (Servidor)     *
#****************************************************************************************************************************************************************************************************************************************************************
Imprimir()
#****************************************************************************************************************************************************************************************************************************************************************
log.close()
Mymod.mainloop()
#****************************************************************************************************************************************************************************************************************************************************************


