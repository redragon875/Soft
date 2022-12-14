#****************************************************************************************************************************************************************************************************************************************************************
# Nombre: Buscador.Py                                                                                                                                                                                                                                           *       
# Descripcion: Aplicación para poder monitorear y visualizar las tareas diarias a realizar                                                                                                                                                                      *
# #**************************************************************************************************************************************************************************************************************************************************************
from ast import Str
from calendar import weekday
import cmd
from msilib.schema import PatchPackage
import os as os
import string
import sys as sys
from sys import *
from os import replace, system as system
import datetime,time
#****************************************************************************************************************************************************************************************************************************************************************
import tkinter as tk
from tkinter import *
from tkinter import Entry, Grid, Image, StringVar, Text, Variable, messagebox, ttk, scrolledtext, simpledialog, tix, font, commondialog
from tkinter import dialog
from tkinter.ttk import Notebook, Progressbar, setup_master,Sizegrip,Entry,Checkbutton,Treeview
from tkinter.tix import COLUMN, STATUS, LabelEntry,LabelFrame,Meter, ButtonBox,ComboBox
from tkinter.constants import *
from tokenize import Double
from turtle import bgcolor, color, delay, heading, title, width
from typing import Any
#****************************************************************************************************************************************************************************************************************************************************************
import openpyxl
from openpyxl import Workbook,load_workbook
from openpyxl.styles import PatternFill
#****************************************************************************************************************************************************************************************************************************************************************
#                                                                                                               Declaración de Variables                                                                                                                        *
#****************************************************************************************************************************************************************************************************************************************************************
new_user=os.environ['USERPROFILE']                                                                                                  # Identifico el "UserProfile" de la pc para poder encontrar las carpetas instaladas en el Setup.py                          *
user=new_user.replace("\\","/")                                                                                                     # Reemplazo "\" por "/" dado a que no reconocen la ruta en Python                                                           *

#path=open(new_user + '/Desktop/Codigo/Paths.txt','r')                                                                              # Busco el archivo donde se encuentran las rutas preestablecidas para encontrar los archivos. (Las cuales se pueden         *
path=open(user+'/Desktop/Soft/Path.txt')                                                                              
lineas=path.readlines()                                                                                                             # modificar en caso que asi se quiera)                                                                                      *

log=open(user  + (str(lineas[1])[:-1]),mode="a")                                                                                    # Ruta de Archivo donde se encuentran los Logs de eventos                                                                   *
#****************************************************************************************************************************************************************************************************************************************************************
dia=        datetime.datetime.today().day                                                                                           # Variable de Dia                                                                                                           *
mes=        datetime.datetime.today().month                                                                                         # Variable de Mes                                                                                                           *
week=       datetime.datetime.today().weekday
Hini=       datetime.datetime.now()                                                                                                 # Horario de inicio                                                                                                         *
Mensaje=    str()                                                                                                                   # Variable que utilizo para confeccionar los mensajes a mostrar en la aplicación                                            *
servidor=   str()
#****************************************************************************************************************************************************************************************************************************************************************
#****************************************************************************************************************************************************************************************************************************************************************
Texto=      str()                                                                                                                   # Variable para poder escribir e imprimir en pantalla o CMD                                                                 *
sdia=       int()                                                                                                                   # Variable para buscar "Día"                                                                                                *
smes=       int()                                                                                                                   # Variable para buscar "Mes"                                                                                                *
Step=       int()                                                                                                                   # Variable para marcar el paso a ejecutar en el programa                                                                    *
Fondo=      '#5B5B5B'                                                                                                               # Color Amarillo de Prosegur'#FFCC01'
#****************************************************************************************************************************************************************************************************************************************************************
#   Variables para uso del programa                                                                                                                                                                                                                             *
#****************************************************************************************************************************************************************************************************************************************************************
var=        int()                                                                                                                   # Variable de uso general para pruebas                                                                                      *
dato=       str()                                                                                                                   # Variable para poder obtener datos y poder comparar                                                                        *
num=        int(1)                                                                                                                  # Variable para poder navegar por el listado Excel                                                                          *
limite=     3000                                                                                                                    # Limite de renglones para buscar                                                                                           *
Mensaje=    str()                                                                                                                   # Variable para poder mostrar mensaje                                                                                       *    
#****************************************************************************************************************************************************************************************************************************************************************
#****************************************************************************************************************************************************************************************************************************************************************
# Defina Interfaz Grafica                                                                                                                                                                                                                                       *
#****************************************************************************************************************************************************************************************************************************************************************
#****************************************************************************************************************************************************************************************************************************************************************
myapp = Tk()                                                                                                                        # MyApp es el nombre de la planilla a visualizar                                                                            *                                                                                                             
myapp.title("Diario")                                                                                                               # Defino el titulo del programa                                                                                             *
h=700
w=1350
myapp.minsize(w,h)                                                                                                                  # Defino el Tamaño minimo e inicial de plantilla                                                                            *
#myapp.maxsize(1050,900)                                                                                                            # Defino el Tamaño al Maximar                                                                                               *
myapp.resizable(False,False)
#myapp.geometry('1000x500')                                                                                                         # Defino el Tamaño de la Grilla para poder colocar el resto de los objetos                                                  *
myapp.config(background=Fondo)                                                                                                      # Defino color de Fondo de pantalla "Myapp"                                                                                 *
#****************************************************************************************************************************************************************************************************************************************************************
#   Lineas de comando para poder centrar la ventana en la mitad de la pantalla (En teoria)
#****************************************************************************************************************************************************************************************************************************************************************
screen_width =  myapp.winfo_screenwidth()
screen_height = myapp.winfo_screenheight()
x_cordinate =   int((screen_width/2) - (w/2))
y_cordinate =   int((screen_height/2) - (h/2)-26)
myapp.geometry("{}x{}+{}+{}".format(w,h, x_cordinate, y_cordinate))
#****************************************************************************************************************************************************************************************************************************************************************
myimg=      PhotoImage(file=(user+"/Desktop/Soft/actualizar.png"))                                                                  # Variable para imagen del Boton de Busqueda. Se define la ruta en el programa. Tendria que ver de ponerlo en el Path       *
#****************************************************************************************************************************************************************************************************************************************************************
renglon=        30                                                                                                                  # Defino el valor a utilizar commo Renglones para poder las etiquetas sobre la ventana del soft                             *
#****************************************************************************************************************************************************************************************************************************************************************
mylogo=     PhotoImage(file=(user  + (str(lineas[4]))[:-1]))                                                                        # Variable para imagen del Logo del Icono a usar                                                                            *
lbl_lable=ttk.Label(myapp,image=mylogo,border=0).place(x=-10,y=-10)   
#****************************************************************************************************************************************************************************************************************************************************************
sct=scrolledtext.ScrolledText(myapp)                                                                                                # Creamos el "ScrolledText", es para poder mostrar los mensajes de progreso del soft                                        *
sct.place(x=900,y=350,width=420,height=200)                                                                                         # Configuramos la posición del "ScrolledText"                                                                               *
#****************************************************************************************************************************************************************************************************************************************************************
#****************************************************************************************************************************************************************************************************************************************************************
# Definimos Variables para respuestas                                                                                                                                                                                                                           *
#****************************************************************************************************************************************************************************************************************************************************************
res_celda=  StringVar()                                                                                                             # Variable para respuesta de Etiqueta de "Celda"                                                                            *
res_t1=     StringVar()                                                                                                             # Variable para respuesta de Etiqueta de "T1"                                                                               *
res_t2=     StringVar()                                                                                                             # Variable para respuesta de Etiqueta de "T2"                                                                               *
res_t3=     StringVar()                                                                                                             # Variable para respuesta de Etiqueta de "T3"                                                                               *
res_t4=     StringVar()                                                                                                             # Variable para respuesta de Etiqueta de "T4"                                                                               *
res_t5=     StringVar()                                                                                                             # Variable para respuesta de Etiqueta de "T5"                                                                               *
res_t6=     StringVar()                                                                                                             # Variable para respuesta de Etiqueta de "T6"                                                                               *
res_t7=     StringVar()                                                                                                             # Variable para respuesta de Etiqueta de "T7"                                                                               *
#****************************************************************************************************************************************************************************************************************************************************************
l_t1=       StringVar()                                                                                                             # Variable para manipular el Label relacionado a "T1"                                                                       *
l_t2=       StringVar()                                                                                                             # Variable para manipular el Label relacionado a "T2"                                                                       *
l_t3=       StringVar()                                                                                                             # Variable para manipular el Label relacionado a "T3"                                                                       *
l_t4=       StringVar()                                                                                                             # Variable para manipular el Label relacionado a "T4"                                                                       *
l_t5=       StringVar()                                                                                                             # Variable para manipular el Label relacionado a "T5"                                                                       *
l_t6=       StringVar()                                                                                                             # Variable para manipular el Label relacionado a "T6"                                                                       *
l_t7=       StringVar()                                                                                                             # Variable para manipular el Label relacionado a "T7"                                                                       *
#****************************************************************************************************************************************************************************************************************************************************************
cdia=       IntVar()                                                                                                                # Variable para respuesta de Etiqueta de "cdia"                                                                             *
dcontrol=   IntVar()                                                                                                                # Variable para respuesta de Etiqueta de "dcontrol"                                                                         *
controldia= IntVar()                                                                                                                # Variable para utilizar de control para poder hacer calculo                                                                *
ddia=       StringVar()                                                                                                             # Utilizamos esta variable para mostrar el String el Dia de la semana.                                                      *
#****************************************************************************************************************************************************************************************************************************************************************
if (mes%2)==0:                                                                                                                      # Validamos la cantidad de dias que tiene el mes para poder ajustar en planilla
    print("Mes Par")                                                                                                                # Tomamos el valor del dia tal cual esta para poder buscar en la planilla
    dcontrol=dia+22                                                                                                                 # Manipulamos esta variable para poder tener control del Estado de las tareas realizadas.
    smes=30
    
else:                                                                                                                               # En caso de que el mes sea Impar, imprimo eso en el CMD
    print("Mes Inpar")
    dcontrol=dia+22 
    smes=31   
#****************************************************************************************************************************************************************************************************************************************************************
# Defino dia de la semana para poder visualizar tareas rutinarias                                                                                                                                                                                               *
#****************************************************************************************************************************************************************************************************************************************************************
semana=str(Hini.strftime('%w'))
dsemana=StringVar()

if semana=="1":
    dsemana.set('Lunes')    
elif semana=="2":
    dsemana.set('Martes')
elif semana=="3":
    dsemana.set('Miercoles')
elif semana=="4":
    dsemana.set('Jueves')
elif semana=="5":
    dsemana.set('Viernes')
elif semana=="6":
    dsemana.set('Sabado')
else:
    dsemana.set("Domingo")
    
print ("Hoy es=> " + str(dsemana.get())+" "+str(dia)+" / "+str(mes))      

if dia<<15:
    cdia.set(dia)
else:
    cdia.set(dia-9)

#****************************************************************************************************************************************************************************************************************************************************************    
wb=openpyxl.load_workbook(str(lineas[0])[:-1])                                                                                      # Corresponde a la ruta en caso de tenerlo en la pc en forma local                                                          *
ws=wb["Calendario"]
 
t1=     ws.cell(row=int(semana),column=2)                                                                                           # Tomo el valor de la celda para la variables                                                                               *     
t2=     ws.cell(row=int(semana),column=3)
t3=     ws.cell(row=int(semana),column=4)
t4=     ws.cell(row=int(semana),column=5)
t5=     ws.cell(row=cdia.get(),column=2)
t6=     ws.cell(row=cdia.get(),column=3)
t7=     ws.cell(row=cdia.get(),column=4)
#****************************************************************************************************************************************************************************************************************************************************************
res_t1.set  (t1.value)                                                                                                              # Configuramos el VarString con relacion al valor de la celda                                                               *    
res_t2.set  (t2.value)
res_t3.set  (t3.value)
res_t4.set  (t4.value)
res_t5.set  (t5.value)
res_t6.set  (t6.value)
res_t7.set  (t7.value)
#****************************************************************************************************************************************************************************************************************************************************************
# Servidores a Realizar Mantenimiento
#****************************************************************************************************************************************************************************************************************************************************************
s_t1=   ws.cell(row=int(semana),column=2)
s_t2=   ws.cell(row=int(semana),column=3)
s_t3=   ws.cell(row=int(semana),column=4)
s_t4=   ws.cell(row=int(semana),column=5)
s_t5=   ws.cell(row=cdia.get(),column=2)
s_t6=   ws.cell(row=cdia.get(),column=3)
s_t7=   ws.cell(row=cdia.get(),column=4)       
#****************************************************************************************************************************************************************************************************************************************************************
# Se configura visual de Arbol para poder realizar conexiones a Servidor Directamente
#****************************************************************************************************************************************************************************************************************************************************************
tree=ttk.Treeview(myapp,height=10,show='tree')                                                                                                  # Creamos el Listado del Arbol para poder seleccionar el Servidor a Realizar conexion remota                    *
pos=int(0)

if s_t1.value !="-":
    tree.insert('', tk.END, text=s_t1.value, iid=pos, open=False)
    pos=pos+1
if s_t2.value !="-":
    tree.insert('', tk.END, text=s_t2.value, iid=pos, open=False)
    pos=pos+1
if s_t3.value !="-":
    tree.insert('', tk.END, text=s_t3.value, iid=pos, open=False)
    pos=pos+1
if s_t4.value !="-":
    tree.insert('', tk.END, text=s_t4.value, iid=pos, open=False)
    pos=pos+1
if s_t5.value!="-":
    tree.insert('', tk.END, text=s_t5.value, iid=pos, open=False)
    pos=pos+1
if s_t6.value != "-":
    tree.insert('', tk.END, text=s_t6.value, iid=pos, open=False)
    pos=pos+1
if s_t7.value != "-":
    tree.insert('', tk.END, text=s_t7.value, iid=pos, open=False)
    pos=pos+1
    
tree.place(x=900,y=90)
myapp.update()
#****************************************************************************************************************************************************************************************************************************************************************
# Defino las variables a utilizar para los CheckBox
#****************************************************************************************************************************************************************************************************************************************************************
var_t1=     IntVar()
var_t2=     IntVar()
var_t3=     IntVar()
var_t4=     IntVar()
var_t5=     IntVar()
var_t6=     IntVar()
var_t7=     IntVar()
#****************************************************************************************************************************************************************************************************************************************************************
# Definimos Los Checkbotons a Mostrar 
#****************************************************************************************************************************************************************************************************************************************************************
chk_t1=     tk.Checkbutton(myapp,activebackground="#FF0000",background="#0B0B0B",onvalue=1,offvalue=0,variable=var_t1).place(x=500,y=renglon*3)
chk_t2=     tk.Checkbutton(myapp,activebackground="#FF0000",background="#0B0B0B",onvalue=1,offvalue=0,variable=var_t2).place(x=500,y=renglon*4)
chk_t3=     tk.Checkbutton(myapp,activebackground="#FF0000",background="#0B0B0B",onvalue=1,offvalue=0,variable=var_t3).place(x=500,y=renglon*5)
chk_t4=     tk.Checkbutton(myapp,activebackground="#FF0000",background="#0B0B0B",onvalue=1,offvalue=0,variable=var_t4).place(x=500,y=renglon*6)
chk_t5=     tk.Checkbutton(myapp,activebackground="#FF0000",background="#0B0B0B",onvalue=1,offvalue=0,variable=var_t5).place(x=500,y=renglon*7)
chk_t6=     tk.Checkbutton(myapp,activebackground="#FF0000",background="#0B0B0B",onvalue=1,offvalue=0,variable=var_t6).place(x=500,y=renglon*8)
chk_t7=     tk.Checkbutton(myapp,activebackground="#FF0000",background="#0B0B0B",onvalue=1,offvalue=0,variable=var_t7).place(x=500,y=renglon*9)
#****************************************************************************************************************************************************************************************************************************************************************
# Defino Variables para control de estado de Celdas de Excel
#****************************************************************************************************************************************************************************************************************************************************************
c_t1=       ws.cell(row=dcontrol,column=2)
c_t2=       ws.cell(row=dcontrol,column=3)
c_t3=       ws.cell(row=dcontrol,column=4)
c_t4=       ws.cell(row=dcontrol,column=5)
c_t5=       ws.cell(row=dcontrol,column=7)
c_t6=       ws.cell(row=dcontrol,column=8)
c_t7=       ws.cell(row=dcontrol,column=9)
#****************************************************************************************************************************************************************************************************************************************************************
var_t1.set  (c_t1.value)
var_t2.set  (c_t2.value)
var_t3.set  (c_t3.value)
var_t4.set  (c_t4.value)
var_t5.set  (c_t5.value)
var_t6.set  (c_t6.value)
var_t7.set  (c_t7.value)
#****************************************************************************************************************************************************************************************************************************************************************
#Defino variables para poder limpiar en el Excel las celdas de control marcadas                                                                                                                                                                                 *
#****************************************************************************************************************************************************************************************************************************************************************
clr_t1=       ws.cell(row=(dcontrol+1),column=2)
clr_t2=       ws.cell(row=(dcontrol+1),column=3)
clr_t3=       ws.cell(row=(dcontrol+1),column=4)
clr_t4=       ws.cell(row=(dcontrol+1),column=5)
clr_t5=       ws.cell(row=(dcontrol+1),column=7)
clr_t6=       ws.cell(row=(dcontrol+1),column=8)
clr_t7=       ws.cell(row=(dcontrol+1),column=9)
#****************************************************************************************************************************************************************************************************************************************************************
# Definimos Labels para Agregar Tareas Manuales.                                                                                                                                                                                                                *
#****************************************************************************************************************************************************************************************************************************************************************

#****************************************************************************************************************************************************************************************************************************************************************    
cn=             IntVar()
wb=             openpyxl.load_workbook(str(lineas[0])[:-1])                                                                         # Corresponde a la ruta en caso de tenerlo en la pc en forma local                                                          *
ws=             wb["Tareas"]
dispositivo=    ws.cell(row=num,column=1)

while dispositivo.value!=None:                                                                                                      # Verificamos que el renglon contenga información para poder ver el contenido del mismo. Desde este punto podemos navegar   *
        num=num+1                                                                                                                   #por el listado de tareas Realizadas.
        dispositivo=ws.cell(row=num-1,column=1)
num=num-1
#****************************************************************************************************************************************************************************************************************************************************************
TFecha=     StringVar   (value=(ws.cell(column=1,row=num-1)).value)
TTarea=     StringVar   (value=(ws.cell(column=2,row=num-1)).value)
TEqui=      StringVar   (value=(ws.cell(column=3,row=num-1)).value)
TSector=    StringVar   (value=(ws.cell(column=4,row=num-1)).value)
TStatus=    StringVar   (value=(ws.cell(column=5,row=num-1)).value)
TAsig=      StringVar   (value=(ws.cell(column=6,row=num-1)).value)
TNotas=     StringVar   (value=(ws.cell(column=7,row=num-1)).value)
TInicio=    StringVar   (value=(ws.cell(column=8,row=num-1)).value)
TFin=       StringVar   (value=(ws.cell(column=9,row=num-1)).value)
TStat=      StringVar   ()
#****************************************************************************************************************************************************************************************************************************************************************
print       (TTarea.get())
        
wb=         openpyxl.load_workbook(str(lineas[0])[:-1])                                                                             # Corresponde a la ruta en caso de tenerlo en la pc en forma local                                                          *
ws=         wb["Tareas"]
#****************************************************************************************************************************************************************************************************************************************************************
# Definimos Sub Rutina de Status para poder cambiar el color de Labels                                                                                                                                                                                          *                
#****************************************************************************************************************************************************************************************************************************************************************    
if TStatus.get()=='0':                                                                                                              # Dependiendo del Nivel de Status es que vamos a pintar los Labels de las Tareas                                            *
    TStat.set('Abierto')                                                                                                            # En caso de que figure "0", la tarea esta "Abierto", por lo que se pinta de color Rojo                                     *
    Flbl='#FF0000'
elif TStatus.get()=='1':                                                                                                            # En caso de que figure "1", la tarea esta "Trabajando", por lo que se pinta de color Amarillo                              *
    TStat.set('Trabajando')
    Flbl='#FFFF00'
elif TStatus.get()=='2':                                                                                                            # En caso de que figure "2", la tarea esta "Listo", por lo que se pinta de color Verde                                      *
    TStat.set('Listo')
    Flbl='#00FF00'
else:
    Flbl='#0000FF'                                                                                                                  # En caso de que no figure nada, la tarea esta "Abierto", por lo que se pinta de color Azul                                 *
#****************************************************************************************************************************************************************************************************************************************************************
def Nuevo():
    global num,dispositivo
    
    dispositivo=ws.cell(row=num,column=1)
    while dispositivo.value!=None:
        num=num+1
        dispositivo=ws.cell(row=num-1,column=1)

    print('Nuevo Numero =>'+str(num))
    ws.cell     (column=1,row=num).value=       TFecha  .get()
    ws.cell     (column=2,row=num).value=       TTarea  .get()
    ws.cell     (column=3,row=num).value=       TEqui   .get()
    ws.cell     (column=4,row=num).value=       TSector .get()
    ws.cell     (column=5,row=num).value=       TStatus .get()
    ws.cell     (column=6,row=num).value=       TAsig   .get()
    ws.cell     (column=7,row=num).value=       TTarea  .get()
    ws.cell     (column=8,row=num).value=       TInicio .get()
    ws.cell     (column=9,row=num).value=       TFin    .get()
    ws.cell     (column=10,row=num).value=      TNotas  .get()
    #num=num+1
    wb.save(str(lineas[0])[:-1])

#****************************************************************************************************************************************************************************************************************************************************************    
#   Nombre          Tipo        Codificación                                                                                     Ubicacion                                                                                                                      *
#****************************************************************************************************************************************************************************************************************************************************************
lbl_TFecha=     tk.Label    (myapp,text="Fecha"     ,background=Flbl,font=("Arial",12),width=20)                            .place(x=10,y=renglon*11)
lbl_TEqui=      tk.Label    (myapp,text="Equipo"    ,background=Flbl,font=("Arial",12),width=20)                            .place(x=10,y=renglon*12)
lbl_TSector=    tk.Label    (myapp,text="Sector"    ,background=Flbl,font=("Arial",12),width=20)                            .place(x=10,y=renglon*13)
lbl_TStatus=    tk.Label    (myapp,text="Status"    ,background=Flbl,font=("Arial",12),width=20,justify='center')           .place(x=550,y=renglon*15)
lbl_TAsig=      tk.Label    (myapp,text="Asignado"  ,background=Flbl,font=("Arial",12),width=20)                            .place(x=10,y=renglon*14)
lbl_TTarea=     tk.Label    (myapp,text="Tarea"     ,background=Flbl,font=("Arial",12),width=20)                            .place(x=10,y=renglon*15)
lbl_TInicio=    tk.Label    (myapp,text="Inicio"    ,background=Flbl,font=("Arial",12),width=20)                            .place(x=550,y=renglon*11)
lbl_TFin=       tk.Label    (myapp,text="Fin"       ,background=Flbl,font=("Arial",12),width=20)                            .place(x=550,y=renglon*13)
lbl_TNotas=     tk.Label    (myapp,text="Notas"     ,background=Flbl,font=("Arial",12),width=57)                            .place(x=10,y=renglon*17)

Ent_TFecha=     tk.Entry(myapp,textvariable=TFecha  ,background=Fondo,font=("Arial",12),width=30)                           .place(x=250,y=renglon*11)
Ent_TEqui=      tk.Entry(myapp,textvariable=TEqui   ,background=Fondo,font=("Arial",12),width=30)                           .place(x=250,y=renglon*12)
Ent_TSector=    tk.Entry(myapp,textvariable=TSector ,background=Fondo,font=("Arial",12),width=30)                           .place(x=250,y=renglon*13)
Ent_TStatus=    tk.Entry(myapp,textvariable=TStatus ,background=Fondo,font=("Arial",12),width=20,justify='center')          .place(x=550,y=renglon*16)
Ent_TAsig=      tk.Entry(myapp,textvariable=TAsig   ,background=Fondo,font=("Arial",12),width=30)                           .place(x=250,y=renglon*14)
Ent_TTarea=     tk.Entry(myapp,textvariable=TTarea  ,background=Fondo,font=("Arial",12),width=57)                           .place(x=10,y=renglon*16)
Ent_TInicio=    tk.Entry(myapp,textvariable=TInicio ,background=Fondo,font=("Arial",12),width=20)                           .place(x=550,y=renglon*12)
Ent_TFin=       tk.Entry(myapp,textvariable=TFin    ,background=Fondo,font=("Arial",12),width=20)                           .place(x=550,y=renglon*14)
Ent_TNotas=     tk.Entry(myapp,textvariable=TNotas  ,background=Fondo,font=("Arial",12),width=57)                           .place(x=10,y=renglon*18)          # Imprimo el nombre de Dia en Ingles. Se puede generar el alternativo por medio de un switch*
#****************************************************************************************************************************************************************************************************************************************************************
def NSig():
    global num

    wb=openpyxl.load_workbook(str(lineas[0])[:-1])                                                                                      # Corresponde a la ruta en caso de tenerlo en la pc en forma local                                                          *
    ws=wb["Tareas"]

    num=num+1
#****************************************************************************************************************************************************************************************************************************************************************    
#   Variable    Tipo        Codificación                                                                                                                                                                                                                        *
#****************************************************************************************************************************************************************************************************************************************************************
    TFecha  =   StringVar   (value=(ws.cell(column=1,row=num)).value)       #*
    TTarea  =   StringVar   (value=(ws.cell(column=2,row=num)).value)       #*   
    TEqui   =   StringVar   (value=(ws.cell(column=3,row=num)).value)       #*
    TSector =   StringVar   (value=(ws.cell(column=4,row=num)).value)       #*
    TStatus =   StringVar   (value=(ws.cell(column=5,row=num)).value)       #*
    TAsig   =   StringVar   (value=(ws.cell(column=6,row=num)).value)       #*
    TNotas  =   StringVar   (value=(ws.cell(column=7,row=num)).value)       #*
    TInicio =   StringVar   (value=(ws.cell(column=8,row=num)).value)       #*
    TFin    =   StringVar   (value=(ws.cell(column=9,row=num)).value)       #*
#****************************************************************************************************************************************************************************************************************************************************************    
    print('TTarea => '+TTarea.get())
    print('TStatus =>'+TStatus.get())
    print('Numero =>'+ str(num))

    if TStatus.get()=='0':
        TStat.set('Abierto')
        Flbl='#FF0000'
    elif TStatus.get()=='1':
        TStat.set('Trabajando')
        Flbl='#FFFF00'
    elif TStatus.get()=='2':
        TStat.set('Listo')
        Flbl='#00FF00'
    else:
        Flbl='#0000FF' 
#****************************************************************************************************************************************************************************************************************************************************************    
#   Nombre          Tipo        Codificación                                                                                     Ubicacion                                                                                                                      *
#****************************************************************************************************************************************************************************************************************************************************************
    lbl_TFecha=     tk.Label    (myapp,text="Fecha"     ,background=Flbl,font=("Arial",12),width=20)                            .place(x=10,y=renglon*11)
    lbl_TEqui=      tk.Label    (myapp,text="Equipo"    ,background=Flbl,font=("Arial",12),width=20)                            .place(x=10,y=renglon*12)
    lbl_TSector=    tk.Label    (myapp,text="Sector"    ,background=Flbl,font=("Arial",12),width=20)                            .place(x=10,y=renglon*13)
    lbl_TStatus=    tk.Label    (myapp,text="Status"    ,background=Flbl,font=("Arial",12),width=20,justify='center')           .place(x=550,y=renglon*15)
    lbl_TAsig=      tk.Label    (myapp,text="Asignado"  ,background=Flbl,font=("Arial",12),width=20)                            .place(x=10,y=renglon*14)
    lbl_TTarea=     tk.Label    (myapp,text="Tarea"     ,background=Flbl,font=("Arial",12),width=20)                            .place(x=10,y=renglon*15)
    lbl_TInicio=    tk.Label    (myapp,text="Inicio"    ,background=Flbl,font=("Arial",12),width=20)                            .place(x=550,y=renglon*11)
    lbl_TFin=       tk.Label    (myapp,text="Fin"       ,background=Flbl,font=("Arial",12),width=20)                            .place(x=550,y=renglon*13)
    lbl_TNotas=     tk.Label    (myapp,text="Notas"     ,background=Flbl,font=("Arial",12),width=57)                            .place(x=10,y=renglon*17)

    Ent_TFecha=     tk.Entry(myapp,textvariable=TFecha  ,background=Fondo,font=("Arial",12),width=30)                           .place(x=250,y=renglon*11)
    Ent_TEqui=      tk.Entry(myapp,textvariable=TEqui   ,background=Fondo,font=("Arial",12),width=30)                           .place(x=250,y=renglon*12)
    Ent_TSector=    tk.Entry(myapp,textvariable=TSector ,background=Fondo,font=("Arial",12),width=30)                           .place(x=250,y=renglon*13)
    Ent_TStatus=    tk.Entry(myapp,textvariable=TStatus ,background=Fondo,font=("Arial",12),width=20,justify='center')          .place(x=550,y=renglon*16)
    Ent_TAsig=      tk.Entry(myapp,textvariable=TAsig   ,background=Fondo,font=("Arial",12),width=30)                           .place(x=250,y=renglon*14)
    Ent_TTarea=     tk.Entry(myapp,textvariable=TTarea  ,background=Fondo,font=("Arial",12),width=57)                           .place(x=10,y=renglon*16)
    Ent_TInicio=    tk.Entry(myapp,textvariable=TInicio ,background=Fondo,font=("Arial",12),width=20)                           .place(x=550,y=renglon*12)
    Ent_TFin=       tk.Entry(myapp,textvariable=TFin    ,background=Fondo,font=("Arial",12),width=20)                           .place(x=550,y=renglon*14)
    Ent_TNotas=     tk.Entry(myapp,textvariable=TNotas  ,background=Fondo,font=("Arial",12),width=57)                           .place(x=10,y=renglon*18)          # Imprimo el nombre de Dia en Ingles. Se puede generar el alternativo por medio de un switch*
#****************************************************************************************************************************************************************************************************************************************************************
def NAnt():
    global num
    wb=openpyxl.load_workbook(str(lineas[0])[:-1])                                                                                      # Corresponde a la ruta en caso de tenerlo en la pc en forma local                                                          *
    ws=wb["Tareas"]
    
    num=num-1
#****************************************************************************************************************************************************************************************************************************************************************    
#   Variable    Tipo        Codificación                                                                                                                                                                                                                        *
#****************************************************************************************************************************************************************************************************************************************************************
    TFecha=     StringVar   (value=(ws.cell(column=1,row=num)).value)
    TTarea=     StringVar   (value=(ws.cell(column=2,row=num)).value)
    TEqui=      StringVar   (value=(ws.cell(column=3,row=num)).value)
    TSector=    StringVar   (value=(ws.cell(column=4,row=num)).value)
    TStatus=    StringVar   (value=(ws.cell(column=5,row=num)).value)
    TAsig=      StringVar   (value=(ws.cell(column=6,row=num)).value)
    TNotas=     StringVar   (value=(ws.cell(column=7,row=num)).value)
    TInicio=    StringVar   (value=(ws.cell(column=8,row=num)).value)
    TFin=       StringVar   (value=(ws.cell(column=9,row=num)).value)
#****************************************************************************************************************************************************************************************************************************************************************        
    print('TTarea => '+TTarea.get())
    print('TStatus =>'+TStatus.get())
    print('Numero =>'+ str(num))
    
    if TStatus.get()=='0':
        TStat.set('Abierto')
        Flbl='#FF0000'
    elif TStatus.get()=='1':
        TStat.set('Trabajando')
        Flbl='#FFFF00'
    elif TStatus.get()=='2':
        TStat.set('Listo')
        Flbl='#00FF00'
    else:
        Flbl='#0000FF'            

#****************************************************************************************************************************************************************************************************************************************************************    
#   Nombre          Tipo        Codificación                                                                                     Ubicacion                                                                                                                      *
#****************************************************************************************************************************************************************************************************************************************************************
    lbl_TFecha=     tk.Label    (myapp,text="Fecha"     ,background=Flbl,font=("Arial",12),width=20)                            .place(x=10,y=renglon*11)
    lbl_TEqui=      tk.Label    (myapp,text="Equipo"    ,background=Flbl,font=("Arial",12),width=20)                            .place(x=10,y=renglon*12)
    lbl_TSector=    tk.Label    (myapp,text="Sector"    ,background=Flbl,font=("Arial",12),width=20)                            .place(x=10,y=renglon*13)
    lbl_TStatus=    tk.Label    (myapp,text="Status"    ,background=Flbl,font=("Arial",12),width=20,justify='center')           .place(x=550,y=renglon*15)
    lbl_TAsig=      tk.Label    (myapp,text="Asignado"  ,background=Flbl,font=("Arial",12),width=20)                            .place(x=10,y=renglon*14)
    lbl_TTarea=     tk.Label    (myapp,text="Tarea"     ,background=Flbl,font=("Arial",12),width=20)                            .place(x=10,y=renglon*15)
    lbl_TInicio=    tk.Label    (myapp,text="Inicio"    ,background=Flbl,font=("Arial",12),width=20)                            .place(x=550,y=renglon*11)
    lbl_TFin=       tk.Label    (myapp,text="Fin"       ,background=Flbl,font=("Arial",12),width=20)                            .place(x=550,y=renglon*13)
    lbl_TNotas=     tk.Label    (myapp,text="Notas"     ,background=Flbl,font=("Arial",12),width=57)                            .place(x=10,y=renglon*17)

    Ent_TFecha=     tk.Entry(myapp,textvariable=TFecha  ,background=Fondo,font=("Arial",12),width=30)                           .place(x=250,y=renglon*11)
    Ent_TEqui=      tk.Entry(myapp,textvariable=TEqui   ,background=Fondo,font=("Arial",12),width=30)                           .place(x=250,y=renglon*12)
    Ent_TSector=    tk.Entry(myapp,textvariable=TSector ,background=Fondo,font=("Arial",12),width=30)                           .place(x=250,y=renglon*13)
    Ent_TStatus=    tk.Entry(myapp,textvariable=TStatus ,background=Fondo,font=("Arial",12),width=20,justify='center')          .place(x=550,y=renglon*16)
    Ent_TAsig=      tk.Entry(myapp,textvariable=TAsig   ,background=Fondo,font=("Arial",12),width=30)                           .place(x=250,y=renglon*14)
    Ent_TTarea=     tk.Entry(myapp,textvariable=TTarea  ,background=Fondo,font=("Arial",12),width=57)                           .place(x=10,y=renglon*16)
    Ent_TInicio=    tk.Entry(myapp,textvariable=TInicio ,background=Fondo,font=("Arial",12),width=20)                           .place(x=550,y=renglon*12)
    Ent_TFin=       tk.Entry(myapp,textvariable=TFin    ,background=Fondo,font=("Arial",12),width=20)                           .place(x=550,y=renglon*14)
    Ent_TNotas=     tk.Entry(myapp,textvariable=TNotas  ,background=Fondo,font=("Arial",12),width=57)                           .place(x=10,y=renglon*18)          # Imprimo el nombre de Dia en Ingles. Se puede generar el alternativo por medio de un switch*
#****************************************************************************************************************************************************************************************************************************************************************
def Nguardar():
    global num,dispositivo
    global TFecha,TTarea
    
    dispositivo=ws.cell(row=num,column=1)
    while dispositivo.value!=None:
        num=num+1
        dispositivo=ws.cell(row=num-1,column=1)
    
    VFecha=str(Ent_TFecha)
    VTarea=Str(Ent_TTarea)
    VEqui=Str()
    VSector=Str()
    VStatus=Str()
    VAsig=Str()
    VTarea=Str()
    VInicio=Str()
    VFin=Str()
    Vnotas=Str()

    #TFecha      .set    (str(VFecha.value))
    #TTarea      .set    (value=Ent_TTarea)
    #TEqui       .set    (value=Ent_TEqui)
    #TSector     .set    (value=Ent_TSector)
    #TStatus     .set    (value=Ent_TStatus)
    #TAsig       .set    (value=Ent_TAsig)
    #TTarea      .set    (value=Ent_TTarea)
    #TInicio     .set    (value=Ent_TInicio)
    #TFin        .set    (value=Ent_TFin)
    #TNotas      .set    (value=Ent_TNotas)

    print('Nuevo Numero =>'+str(num)+"valor de entrada "+str(VFecha))
    
    ws.cell     (column=1,row=num).value=       VFecha
    #ws.cell     (column=2,row=num).value=       TTarea  .get()
    #ws.cell     (column=3,row=num).value=       TEqui   .get()
    #ws.cell     (column=4,row=num).value=       TSector .get()
    #ws.cell     (column=5,row=num).value=       TStatus .get()
    #ws.cell     (column=6,row=num).value=       TAsig   .get()
    #ws.cell     (column=7,row=num).value=       TTarea  .get()
    #ws.cell     (column=8,row=num).value=       TInicio .get()
    #ws.cell     (column=9,row=num).value=       TFin    .get()
    #ws.cell     (column=10,row=num).value=      TNotas  .get()
    
    print('Nuevo Numero =>'+str(num)+" Fecha "+str(VFecha))
    
    num=num+1
    messagebox.showinfo(title="Guardado",message="Se guardo información")
    
    wb.save(str(lineas[0])[:-1])
    myapp.update()
    
#****************************************************************************************************************************************************************************************************************************************************************
#print("\nValores Iniciales:\nValor de C_T1 => "+ str(c_t1.value)+"\nValor de C_T2 => "+ str(c_t2.value)+"\nValor de C_T3 => "+ str(c_t3.value)+"\nValor de C_T4 => "+ str(c_t4.value)+"\nValor de C_T5 => "+ str(c_t5.value)+"\nValor de C_T6 => "+ str(c_t6.value))
#****************************************************************************************************************************************************************************************************************************************************************
# Funciones del Programa a utilizar                                                                                                                                                                                                                             *
#****************************************************************************************************************************************************************************************************************************************************************   
def Ayuda():
    system('start %userprofile%\Desktop\Soft\Help.txt')                                                                             # Abro por Sistema el Txt de ayuda del programa                                                                             *
    return
#****************************************************************************************************************************************************************************************************************************************************************                        
def Siguiente():
    
    global var_t1,var_t2,var_t3,var_t4,var_t5,var_t6
    global t1,t2,t3,t4,t5,t6
    global s_t1,s_t2,s_t3,s_t4,s_t5,s_t6
    global num,celda                                                                                                                # Tomo las Variables Globales y le permito el uso dentro del subprograma                                                    *
    pos=1
    wb=openpyxl.load_workbook(str(lineas[0])[:-1])                                                                                      # Corresponde a la ruta en caso de tenerlo en la pc en forma local                                                          *
    ws=wb["Calendario"]
    
    control=cdia.get()
    control=control+1
    
    if smes==30:
        cdia.set(control)
        print ("Smes => "+str(control))
    else:
        cdia.set(control-1)
        print("Mes 31 dias"+str(control))

    
    if control<smes:
        cdia.set(control)
        print ("control "+str(control))
    else:
        control=1
        cdia.set(control)
        print (control) 

#****************************************************************************************************************************************************************************************************************************************************************
#   Nombre   /  Valor                                                                                                                                                                                                                                           *
#****************************************************************************************************************************************************************************************************************************************************************   
    
    t1=     ws.cell(row=int(semana),column=2)                                                                                           # Tomo el valor de la celda para la variables                                                                               *     
    t2=     ws.cell(row=int(semana),column=3)
    t3=     ws.cell(row=int(semana),column=4)
    t4=     ws.cell(row=int(semana),column=5)
    t5=     ws.cell(row=cdia.get(),column=2)
    t6=     ws.cell(row=cdia.get(),column=3)
    t7=     ws.cell(row=cdia.get(),column=4)
    
    res_t1.set  (t1.value)
    res_t2.set  (t2.value)
    res_t3.set  (t3.value)
    res_t4.set  (t4.value)
    res_t5.set  (t5.value)
    res_t6.set  (t6.value)
    res_t7.set  (t7.value)
      
    s_t1=   ws.cell(row=int(semana),column=2)
    s_t2=   ws.cell(row=int(semana),column=3)
    s_t3=   ws.cell(row=int(semana),column=4)
    s_t4=   ws.cell(row=int(semana),column=5)
    s_t5=   ws.cell(row=cdia.get(),column=2)
    s_t6=   ws.cell(row=cdia.get(),column=3)
    s_t7=   ws.cell(row=cdia.get(),column=4)

    for i in tree.get_children():
         tree.delete(i)
    myapp.update()
    pos=int(0)
    
    if s_t1.value !="-":
        tree.insert('', tk.END, text=s_t1.value, iid=pos, open=False,)
        pos=pos+1     
    if s_t2.value !="-":
        tree.insert('', tk.END, text=s_t2.value, iid=pos, open=False)
        pos=pos+1
    if s_t3.value !="-":
        tree.insert('', tk.END, text=s_t3.value, iid=pos, open=False)
        pos=pos+1
    if s_t4.value !="-":
        tree.insert('', tk.END, text=s_t4.value, iid=pos, open=False)
        pos=pos+1
    if s_t5.value!="-":
        tree.insert('', tk.END, text=s_t5.value, iid=pos, open=False)
        pos=pos+1
    if s_t6.value != "-":
        tree.insert('', tk.END, text=s_t6.value, iid=pos, open=False)
        pos=pos+1
    if s_t7.value != "-":
        tree.insert('', tk.END, text=s_t7.value, iid=pos, open=False)
        pos=pos+1
    
    myapp.update() 
    return 
#****************************************************************************************************************************************************************************************************************************************************************    
def Anterior():
       
    global var_t1,var_t2,var_t3,var_t4,var_t5,var_t6
    global t1,t2,t3,t4,t5,t6
    global num,celda                                                                                                                # Tomo las Variables Globales y le permito el uso dentro del subprograma                                                    *
    pos=int(0)
    
    wb=openpyxl.load_workbook(str(lineas[0])[:-1])                                                                                      # Corresponde a la ruta en caso de tenerlo en la pc en forma local                                                          *
    ws=wb["Calendario"]
    num=1
    control=cdia.get()
    cdia.set(control-num)
    if (cdia.get()<1) & (smes==30):
        cdia.set(control+smes)
    elif (cdia.get()<1) & (smes==31):
        cdia.set(control+29)  
#****************************************************************************************************************************************************************************************************************************************************************
#   Nombre    / Valor                                                                                                                                                                                                                                           *
#****************************************************************************************************************************************************************************************************************************************************************
    t1=     ws.cell(row=int(semana),column=2)                                                                                           # Tomo el valor de la celda para la variables                                                                               *     
    t2=     ws.cell(row=int(semana),column=3)
    t3=     ws.cell(row=int(semana),column=4)
    t4=     ws.cell(row=int(semana),column=5)
    t5=     ws.cell(row=cdia.get(),column=2)
    t6=     ws.cell(row=cdia.get(),column=3)
    t7=     ws.cell(row=cdia.get(),column=4)
    
    res_t1.set  (t1.value)
    res_t2.set  (t2.value)
    res_t3.set  (t3.value)
    res_t4.set  (t4.value)
    res_t5.set  (t5.value)
    res_t6.set  (t6.value)
    res_t7.set  (t7.value)
      
    s_t1=   ws.cell(row=int(semana),column=2)
    s_t2=   ws.cell(row=int(semana),column=3)
    s_t3=   ws.cell(row=int(semana),column=4)
    s_t4=   ws.cell(row=int(semana),column=5)
    s_t5=   ws.cell(row=cdia.get(),column=2)
    s_t6=   ws.cell(row=cdia.get(),column=3)
    s_t7=   ws.cell(row=cdia.get(),column=4)
    
    for i in tree.get_children():
         tree.delete(i)
    myapp.update()
    pos=int(0)
    
    if s_t1.value !="-":
        tree.insert('', tk.END, text=s_t1.value, iid=pos, open=False,)
        pos=pos+1     
    if s_t2.value !="-":
        tree.insert('', tk.END, text=s_t2.value, iid=pos, open=False)
        pos=pos+1
    if s_t3.value !="-":
        tree.insert('', tk.END, text=s_t3.value, iid=pos, open=False)
        pos=pos+1
    if s_t4.value !="-":
        tree.insert('', tk.END, text=s_t4.value, iid=pos, open=False)
        pos=pos+1
    if s_t5.value!="-":
        tree.insert('', tk.END, text=s_t5.value, iid=pos, open=False)
        pos=pos+1
    if s_t6.value != "-":
        tree.insert('', tk.END, text=s_t6.value, iid=pos, open=False)
        pos=pos+1
    if s_t7.value != "-":
        tree.insert('', tk.END, text=s_t7.value, iid=pos, open=False)
        pos=pos+1   
           
    myapp.update() 
#****************************************************************************************************************************************************************************************************************************************************************
def Actualizar():
    
    global var_t1,var_t2,var_t3,var_t4,var_t5,var_t6,var_t7
    global t1,t2,t3,t4,t5,t6
    global num,fill_cell
   
    sct.delete('1.0', END)                                                                                                          # Limpieamos el Scrolltext, para poder visualizar las tareas correctamente armadas
    sct.insert(INSERT,"Se Realiza Mantenimiento de: ")                                                                              # Incertamos la primer linea del Scroll text
                   
    wb=openpyxl.load_workbook(str(lineas[0])[:-1])                                                                                  # Abrimos el Excel de Base para
    ws=wb["Calendario"]                                                 
    
#****************************************************************************************************************************************************************************************************************************************************************
#   Validacion Visual de Celdas en Excel
#****************************************************************************************************************************************************************************************************************************************************************
    if var_t1.get()==1:                                                                                                                     # Validacion de tarea de renglon T1, en caso de que sea Verdadera, se pinta la tarea de Verde en Excel               *        
        t1=ws.cell(row=int(semana),column=2)
        ws[t1.coordinate].fill=PatternFill(patternType='solid',fgColor='FF0000')
        lbl_t1=tk.Label(myapp,textvariable=res_t1,background="#00FF00",font=("Arial",12),width=50).place(x=10,y=renglon*3)
        sct.insert(INSERT,"\n \t* "+t1.value)
        c_t1.value="1"
        ws[c_t1.coordinate]="1"
    else:                                                                                                                                   # En caso de que sea Falso, se pinta la tarea de Rojo                                                             *
        t1=ws.cell(row=int(semana),column=2)
        ws[t1.coordinate].fill=PatternFill(patternType='solid', fgColor='FF0000')
        lbl_t1=tk.Label(myapp,textvariable=res_t1,background='#FF0000',font=("Arial",12),width=50).place(x=10,y=renglon*3)
        c_t1.value="0" 
        ws[c_t1.coordinate]="0"
        
    if var_t2.get()==1:                                                                                                                     # Validacion del resto de las tareas de renglones T2 a T8                                                           * 
        t2=ws.cell(row=int(semana),column=3)
        ws[t2.coordinate].fill=PatternFill(patternType='solid', fgColor='00FF00')
        lbl_t2=tk.Label(myapp,textvariable=res_t2,background="#00FF00",font=("Arial",12),width=50).place(x=10,y=renglon*4)
        sct.insert(INSERT,"\n \t* "+t2.value)
        c_t2.value="1"
        ws[c_t2.coordinate]="1"
    else:                                                                                                                                   
        t2=ws.cell(row=int(semana),column=3)
        ws[t2.coordinate].fill=PatternFill(patternType='solid', fgColor='FF0000')
        lbl_t2=tk.Label(myapp,textvariable=res_t2,background="#FF0000",font=("Arial",12),width=50).place(x=10,y=renglon*4)
        c_t2.value="0"
        ws[c_t2.coordinate]="1"
         
    if var_t3.get()==1:                                                                                                                 
        t3=ws.cell(row=int(semana),column=4)
        ws[t3.coordinate].fill=PatternFill(patternType='solid', fgColor='00FF00')
        lbl_t3=tk.Label(myapp,textvariable=res_t3,background="#00FF00",font=("Arial",12),width=50).place(x=10,y=renglon*5)
        sct.insert(INSERT,"\n \t* "+t3.value)
        c_t3.value="1"
        ws[c_t3.coordinate]="1"
    else:                                                                                                                               
        t3=ws.cell(row=int(semana),column=4)
        ws[t3.coordinate].fill=PatternFill(patternType='solid', fgColor='FF0000')
        lbl_t3=tk.Label(myapp,textvariable=res_t3,background="#FF0000",font=("Arial",12),width=50).place(x=10,y=renglon*5)
        c_t3.value="0"
        ws[c_t3.coordinate]="0"
        
    if var_t4.get()==1:                                                                                                                 
        t4=ws.cell(row=int(semana),column=5)
        ws[t4.coordinate].fill=PatternFill(patternType='solid', fgColor='00FF00')
        lbl_t4=tk.Label(myapp,textvariable=res_t4,background="#00FF00",font=("Arial",12),width=50).place(x=10,y=renglon*6)
        sct.insert(INSERT,"\n \t* "+t4.value)
        c_t4.value="1"
        ws[c_t4.coordinate]="1"
    else:                                                                                                                               
        t4=ws.cell(row=int(semana),column=5)
        ws[t4.coordinate].fill=PatternFill(patternType='solid', fgColor='FF0000')
        lbl_t4=tk.Label(myapp,textvariable=res_t4,background="#FF0000",font=("Arial",12),width=50).place(x=10,y=renglon*6)
        c_t4.value="0"
        ws[c_t4.coordinate]="0"
    
    if var_t5.get()==1:                                                                                                                 
        t5 =ws.cell(row=cdia.get(),column=2)
        ws[t5.coordinate].fill=PatternFill(patternType='solid', fgColor='00FF00')
        lbl_t5=tk.Label(myapp,textvariable=res_t5,background="#00FF00",font=("Arial",12),width=50).place(x=10,y=renglon*7)
        sct.insert(INSERT,"\n \t* "+t5.value)
        c_t5.value="1"
        ws[c_t5.coordinate]="1"
    else:                                                                                                                               
        t5=ws.cell(row=cdia.get(),column=2)
        ws[t5.coordinate].fill=PatternFill(patternType='solid', fgColor='FF0000')
        lbl_t5=tk.Label(myapp,textvariable=res_t5,background="#FF0000",font=("Arial",12),width=50).place(x=10,y=renglon*7)
        c_t5.value="0"
        ws[c_t5.coordinate]="0"

    if var_t6.get()==1:                                                                                                                 
        t6=ws.cell(row=cdia.get(),column=3)
        ws[t6.coordinate].fill=PatternFill(patternType='solid', fgColor='00FF00')
        lbl_t6=tk.Label(myapp,textvariable=res_t6,background="#00FF00",font=("Arial",12),width=50).place(x=10,y=renglon*8)
        sct.insert(INSERT,"\n \t* "+t6.value)
        c_t6.value="1"
        ws[c_t6.coordinate]="1"
    else:                                                                                                                               
        t6=ws.cell(row=cdia.get(),column=3)
        ws[t6.coordinate].fill=PatternFill(patternType='solid', fgColor='FF0000')
        lbl_t6=tk.Label(myapp,textvariable=res_t6,background="#FF0000",font=("Arial",12),width=50).place(x=10,y=renglon*8)
        c_t6.value="0"
        ws[c_t6.coordinate]="0"
    
    if var_t7.get()==1:                                                                                                                 
        t7=ws.cell(row=cdia.get(),column=4)
        ws[t7.coordinate].fill=PatternFill(patternType='solid', fgColor='00FF00')
        lbl_t7=tk.Label(myapp,textvariable=res_t7,background="#00FF00",font=("Arial",12),width=50).place(x=10,y=renglon*9)
        sct.insert(INSERT,"\n \t* "+t7.value)
        c_t7.value="1"
        ws[c_t7.coordinate]="1"
    else:                                                                                                                               
        t7=ws.cell(row=cdia.get(),column=4)
        ws[t7.coordinate].fill=PatternFill(patternType='solid', fgColor='FF0000')
        lbl_t7=tk.Label(myapp,textvariable=res_t7,background="#FF0000",font=("Arial",12),width=50).place(x=10,y=renglon*9)
        c_t7.value="0"
        ws[c_t7.coordinate]="0"
#****************************************************************************************************************************************************************************************************************************************************************
    res_t1.set      (t1.value)
    res_t2.set      (t2.value)
    res_t3.set      (t3.value)
    res_t4.set      (t4.value)
    res_t5.set      (t5.value)
    res_t6.set      (t6.value)  
    res_t7.set      (t7.value) 
#****************************************************************************************************************************************************************************************************************************************************************
    l_t1.set        (ws.cell(row=(int(semana)),column=2))
    l_t2.set        (ws.cell(row=(int(semana)),column=3))
    l_t3.set        (ws.cell(row=(int(semana)),column=4))
    l_t4.set        (ws.cell(row=(int(semana)),column=5))
    l_t5.set        (ws.cell(row=(cdia.get()),column=2))
    l_t6.set        (ws.cell(row=(cdia.get()),column=3))
    l_t7.set        (ws.cell(row=(cdia.get()),column=4))
#****************************************************************************************************************************************************************************************************************************************************************
#    
    ws.cell         (row=(int(semana)+1),column=2).fill=      PatternFill(patternType='solid', fgColor='FFFFFF')
    ws.cell         (row=(int(semana)+1),column=3).fill=      PatternFill(patternType='solid', fgColor='FFFFFF')
    ws.cell         (row=(int(semana)+1),column=4).fill=      PatternFill(patternType='solid', fgColor='FFFFFF')
    ws.cell         (row=(int(semana)+1),column=5).fill=      PatternFill(patternType='solid', fgColor='FFFFFF')
    ws.cell         (row=(cdia.get()+1),column=2).fill=       PatternFill(patternType='solid', fgColor='FFFFFF')
    ws.cell         (row=(cdia.get()+1),column=3).fill=       PatternFill(patternType='solid', fgColor='FFFFFF')
    ws.cell         (row=(cdia.get()+1),column=4).fill=       PatternFill(patternType='solid', fgColor='FFFFFF')
    ws.cell         (row=(cdia.get()+1),column=5).fill=       PatternFill(patternType='solid', fgColor='FFFFFF')
#****************************************************************************************************************************************************************************************************************************************************************
    myapp.update()  
    wb.save(str(lineas[0])[:-1])
#****************************************************************************************************************************************************************************************************************************************************************       
    #print('\nValor Actualizados\ndia '+str(cdia.get())+" mes "+str(mes) + "\nt1= " + t1.value +" => "+str(var_t1.get())+"\nt2= " + t2.value+" => "+str(var_t2.get())+"\nt3= "+ str(t3.value)+" => "+str(var_t3.get())+"\nt4= "+str(t4.value)+" => "+str(var_t4.get())
    #        +"\nt5= "+str(t5.value)+" => "+str(var_t5.get())+"\nt6= "+str(t6.value)+" => "+str(var_t6.get()))
    #print("\nValor de C_T1 => "+ str(c_t1.value)+"\nValor de C_T2 => "+ str(c_t2.value)+"\nValor de C_T3 => "+ str(c_t3.value)+"\nValor de C_T4 => "+ str(c_t4.value)+"\nValor de C_T5 => "+ str(c_t5.value)+"\nValor de C_T6 => "+ str(c_t6.value))
#****************************************************************************************************************************************************************************************************************************************************************
# Definimos Los Labels a Mostrar 
#****************************************************************************************************************************************************************************************************************************************************************
#****************************************************************************************************************************************************************************************************************************************************************    
lbl_res_celda=  tk.Label    (myapp,text="Dia",background=Fondo,font=("Arial",12))                       .place(x=10,y=renglon)
lbl_dia=        tk.Label    (myapp,textvariable=cdia,background=Fondo,font=("Arial",12),width=25)       .place(x=50,y=renglon)
lbl_t1=         tk.Label    (myapp,textvariable=res_t1,background=Fondo,font=("Arial",12),width=50)     .place(x=10,y=renglon*3)
lbl_t2=         tk.Label    (myapp,textvariable=res_t2,background=Fondo,font=("Arial",12),width=50)     .place(x=10,y=renglon*4)
lbl_t3=         tk.Label    (myapp,textvariable=res_t3,background=Fondo,font=("Arial",12),width=50)     .place(x=10,y=renglon*5)
lbl_t4=         tk.Label    (myapp,textvariable=res_t4,background=Fondo,font=("Arial",12),width=50)     .place(x=10,y=renglon*6)
lbl_t5=         tk.Label    (myapp,textvariable=res_t5,background=Fondo,font=("Arial",12),width=50)     .place(x=10,y=renglon*7)
lbl_t6=         tk.Label    (myapp,textvariable=res_t6,background=Fondo,font=("Arial",12),width=50)     .place(x=10,y=renglon*8)
lbl_t7=         tk.Label    (myapp,textvariable=res_t7,background=Fondo,font=("Arial",12),width=50)     .place(x=10,y=renglon*9)
#****************************************************************************************************************************************************************************************************************************************************************
myapp.update()
#****************************************************************************************************************************************************************************************************************************************************************

#****************************************************************************************************************************************************************************************************************************************************************        
def Imprimir():    
    
    global dato,num,dispositivo
    
    wb=openpyxl.load_workbook(str(lineas[0])[:-1])
    ws=wb["Calendario"]
    
    sct.insert(INSERT,"\n Se imprime informe ")
    
    Mensaje=str(Hini)+" => "+user[9:30]+ " " + sct.get("1.0",END) + "\n"+(100*"*")+"\n"
    log=open(user  + (str(lineas[1])[:-1]),mode="a")
    log.write(Mensaje)
    log.close()
#****************************************************************************************************************************************************************************************************************************************************************     
def Conectar():
    
    global tree,sct
    
    current_item = tree.focus()    
    data = tree.item(current_item,option='text') 
    server=data
    log=open(user  + (str(lineas[1])[:-1]),mode="a")
    log.write(str(Hini)+" => "+ user[9:]+"Se conecta al servidor " + str(server) + "\n")
    log.close()
    servidor='cmd /k "mstsc -v ' + str(server) + ':4489'
    sct.insert(INSERT,"\n\t* "+ server)
    time.sleep(1)
    #myapp.destroy()
    os.system(servidor)                                                                                                                 #\n Se realiza mantenimiento remoto en:                                                                                 *
    return                                                                                                                                                                                                                                                     #*
#****************************************************************************************************************************************************************************************************************************************************************    
def Salir():
    
    log=open(user  + (str(lineas[1])[:-1]),mode="a")
    log.write(str(Hini)+" => Se cierra aplicación "+user[9:]+"\n")
    log.close()
    Destroy()
#****************************************************************************************************************************************************************************************************************************************************************    
def Destroy():
    myapp.destroy()

#****************************************************************************************************************************************************************************************************************************************************************
# Definimos los botoes a ver en la ventana inicial                                                                                                                                                                                                              *
#****************************************************************************************************************************************************************************************************************************************************************
boton=          tk.Button(myapp,text="Buscar",activebackground="#ABCDEF",background="#838B8B",command=Actualizar,width=155,image=myimg)         .place(x=1150,y=60)                  # Creo Boton "planilla" para procesar las plantillas Requeridas para informe*
salir=          tk.Button(myapp,text="Salir",activebackground="#BABABA",command=Salir,justify='center',width=7)                                 .place(x=1250,y=610)                 # Creo un Boton para cerrar la aplicación                                   *
bhelp=          tk.Button(myapp,text="Ayuda",background="#838383",command=Ayuda,width=5)                                                        .place(x=1200,y=610)                 # Creo el Boton de "Ayuda" para mostrar el Txt correspondiente              *
bsiguiente=     tk.Button(myapp,text="Siguiente",activebackground="#ABABAB",background="#838383",command=Siguiente,width=10,state='active')     .place(x=1230,y=150)
banterios=      tk.Button(myapp,text="Previo",activebackground="#ABABAB",background="#838383",command=Anterior,width=10,state='active')         .place(x=1150,y=150)
bconectar=      tk.Button(myapp,text="Conectar",activebackground="#ABABAB",background="#838383",command=Conectar,width=22,state='active')       .place(x=1150,y=180)
bimprimir=      tk.Button(myapp,text="Imprimir",activebackground="#ABABAB",background="#838383",command=Imprimir,width=22,state='active')       .place(x=1150,y=250)
#****************************************************************************************************************************************************************************************************************************************************************                     
bnuevo=         tk.Button(myapp,text="Nueva Linea",activebackground="#ABABAB",background="#838383",command=Nuevo,width=15,state='active')       .place(x=10,y=640)
bn_siguente=    tk.Button(myapp,text="Siguiente Linea",activebackground="#ABABAB",background="#838383",command=NSig,width=15,state='active')    .place(x=240,y=640)
bn_anterior=    tk.Button(myapp,text="Linea Anterior",activebackground="#ABABAB",background="#838383",command=NAnt,width=15,state='active')     .place(x=360,y=640)
bn_guardar=     tk.Button(myapp,text="Guardar",activebackground="#ABABAB",background="#838383",command=Nguardar,width=6,state='active')         .place(x=475,y=640)
#****************************************************************************************************************************************************************************************************************************************************************                     
#****************************************************************************************************************************************************************************************************************************************************************
log.close()
myapp.mainloop()
#****************************************************************************************************************************************************************************************************************************************************************